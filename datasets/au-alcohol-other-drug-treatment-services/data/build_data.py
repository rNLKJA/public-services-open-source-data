"""
Converts the two raw AIHW AODTS NMDS "state and territory" workbooks into tidy CSVs,
isolating only the South Australia sheets from each (each workbook covers all
states/territories; SA is one block of sheets among them).

Source workbooks (raw/, verbatim as downloaded):
  - aihw-hse-258-2425-SCR-client-state-territory-numbers-tables.xlsx
    "SCR.Clients (state and territories numbers)" - client counts, FY2013-14 to 2024-25
  - aihw-hse-258-2425-ST-state-territory-episode-tables-25062026.xlsx
    "ST.State and territories (episodes)" - closed treatment episode counts

Each source sheet is one of two shapes:
  Type A ("year-as-columns"): financial years run across the columns, 1-2 row-label
    columns identify the breakdown category (e.g. sex, age group, principal drug).
  Type B ("category-as-columns"): a second categorical variable (e.g. referral source,
    duration band) runs across the columns instead of years. Where the sheet stacks
    a repeating block per financial year, the year appears as the first row label.

No value is recalculated - every number here is exactly what AIHW published, reshaped
from wide to long with a `source_table` column identifying which numbered AIHW table
(and therefore which exact row/column definition) each row came from.
"""
import csv
import re
from pathlib import Path

import openpyxl

RAW = Path(__file__).resolve().parent.parent / "raw"
OUT = Path(__file__).resolve().parent

YEAR_RE = re.compile(r"^\d{4}–\d{2}$")
TITLE_YEAR_RE = re.compile(r"(\d{4}–\d{2})\s*$")
TITLE_DESC_RE = re.compile(r"^Table \S+ SA\.\d+:\s*(.+?),\s*South Australia,")


def clean(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def extract_table(ws):
    rows = list(ws.iter_rows(values_only=True))
    title = rows[0][0] or ""
    desc_match = TITLE_DESC_RE.match(title)
    description = desc_match.group(1) if desc_match else title
    header_row = rows[2]

    # Locate the contiguous block of non-empty header cells (from col C / idx 2 onward)
    populated = [i for i, v in enumerate(header_row) if v not in (None, "") and i >= 2]
    if not populated:
        return []
    value_start = min(populated)
    value_end = max(populated)
    value_cols = list(range(value_start, value_end + 1))

    year_value_cols = [c for c in value_cols if YEAR_RE.match(str(header_row[c] or ""))]
    is_type_a = len(year_value_cols) >= max(2, len(value_cols) // 2)

    records = []

    if is_type_a:
        label_cols = list(range(2, value_start))
        last_seen = {c: None for c in label_cols}
        for row in rows[3:]:
            if row[0] and str(row[0]).strip().lower().startswith("note"):
                break
            if all(v is None for v in row):
                break
            for c in label_cols:
                if row[c] not in (None, ""):
                    last_seen[c] = row[c]
            labels = [last_seen[c] for c in label_cols]
            for c in year_value_cols:
                val = row[c]
                if val is None:
                    continue
                year = header_row[c]
                rec = {
                    "source_table": title.split(":")[0].replace("Table ", ""),
                    "breakdown": description,
                    "financial_year": year,
                    "category_1": clean(labels[0]) if len(labels) > 0 else "",
                    "category_2": clean(labels[1]) if len(labels) > 1 else "",
                    "column_category": "",
                    "count": clean(val),
                }
                records.append(rec)
    else:
        label_cols = list(range(2, value_start))
        last_seen = {c: None for c in label_cols}
        year_col = label_cols[0] if label_cols else None
        # decide whether the first label column actually carries a year value anywhere
        year_col_is_year = False
        if year_col is not None:
            for row in rows[3:]:
                if row[year_col] not in (None, "") and YEAR_RE.match(str(row[year_col])):
                    year_col_is_year = True
                    break
        fixed_year = None
        if not year_col_is_year:
            m = TITLE_YEAR_RE.search(title)
            fixed_year = m.group(1) if m else ""

        cat_label_cols = label_cols[1:] if year_col_is_year else label_cols

        for row in rows[3:]:
            if row[0] and str(row[0]).strip().lower().startswith("note"):
                break
            if all(v is None for v in row):
                break
            for c in label_cols:
                if row[c] not in (None, ""):
                    last_seen[c] = row[c]
            year = last_seen[year_col] if year_col_is_year else fixed_year
            labels = [last_seen[c] for c in cat_label_cols]
            for c in value_cols:
                val = row[c]
                if val is None:
                    continue
                col_category = header_row[c]
                rec = {
                    "source_table": title.split(":")[0].replace("Table ", ""),
                    "breakdown": description,
                    "financial_year": year,
                    "category_1": clean(labels[0]) if len(labels) > 0 else "",
                    "category_2": clean(labels[1]) if len(labels) > 1 else "",
                    "column_category": clean(col_category),
                    "count": clean(val),
                }
                records.append(rec)

    return records


def process(workbook_path, sheet_prefix):
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    all_records = []
    for name in wb.sheetnames:
        if not name.startswith(f"Table {sheet_prefix} SA."):
            continue
        all_records.extend(extract_table(wb[name]))
    return all_records


def write_csv(records, path, extra_field=None):
    fields = ["source_table", "breakdown", "financial_year", "category_1", "category_2"]
    if extra_field:
        fields.append(extra_field)
    fields.append("count")
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for rec in records:
            writer.writerow({k: rec[k] for k in fields})


def main():
    clients = process(RAW / "aihw-hse-258-2425-SCR-client-state-territory-numbers-tables.xlsx", "SCR")
    episodes = process(RAW / "aihw-hse-258-2425-ST-state-territory-episode-tables-25062026.xlsx", "ST")

    write_csv(clients, OUT / "sa-aod-clients-by-characteristic.csv", extra_field="column_category")
    write_csv(episodes, OUT / "sa-aod-closed-treatment-episodes.csv", extra_field="column_category")

    print(f"clients: {len(clients)} rows -> sa-aod-clients-by-characteristic.csv")
    print(f"episodes: {len(episodes)} rows -> sa-aod-closed-treatment-episodes.csv")


if __name__ == "__main__":
    main()
