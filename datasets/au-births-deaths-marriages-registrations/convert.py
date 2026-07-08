#!/usr/bin/env python3
"""Convert ABS registration statistics (births, deaths, marriages, divorces) into tidy CSVs.

Source files (raw/):
  - abs-births-summary-by-state-1975-2024.csv (ABS.Stat SDMX data API, dataflow BIRTHS_SUMMARY)
  - abs-deaths-summary-by-state-1971-2024.csv (ABS.Stat SDMX data API, dataflow DEATHS_SUMMARY)
  - abs-marriages-and-divorces-australia-2024.xlsx (Marriages and Divorces, Australia, 2024 data cube)

The two ABS.Stat CSVs come pre-coded (numeric MEASURE/REGION/SEX codes); this
script decodes them against the codelists published by ABS.Stat's own
datastructure endpoint for these two dataflows (CL_BIRTHS_MEASURE, CL_DEATHS_MEASURE,
CL_STATE, CL_SEX), fetched and cross-checked directly rather than assumed.
"""
import csv
import openpyxl

RAW = "raw"
DATA = "data"

STATE = {
    "AUS": "Australia", "1": "New South Wales", "2": "Victoria", "3": "Queensland",
    "4": "South Australia", "5": "Western Australia", "6": "Tasmania",
    "7": "Northern Territory", "8": "Australian Capital Territory", "9": "Other Territories",
}

SEX = {"3": "Persons", "1": "Males", "2": "Females"}

BIRTHS_MEASURE = {
    "1": "Births", "2": "Population", "3": "Crude birth rate", "4": "Male births",
    "5": "Female births", "6": "Sex ratio", "7": "Nuptial births", "8": "Ex-nuptial births",
    "9": "Ex-nuptial, paternity acknowledged births", "10": "Ex-nuptial, paternity not acknowledged births",
    "11": "Fertility rate", "12": "Paternity rate", "13": "Confinements",
    "14": "Median age of mother", "15": "Median age of father", "16": "Median age of first-time mother",
    "17": "Median duration of marriage", "18": "Births where both parents Aboriginal or Torres Strait Islander",
    "19": "Births where only mother Aboriginal or Torres Strait Islander",
    "20": "Births where only father Aboriginal or Torres Strait Islander",
}

DEATHS_MEASURE = {
    "1": "Population", "2": "Median age - population", "3": "Births", "4": "Deaths",
    "5": "Crude death rates", "6": "Median age - deaths", "7": "Sex ratio - deaths",
    "8": "Infant deaths", "9": "Infant mortality rates", "10": "Sex ratio - infant deaths",
    "11": "Standardised death rates", "12": "Age-specific death rate", "13": "Rate ratio",
}


def write_csv(path, header, rows):
    with open(f"{DATA}/{path}", "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)
    print(f"wrote {path}: {len(rows)} rows")


def births_summary():
    rows = []
    with open(f"{RAW}/abs-births-summary-by-state-1975-2024.csv") as f:
        for r in csv.DictReader(f):
            rows.append([
                int(r["TIME_PERIOD"]),
                STATE[r["REGION"]],
                BIRTHS_MEASURE[r["MEASURE"]],
                r["OBS_VALUE"],
                r["UNIT_MEASURE"],
                r["OBS_STATUS"] or "",
            ])
    write_csv(
        "au-births-registrations-by-state-1975-2024.csv",
        ["year", "state", "measure", "value", "unit", "obs_status_flag"],
        rows,
    )
    sa_2024 = [r for r in rows if r[1] == "South Australia" and r[2] == "Births" and r[0] == 2024]
    assert sa_2024 and sa_2024[0][3] == "18444", f"SA 2024 births spot-check failed: {sa_2024}"


def deaths_summary():
    rows = []
    with open(f"{RAW}/abs-deaths-summary-by-state-1971-2024.csv") as f:
        for r in csv.DictReader(f):
            rows.append([
                int(r["TIME_PERIOD"]),
                STATE[r["REGION"]],
                SEX[r["SEX"]],
                DEATHS_MEASURE[r["MEASURE"]],
                r["OBS_VALUE"],
                r["UNIT_MEASURE"],
                r["OBS_STATUS"] or "",
            ])
    write_csv(
        "au-deaths-registrations-by-state-1971-2024.csv",
        ["year", "state", "sex", "measure", "value", "unit", "obs_status_flag"],
        rows,
    )
    sa_2024 = [r for r in rows if r[1] == "South Australia" and r[2] == "Persons" and r[3] == "Deaths" and r[0] == 2024]
    assert sa_2024 and sa_2024[0][4] == "15739", f"SA 2024 deaths spot-check failed: {sa_2024}"


def _read_indicator_table(ws, years, section_header_prefixes):
    """Walk an ABS 'Table N: Selected ... indicators' sheet and emit long-format rows.

    Each row in the sheet is either a section header (single non-empty cell in
    column A, no unit/values) or an indicator row (label, unit, one value per year).
    Sub-headings that are themselves section markers (e.g. 'Age group (years)')
    are folded into the current section rather than treated as new sections.
    """
    rows = []
    section = None
    for row in ws.iter_rows(min_row=1, values_only=True):
        label = row[0]
        unit = row[1]
        values = row[2:2 + len(years)]
        if label is None:
            continue
        if unit is None and all(v is None for v in values):
            # section header row (or a footnote row after the data ends)
            if any(label.startswith(p) for p in section_header_prefixes):
                section = label
            continue
        if section is None:
            continue
        for year, value in zip(years, values):
            if value is None:
                continue
            rows.append([section, label, unit, year, value])
    return rows


def marriages_and_divorces():
    wb = openpyxl.load_workbook(f"{RAW}/abs-marriages-and-divorces-australia-2024.xlsx", data_only=True)
    years = [2020, 2021, 2022, 2023, 2024]
    sections = [
        "All marriages", "State or territory of registration", "Marriages including same and non-binary gender",
        "Males", "Male marriages registered by age", "Male age-specific marriage rates",
        "Females", "Female marriages registered by age", "Female age-specific marriage rates",
        "All divorces", "Median duration of marriage", "Type of applicant", "State or territory",
        "Divorces including couples of the same gender", "Divorces involving children (under 18 years)",
        "Male divorces granted by age", "Male age-specific divorce rates",
        "Female divorces granted by age", "Female age-specific divorce rates",
    ]

    marriage_rows = _read_indicator_table(wb["Table 1"], years, sections)
    write_csv(
        "au-marriage-indicators-by-state-2020-2024.csv",
        ["section", "indicator", "unit", "year", "value"],
        marriage_rows,
    )
    sa = [r for r in marriage_rows if r[0] == "State or territory of registration" and r[1] == "South Australia" and r[3] == 2024]
    assert sa and sa[0][4] == 7500, f"SA 2024 marriages spot-check failed: {sa}"

    divorce_rows = _read_indicator_table(wb["Table 3"], years, sections)
    write_csv(
        "au-divorce-indicators-by-state-2020-2024.csv",
        ["section", "indicator", "unit", "year", "value"],
        divorce_rows,
    )
    sa_d = [r for r in divorce_rows if r[0] == "State or territory" and r[1] == "South Australia" and r[3] == 2024]
    assert sa_d and sa_d[0][4] == 3147, f"SA 2024 divorces spot-check failed: {sa_d}"

    # Table 2: marriages by month of occurrence -- national by year, and 2024 by state
    ws2 = wb["Table 2"]
    months = ["January", "February", "March", "April", "May", "June", "July", "August",
              "September", "October", "November", "December"]
    national_rows, state_rows = [], []
    section = None
    for row in ws2.iter_rows(min_row=1, values_only=True):
        label, unit = row[0], row[1]
        values = row[2:14]
        if label is None:
            continue
        if label == "2024 occurrence by state or territory of registration":
            section = "state"
            continue
        if isinstance(label, int) and section is None:
            for month, value in zip(months, values):
                if value is not None:
                    national_rows.append([label, month, value])
        elif isinstance(label, str) and section == "state" and unit == "no.":
            for month, value in zip(months, values):
                if value is not None:
                    state_rows.append([label, month, value])
    write_csv(
        "au-marriages-by-month-of-occurrence-2020-2024.csv",
        ["year", "month", "marriages"],
        national_rows,
    )
    write_csv(
        "au-marriages-by-month-of-occurrence-by-state-2024.csv",
        ["state", "month", "marriages"],
        state_rows,
    )
    sa_month = [r for r in state_rows if r[0] == "South Australia" and r[1] == "January"]
    assert sa_month and sa_month[0][2] == 560, f"SA January 2024 marriages spot-check failed: {sa_month}"


if __name__ == "__main__":
    births_summary()
    deaths_summary()
    marriages_and_divorces()
    print("All spot-checks passed.")
