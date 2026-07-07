#!/usr/bin/env python3
"""Convert the 5 ABS "Building Approvals, Australia" time-series workbooks
mirrored in raw/ into tidy long-format CSVs.

Each source workbook is a standard ABS Time Series Workbook: sheet "Data1"
has one column per (measure, [dwelling-type,] geography, series-type)
combination, with:
  - row 1: column description, semicolon-separated, e.g.
           "Total number of dwelling units ;  South Australia ;"
           or (table 10 only) "... ;  Houses ;  Greater Adelaide ;"
  - row 2: Unit (e.g. "Number", "$'000")
  - row 3: Series Type (Original / Seasonally Adjusted / Trend)
  - row 10: Series ID (ABS's own time-series identifier)
  - row 11 onward: one row per month, column A = reference month, other
    columns = the value for that column's series.

This script only unpivots wide (one column per series) into long (one row
per observation) and reads the source's own header metadata rows to label
each row - it does not recalculate, re-derive, seasonally adjust or
reinterpret any figure. Cell values are copied exactly as stored in the
workbook (ABS publishes original/seasonally-adjusted/trend as separate
already-computed series, not something this script derives).
"""
import csv
from pathlib import Path

import openpyxl

RAW = Path(__file__).parent / "raw"
OUT = Path(__file__).parent / "data"

# (raw filename, output stem, table number, table title, geography level)
TABLES = [
    (
        "8731007-total-dwelling-units-approved-states-and-territories.xlsx",
        "table-07-dwelling-units-by-state",
        "07",
        "Total number of dwelling units approved - states and territories",
        "state",
    ),
    (
        "87310039-value-of-total-building-approved-states-and-territories.xlsx",
        "table-39-value-of-total-building-by-state",
        "39",
        "Value of total building approved - states and territories",
        "state",
    ),
    (
        "87310040-value-of-residential-building-approved-states-and-territories.xlsx",
        "table-40-value-of-residential-building-by-state",
        "40",
        "Value of residential building approved - states and territories",
        "state",
    ),
    (
        "87310041-value-of-non-residential-building-approved-states-and-territories.xlsx",
        "table-41-value-of-non-residential-building-by-state",
        "41",
        "Value of non-residential building approved - states and territories",
        "state",
    ),
    (
        "87310010-dwelling-units-approved-by-gccsa-original.xlsx",
        "table-10-dwelling-units-by-gccsa",
        "10",
        "Number of dwelling units approved, by Greater Capital City Statistical Area - original",
        "gccsa",
    ),
]

LONG_HEADER = [
    "table",
    "table_title",
    "reference_month",
    "geography_level",
    "geography",
    "dwelling_type",
    "series_type",
    "unit",
    "series_id",
    "value",
]

# South Australia is identified by exact match on either the state name or,
# for table 10 (GCCSA), the SA capital-city statistical area name.
SA_GEOGRAPHIES = {"South Australia", "Greater Adelaide"}


def parse_header_desc(desc):
    """Split row-1's semicolon-separated column description into
    (dwelling_type, geography). Table 07 headers have 2 parts:
    'Total number of dwelling units ;  South Australia ;' (measure ; geography;
    no dwelling/building-sector type). Tables 10/39/40/41 headers have 3 parts,
    e.g. 'Total number of dwelling units ;  Houses ;  Greater Adelaide ;' or
    'Total value of building jobs ;  Total Non-residential ;  South Australia ;'
    (measure ; dwelling_or_sector_type ; geography ;). Detected by part count,
    not by table/geography level, since the 3-part shape appears in both the
    state tables (39/40/41, where the middle part is a building-sector type
    like "Total Residential") and the GCCSA table (10, dwelling type).
    """
    parts = [p.strip() for p in desc.split(";") if p.strip() != ""]
    if len(parts) == 3:
        dwelling_type = parts[1]
        geography = parts[2]
    else:
        dwelling_type = ""
        geography = parts[1]
    return dwelling_type, geography


def convert_one(fname, stem, table_no, table_title, geography_level):
    wb = openpyxl.load_workbook(RAW / fname, data_only=True)
    ws = wb["Data1"]
    ncols = ws.max_column

    descs = [ws.cell(row=1, column=c).value for c in range(2, ncols + 1)]
    units = [ws.cell(row=2, column=c).value for c in range(2, ncols + 1)]
    series_types = [ws.cell(row=3, column=c).value for c in range(2, ncols + 1)]
    series_ids = [ws.cell(row=10, column=c).value for c in range(2, ncols + 1)]

    col_meta = []
    for desc, unit, stype, sid in zip(descs, units, series_types, series_ids):
        dwelling_type, geography = parse_header_desc(desc)
        col_meta.append(
            {
                "dwelling_type": dwelling_type,
                "geography": geography,
                "series_type": stype,
                "unit": unit,
                "series_id": sid,
            }
        )

    rows_out = []
    for r in range(11, ws.max_row + 1):
        ref_month_cell = ws.cell(row=r, column=1).value
        if ref_month_cell is None:
            continue
        reference_month = ref_month_cell.strftime("%Y-%m")
        for i, meta in enumerate(col_meta):
            value = ws.cell(row=r, column=i + 2).value
            if value is None or value == "":
                continue
            rows_out.append(
                {
                    "table": table_no,
                    "table_title": table_title,
                    "reference_month": reference_month,
                    "geography_level": geography_level,
                    "geography": meta["geography"],
                    "dwelling_type": meta["dwelling_type"],
                    "series_type": meta["series_type"],
                    "unit": meta["unit"],
                    "series_id": meta["series_id"],
                    "value": value,
                }
            )

    out_file = OUT / f"{stem}.csv"
    with out_file.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=LONG_HEADER)
        w.writeheader()
        w.writerows(rows_out)

    return rows_out


def main():
    OUT.mkdir(exist_ok=True)
    all_rows = []
    table_index = []

    for fname, stem, table_no, table_title, geography_level in TABLES:
        rows_out = convert_one(fname, stem, table_no, table_title, geography_level)
        all_rows.extend(rows_out)
        table_index.append(
            {
                "table": table_no,
                "title": table_title,
                "geography_level": geography_level,
                "rows": len(rows_out),
                "file": f"{stem}.csv",
            }
        )
        print(f"table {table_no}: {len(rows_out)} rows -> {stem}.csv")

    with (OUT / "all-tables-long.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=LONG_HEADER)
        w.writeheader()
        w.writerows(all_rows)

    sa_rows = [r for r in all_rows if r["geography"] in SA_GEOGRAPHIES]
    with (OUT / "south-australia.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=LONG_HEADER)
        w.writeheader()
        w.writerows(sa_rows)

    with (OUT / "table-index.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["table", "title", "geography_level", "rows", "file"])
        for t in table_index:
            w.writerow([t["table"], t["title"], t["geography_level"], t["rows"], t["file"]])

    print(f"\nWrote {len(TABLES)} per-table files, {len(all_rows)} total long rows, "
          f"{len(sa_rows)} South Australia / Greater Adelaide rows.")


if __name__ == "__main__":
    main()
