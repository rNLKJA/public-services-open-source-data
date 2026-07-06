"""
Flatten the Productivity Commission's Report on Government Services (RoGS)
"Child protection services" data tables (raw/*.xlsx, sheets "Table 16A.1"
through "Table 16A.44") into tidy long-format CSVs under data/.

Every sheet shares the same layout: a title row, a header row with "Unit" at
column L (index 11) followed by that table's breakdown columns (jurisdiction
abbreviation for most tables, financial year for the per-jurisdiction
activity-group unit-cost tables 16A.27-16A.35 and 16A.39, or an indicator
name for 16A.11), then data rows where a row's label sits in whichever of
columns A-K (indices 0-10) matches its indentation depth - the same column
holds every row at that depth, deeper rows use a column further right. This
script walks each row, tracks the current label at each depth in a stack, and
whenever a row also carries a unit + values, emits one record per non-empty
value using the full label path built from the stack. No value is
recalculated or reinterpreted - only reshaped from a wide indented grid into
one row per observation.
"""
import csv
import re
import openpyxl
from pathlib import Path

RAW = Path(__file__).parent / "raw" / "rogs-2026-partf-section16-child-protection-data-tables.xlsx"
OUT = Path(__file__).parent / "data"

FOOTNOTE_RE = re.compile(r"\xa0\(")


def clean_text(value):
    """Strip trailing footnote markers like " (a), (b)" and normalise nbsp to a space."""
    if value is None:
        return None
    text = str(value)
    text = FOOTNOTE_RE.split(text)[0]
    return text.replace("\xa0", " ").strip()


def parse_contents(ws):
    index = []
    for row in ws.iter_rows(min_row=1, max_row=100, values_only=True):
        cell = row[0]
        if not isinstance(cell, str) or not cell.startswith("Table "):
            continue
        index.append({
            "table": cell.replace("Table ", ""),
            "latest_update": row[1],
            "title": clean_text(row[2]),
        })
    return index


def parse_table(ws):
    rows = list(ws.iter_rows(values_only=True))

    header_idx = next(
        i for i, row in enumerate(rows)
        if len(row) > 11 and row[11] is not None and str(row[11]).strip() == "Unit"
    )
    header_row = rows[header_idx]
    columns = [
        (j, clean_text(header_row[j]))
        for j in range(12, len(header_row))
        if header_row[j] is not None
    ]

    records = []
    stack = {}
    for row in rows[header_idx + 1:]:
        label_col, label_text = None, None
        for c in range(0, 11):
            if c < len(row) and row[c] is not None:
                label_col, label_text = c, clean_text(row[c])
                break
        if label_col is not None:
            for depth in [d for d in stack if d > label_col]:
                del stack[depth]
            stack[label_col] = label_text

        unit = row[11] if len(row) > 11 else None
        if unit is None:
            continue
        values = [(colname, row[j]) for j, colname in columns if j < len(row) and row[j] is not None]
        if not values:
            continue

        category = " > ".join(stack[d] for d in sorted(stack))
        unit_clean = clean_text(unit)
        for colname, value in values:
            records.append({
                "category": category,
                "unit": unit_clean,
                "breakdown": colname,
                "value": value,
            })
    return records


def main():
    OUT.mkdir(exist_ok=True)
    wb = openpyxl.load_workbook(RAW, read_only=True, data_only=True)

    index = parse_contents(wb["Contents"])
    titles = {e["table"]: e["title"] for e in index}

    all_rows = []
    for tname in wb.sheetnames:
        if not tname.startswith("Table "):
            continue
        table_id = tname.replace("Table ", "")
        title = titles.get(table_id, "")
        records = parse_table(wb[tname])
        slug = "table-" + table_id.replace(".", "-").lower()
        with open(OUT / f"{slug}.csv", "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=["category", "unit", "breakdown", "value"])
            w.writeheader()
            w.writerows(records)
        for r in records:
            all_rows.append({"table": table_id, "title": title, **r})
        print(f"Table {table_id}: {len(records)} rows -> {slug}.csv")

    with open(OUT / "all-tables-long.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["table", "title", "category", "unit", "breakdown", "value"])
        w.writeheader()
        w.writerows(all_rows)
    print(f"all-tables-long.csv: {len(all_rows)} rows")

    with open(OUT / "table-index.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["table", "title", "latest_update"])
        w.writeheader()
        w.writerows(index)
    print(f"table-index.csv: {len(index)} rows")


if __name__ == "__main__":
    main()
