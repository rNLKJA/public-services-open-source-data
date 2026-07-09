"""
Converts the AIHW "Immunisation rates for children in 2016-17" source workbook
(raw/aihw-immunisation-rates-for-children-2016-17-data-tables.xlsx) into tidy,
directly-loadable CSVs. No figures are recalculated - only reshaped, decoded
and given consistent snake_case column names.

Run from the dataset's data/ folder: python3 convert.py
"""
import csv
import re
from pathlib import Path

import openpyxl

RAW = Path(__file__).parent.parent / "raw" / "aihw-immunisation-rates-for-children-2016-17-data-tables.xlsx"
OUT = Path(__file__).parent


def sa_related(state_code):
    if not state_code:
        return False
    tokens = [t.strip() for t in re.split(r"[/]", str(state_code))]
    return "SA" in tokens


def find_header_row(ws, marker="State", min_row=1, max_row=25):
    for i, row in enumerate(ws.iter_rows(min_row=min_row, max_row=max_row, values_only=True), start=min_row):
        if row and row[0] == marker:
            return i
    raise ValueError(f"header row not found in {ws.title}")


def data_rows(ws, header_row):
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row[0] is None or row[0] == "":
            continue
        yield row


def write_csv(path, header, rows):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)
    print(f"wrote {path.name}: {len(rows)} rows")


def main():
    wb = openpyxl.load_workbook(RAW, data_only=True)

    # --- TAB 1: national, both population groups side by side -> long format ---
    ws = wb["TAB 1"]
    header_row = find_header_row(ws, marker="Reporting Year")
    out = []
    for row in data_rows(ws, header_row):
        reporting_year, age_group, all_reg, all_full, all_notfull, all_pct, _, atsi_reg, atsi_full, atsi_notfull, atsi_pct = row[:11]
        out.append([reporting_year, age_group, "All children", all_reg, all_full, all_notfull, all_pct])
        out.append([reporting_year, age_group, "Aboriginal and Torres Strait Islander children", atsi_reg, atsi_full, atsi_notfull, atsi_pct])
    write_csv(
        OUT / "immunisation-national.csv",
        ["reporting_year", "age_group", "population_group", "num_registered_children", "num_fully_immunised", "num_not_fully_immunised", "percent_fully_immunised"],
        out,
    )

    # --- TAB 2 (All children, by PHN) + TAB 5 (ATSI children, by PHN) merged ---
    out = []
    ws2 = wb["TAB 2"]
    h2 = find_header_row(ws2, marker="State")
    for row in data_rows(ws2, h2):
        state, phn_code, phn_name, reporting_year, age_group, reg, full, notfull, pct = row[:9]
        out.append([state, sa_related(state), phn_code, phn_name, reporting_year, age_group, "All children", reg, full, notfull, pct, ""])

    ws5 = wb["TAB 5"]
    h5 = find_header_row(ws5, marker="State")
    for row in data_rows(ws5, h5):
        state, phn_code, phn_name, reporting_year, age_group, reg, full, notfull, pct, caution = (list(row) + [None] * 10)[:10]
        out.append([state, sa_related(state), phn_code, phn_name, reporting_year, age_group, "Aboriginal and Torres Strait Islander children", reg, full, notfull, pct, caution or ""])

    write_csv(
        OUT / "immunisation-by-phn.csv",
        ["state", "south_australia_related", "phn_code", "phn_area_name", "reporting_year", "age_group", "population_group", "num_registered_children", "num_fully_immunised", "num_not_fully_immunised", "percent_fully_immunised", "interpret_with_caution"],
        out,
    )

    # --- TAB 3: All children, by SA3 ---
    ws = wb["TAB 3"]
    header_row = find_header_row(ws, marker="State")
    out = []
    for row in data_rows(ws, header_row):
        state, sa3_code, sa3_name, reporting_year, age_group, reg, full, notfull, pct, caution = (list(row) + [None] * 10)[:10]
        out.append([state, sa_related(state), sa3_code, sa3_name, reporting_year, age_group, reg, full, notfull, pct, caution or ""])
    write_csv(
        OUT / "immunisation-by-sa3.csv",
        ["state", "south_australia_related", "sa3_code", "sa3_name", "reporting_year", "age_group", "num_registered_children", "num_fully_immunised", "num_not_fully_immunised", "percent_fully_immunised", "interpret_with_caution"],
        out,
    )

    # --- TAB 4: All children, by postcode (banded percent only, no raw counts) ---
    ws = wb["TAB 4"]
    header_row = find_header_row(ws, marker="State")
    out = []
    for row in data_rows(ws, header_row):
        state, postcode, areas, reporting_year, age_group, pct_band, caution = (list(row) + [None] * 7)[:7]
        out.append([state, sa_related(state), postcode, areas, reporting_year, age_group, pct_band, caution or ""])
    write_csv(
        OUT / "immunisation-by-postcode.csv",
        ["state", "south_australia_related", "postcode", "associated_residential_areas", "reporting_year", "age_group", "percent_fully_immunised_band", "interpret_with_caution"],
        out,
    )

    # --- TAB 6: ATSI children, by SA4 ---
    ws = wb["TAB 6"]
    header_row = find_header_row(ws, marker="State")
    out = []
    for row in data_rows(ws, header_row):
        state, sa4_code, sa4_name, reporting_year, age_group, reg, full, notfull, pct, caution = (list(row) + [None] * 10)[:10]
        out.append([state, sa_related(state), sa4_code, sa4_name, reporting_year, age_group, reg, full, notfull, pct, caution or ""])
    write_csv(
        OUT / "immunisation-atsi-by-sa4.csv",
        ["state", "south_australia_related", "sa4_code", "sa4_name", "reporting_year", "age_group", "num_registered_children", "num_fully_immunised", "num_not_fully_immunised", "percent_fully_immunised", "interpret_with_caution"],
        out,
    )


if __name__ == "__main__":
    main()
