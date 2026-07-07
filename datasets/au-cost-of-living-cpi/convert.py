"""
Convert ABS Consumer Price Index, Australia (May 2026 release) raw time-series
workbooks into tidy CSVs focused on the Adelaide / South Australia breakdown.

Source tables (raw/, all CC BY 4.0, ABS catalogue 6401.0):
  640101  TABLE 1.  CPI: All Groups, Index numbers and Percentage change
  640109  TABLE 9.  CPI: Groups, Index Numbers by Capital City
  6401011 TABLE 11. CPI: Group, Sub-group and Expenditure Class,
                     Annual percentage change, by Capital City
  6401012 TABLE 12. CPI: Group, Sub-group and Expenditure Class,
                     Monthly percentage change by Capital City

Each is a standard ABS "Time Series Workbook": an Index sheet describing every
series (one row per series, format "<measure> ;  <item> ;  <city> ;"), and one
or more Data sheets where row 1 repeats that description per column, row 10
holds the Series ID, and data rows below are (date, value) pairs.
"""
import csv
import re
from pathlib import Path

import openpyxl

RAW = Path(__file__).parent / "raw"
OUT = Path(__file__).parent / "data"
OUT.mkdir(exist_ok=True)


def read_series(path, target_city):
    """Return {(measure, item): {date: value}} for one workbook, filtered to target_city."""
    wb = openpyxl.load_workbook(path, data_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("Data"):
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        data_rows = rows[10:]
        for col_idx in range(1, len(header)):
            desc = header[col_idx]
            if not desc or target_city not in desc:
                continue
            parts = [p.strip() for p in desc.split(";") if p.strip()]
            if len(parts) < 3:
                continue
            measure, item, city = parts[0], parts[1], parts[2]
            if city != target_city:
                continue
            series = {}
            for row in data_rows:
                date = row[0]
                value = row[col_idx] if col_idx < len(row) else None
                if date is None or value is None:
                    continue
                series[date.date().isoformat()[:7]] = value
            result[(measure, item)] = series
    return result


def build_adelaide_item_table():
    city = "Adelaide"
    index_numbers = read_series(RAW / "640109-groups-index-numbers-by-capital-city.xlsx", city)
    annual_pct = read_series(RAW / "6401011-annual-pct-change-by-capital-city.xlsx", city)
    monthly_pct = read_series(RAW / "6401012-monthly-pct-change-by-capital-city.xlsx", city)

    group_level_items = {item for (measure, item) in index_numbers if measure == "Index Numbers"}

    # Collect the full (item, date) key space across all three measures.
    keys = set()
    idx_by_item = {}
    for (measure, item), series in index_numbers.items():
        if measure != "Index Numbers":
            continue
        idx_by_item[item] = series
        for date in series:
            keys.add((item, date))

    ann_by_item = {}
    for (measure, item), series in annual_pct.items():
        if measure != "Percentage Change from Corresponding Month of Previous Year":
            continue
        ann_by_item[item] = series
        for date in series:
            keys.add((item, date))

    mth_by_item = {}
    for (measure, item), series in monthly_pct.items():
        if measure != "Percentage Change from Previous Period":
            continue
        mth_by_item[item] = series
        for date in series:
            keys.add((item, date))

    rows = []
    for item, date in sorted(keys):
        rows.append({
            "date": date,
            "item": item,
            "is_cpi_group_level": item in group_level_items,
            "index_number": idx_by_item.get(item, {}).get(date, ""),
            "annual_pct_change": ann_by_item.get(item, {}).get(date, ""),
            "monthly_pct_change": mth_by_item.get(item, {}).get(date, ""),
        })

    out_path = OUT / "adelaide-cpi-by-item.csv"
    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "date", "item", "is_cpi_group_level",
            "index_number", "annual_pct_change", "monthly_pct_change",
        ])
        writer.writeheader()
        writer.writerows(rows)
    distinct_items = set(idx_by_item) | set(ann_by_item) | set(mth_by_item)
    print(f"Wrote {out_path} ({len(rows)} rows, {len(distinct_items)} distinct items)")
    return rows


def build_headline_comparison_table():
    cities = ["Australia", "Sydney", "Melbourne", "Brisbane", "Adelaide", "Perth", "Hobart", "Darwin", "Canberra"]
    path = RAW / "640101-all-groups-index-numbers-pct-change.xlsx"

    rows_by_key = {}
    for city in cities:
        series_by_measure = read_series(path, city)
        for (measure, item), series in series_by_measure.items():
            if item != "All groups CPI":
                continue
            for date, value in series.items():
                key = (date, city)
                rows_by_key.setdefault(key, {"date": date, "location": city,
                                              "index_number": "", "annual_pct_change": "", "monthly_pct_change": ""})
                if measure == "Index Numbers":
                    rows_by_key[key]["index_number"] = value
                elif measure == "Percentage Change from Corresponding Month of Previous Year":
                    rows_by_key[key]["annual_pct_change"] = value
                elif measure == "Percentage Change from Previous Period":
                    rows_by_key[key]["monthly_pct_change"] = value

    rows = sorted(rows_by_key.values(), key=lambda r: (r["date"], r["location"]))
    out_path = OUT / "all-capital-cities-headline-cpi.csv"
    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["date", "location", "index_number", "annual_pct_change", "monthly_pct_change"])
        writer.writeheader()
        writer.writerows(rows)
    print(f"Wrote {out_path} ({len(rows)} rows)")
    return rows


if __name__ == "__main__":
    item_rows = build_adelaide_item_table()
    headline_rows = build_headline_comparison_table()

    # Spot checks against the raw workbooks (values sourced directly from the
    # Index/Data sheets above, not recomputed) before trusting the output.
    adelaide_all_groups_may2026 = [r for r in item_rows if r["item"] == "All groups CPI" and r["date"] == "2026-05"][0]
    assert adelaide_all_groups_may2026["index_number"] == 102.67, adelaide_all_groups_may2026
    assert round(adelaide_all_groups_may2026["annual_pct_change"], 1) == 4.4, adelaide_all_groups_may2026

    headline_adelaide_may2026 = [r for r in headline_rows if r["location"] == "Adelaide" and r["date"] == "2026-05"][0]
    assert headline_adelaide_may2026["index_number"] == adelaide_all_groups_may2026["index_number"]

    print("Spot checks passed.")
