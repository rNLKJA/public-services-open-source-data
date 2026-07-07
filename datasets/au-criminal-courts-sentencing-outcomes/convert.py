#!/usr/bin/env python3
"""
Converts the ABS "Criminal Courts, Australia, 2024-25" workbooks mirrored in
raw/ into tidy, directly-loadable long-format CSVs, without altering,
recalculating or reinterpreting any published figures.

Source files (mirrored verbatim in raw/):
  7-defendants-finalised-south-australia-tables-39-to-44.xlsx
      Tables 39-44: defendants finalised in South Australia's courts -
      summary characteristics by court level and year (Table 39), principal
      offence by method of finalisation (Table 40), and summary outcomes
      (method of finalisation + principal sentence) by principal offence for
      All Courts / Higher Courts / Magistrates' Courts / Children's Courts
      (Tables 41-44).
  13-sentence-length-and-fine-amount-australia-tables-78-to-84.xlsx
      National sentence-length/fine-amount cube. Table 78 is national-only
      (All Courts / Higher / Magistrates' / Children's, no state split) and is
      NOT extracted into a South-Australia-labelled file because it carries
      no state dimension at all; it is still converted in full (national
      figures only) for reference. Tables 79-84 each break every measure down
      by state/territory in stacked row-blocks (Australia, New South Wales,
      Victoria, Queensland, South Australia, Western Australia, Tasmania,
      Northern Territory, Australian Capital Territory) - for these, this
      script writes one CSV containing every state (so an Australia-wide
      comparison stays available) AND a filtered South-Australia-only CSV,
      since the SA rows are the ones this repository is built around.

Two structurally distinct table shapes appear across both workbooks:

  A) "Nested row-group" tables (Table 39; Tables 79-84's per-state blocks) -
     a single header row (or joined multi-row header) of column labels, with
     the row axis broken into named sections (e.g. "Sex", "Age", "Principal
     offence (ANZSOC 2023)", or a state/territory name) that this script
     forward-fills into a `row_group` column so the section survives outside
     the spreadsheet's visual layout. Converted to:
       table, table_title, section, row_group, row_label, column, value

  B) "Repeated year-block" tables (Tables 41-44) - one shared set of offence
     columns (from row 4), with the same two row-groups (Method of
     finalisation, Principal sentence) repeated once per financial year
     (2024-25, then 2023-24). The year is recovered from the row that
     introduces each repeated block and added as its own field.
       table, table_title, financial_year, row_group, row_label, column, value

  Table 40 is a two-level offence-hierarchy x method-of-finalisation grid
  with a single header row and no row-group sections at all; it uses the
  same schema as shape (A) with `section`/`row_group` left blank and an
  `offence_division`/`offence_subdivision` breakdown folded into row_label.

No totals are recomputed, no percentages derived, no cell values changed -
this script only reshapes wide ABS tables into long rows, joins multi-row
headers, and forward-fills the section/row-group labels ABS already prints
in the spreadsheet. ANZSOC offence codes and all other category labels are
already human-readable text in the source (e.g. "01 Homicide", "021 Serious
assault") - there is no separate numeric-code lookup to decode.
"""
import csv
import re
from pathlib import Path

import openpyxl

RAW_DIR = Path(__file__).parent / "raw"
SA_FILE = RAW_DIR / "7-defendants-finalised-south-australia-tables-39-to-44.xlsx"
NATIONAL_FILE = RAW_DIR / "13-sentence-length-and-fine-amount-australia-tables-78-to-84.xlsx"
OUT_DIR = Path(__file__).parent / "data"
OUT_DIR.mkdir(parents=True, exist_ok=True)

FOOTNOTE_RE = re.compile(r"\s*\([a-z]\)")  # strips trailing " (a)", " (b)(c)" etc. from titles/labels
FIELDS_A = ["table", "table_title", "section", "row_group", "row_label", "column", "value"]
FIELDS_B = ["table", "table_title", "financial_year", "row_group", "row_label", "column", "value"]


def clean(s):
    text = str(s).replace("\xa0", " ").replace("\n", " ")
    text = FOOTNOTE_RE.sub("", text)
    return re.sub(r"\s+", " ", text).strip()


def trim_row(row):
    vals = list(row)
    while vals and vals[-1] is None:
        vals.pop()
    return vals


def get_title(values_rows):
    raw = values_rows[1][0] if len(values_rows) > 1 and values_rows[1] else ""
    # Title rows are formatted "Table NN <plain-English title>"
    m = re.match(r"Table\s+\d+\s+(.*)", str(raw))
    return clean(m.group(1)) if m else clean(raw)


def build_header(header_rows, n_cols):
    """Join up to 2 stacked header rows into one column label per column,
    forward-filling merged-cell blanks left to right (openpyxl returns None
    for non-top-left cells of a merged range)."""
    top = []
    last = None
    for i in range(n_cols):
        v = header_rows[0][i] if i < len(header_rows[0]) else None
        if v is not None and str(v).strip() != "":
            last = clean(v)
        top.append(last if i > 0 else (clean(v) if v else None))

    filled_rows = [top]
    for hr in header_rows[1:]:
        filled = []
        last = None
        last_group = None
        for i in range(n_cols):
            v = hr[i] if i < len(hr) else None
            group = top[i]
            if group != last_group:
                last = None
                last_group = group
            if v is not None and str(v).strip() != "":
                last = clean(v)
                filled.append(last)
            else:
                filled.append(last if i > 0 else None)
        filled_rows.append(filled)

    columns = []
    for i in range(n_cols):
        parts = []
        for fr in filled_rows:
            v = fr[i]
            if v and v not in parts:
                parts.append(v)
        columns.append(" - ".join(parts) if parts else f"col_{i}")
    return columns


def is_footnote_or_copyright(label):
    low = label.lower()
    return (
        label.startswith("©")
        or label.startswith("(")
        or low.startswith("footnotes")
        or low.startswith("due to perturbation")
        or low.startswith("differences in legislation")
        or low.startswith("users are advised")
        or low in (".. not applicable", "np not published", "na not available", "footnotes")
    )


def convert_nested_sections(ws, section_rows, header_row_idx, header_span, data_start):
    """Shape A: forward-fills a stack of (section, row_group) labels seen as
    header-only rows (a row with a label in column 0 and no data in the rest
    of the row) ahead of each data row. `section_rows` gives the row indices
    (0-based) that mark the start of a brand-new top-level section (e.g. a
    state/territory name, or "All Courts"/"Higher Courts" court-level block)
    so that a plain sub-heading like "Sex" nested under it isn't confused
    with a new top-level section."""
    values_rows = [trim_row(r) for r in ws.iter_rows(values_only=True)]
    header_rows = values_rows[header_row_idx:header_row_idx + header_span]
    n_cols = max(
        max((len(r) for r in header_rows), default=0),
        max((len(r) for r in values_rows[data_start:] if r), default=0),
    )
    columns = build_header(header_rows, n_cols)

    out_rows = []
    section = ""
    row_group = ""
    for idx in range(data_start, len(values_rows)):
        r = values_rows[idx]
        if idx in section_rows:
            # A brand new top-level section (state/territory name or court-level
            # block). Marker rows have the section name in column 1, not column 0.
            section = clean(r[1]) if len(r) > 1 and r[1] not in (None, "") else clean(r[0])
            row_group = ""
            continue
        if not r or r[0] is None or str(r[0]).strip() == "":
            continue
        label_raw = clean(r[0])
        if is_footnote_or_copyright(label_raw):
            continue
        has_data = len(r) > 1 and any(v is not None and str(v).strip() != "" for v in r[1:])
        if not has_data:
            row_group = label_raw
            continue
        for i in range(1, len(r)):
            v = r[i]
            if v is None or str(v).strip() == "":
                continue
            out_rows.append({
                "section": section,
                "row_group": row_group,
                "row_label": label_raw,
                "column": columns[i] if i < len(columns) else f"col_{i}",
                "value": v,
            })
    return out_rows


def convert_table_39(ws):
    """Table 39: 4 court-level sections (All Courts, Higher Courts,
    Magistrates' Courts, Children's Courts), each with sub-groups Sex / Age /
    Principal offence / Duration / Method of finalisation / Principal
    sentence, columns = financial years 2010-11 to 2024-25."""
    values_rows = [trim_row(r) for r in ws.iter_rows(values_only=True)]
    court_level_rows = {
        idx for idx, r in enumerate(values_rows)
        if r and r[0] is None and len(r) > 1 and r[1] not in (None, "") and idx > 4
        and str(r[1]).strip() in ("All Courts", "Higher Courts", "Magistrates' Courts", "Children's Courts")
    }
    return convert_nested_sections(ws, court_level_rows, header_row_idx=4, header_span=1, data_start=5)


def convert_table_40(ws):
    """Table 40: single offence-hierarchy x method-of-finalisation grid, no
    row-group sections. Uses shape-A schema with section/row_group blank."""
    return convert_nested_sections(ws, section_rows=set(), header_row_idx=4, header_span=1, data_start=5)


def convert_year_block_table(ws):
    """Shape B: Tables 41-44. Row 4 = offence-category column header (single
    row). Data is organised as repeated year-blocks: a row with column 0
    blank and column 1 holding the year label (e.g. "2024-25") introduces
    each block; within a block, two row-groups appear ("Method of
    finalisation", "Principal sentence") each with their own data rows."""
    values_rows = [trim_row(r) for r in ws.iter_rows(values_only=True)]
    header_rows = [values_rows[4]]
    n_cols = max(len(values_rows[4]), max((len(r) for r in values_rows[6:] if r), default=0))
    columns = build_header(header_rows, n_cols)

    out_rows = []
    year = ""
    row_group = ""
    for idx in range(5, len(values_rows)):
        r = values_rows[idx]
        if not r or r[0] is None or str(r[0]).strip() == "":
            # A blank-column-0 row with a value in column 1 introduces a new year block.
            if r and len(r) > 1 and r[1] not in (None, ""):
                year = clean(r[1])
                row_group = ""
            continue
        label_raw = clean(r[0])
        if is_footnote_or_copyright(label_raw):
            continue
        has_data = len(r) > 1 and any(v is not None and str(v).strip() != "" for v in r[1:])
        if not has_data:
            row_group = label_raw
            continue
        for i in range(1, len(r)):
            v = r[i]
            if v is None or str(v).strip() == "":
                continue
            out_rows.append({
                "financial_year": year,
                "row_group": row_group,
                "row_label": label_raw,
                "column": columns[i] if i < len(columns) else f"col_{i}",
                "value": v,
            })
    return out_rows


def convert_state_breakdown_table(ws):
    """Shape A variant for Tables 79-84 (and 78, treated as a single implicit
    'Australia' section): each state/territory name is a top-level section,
    rows underneath are offence categories (or fine/sentence-length bands for
    Table 78 style), columns from the (up to 2-row) header block starting at
    row 4."""
    values_rows = [trim_row(r) for r in ws.iter_rows(values_only=True)]
    # Header block: row 4 always present; row 5 is a continuation header iff
    # its own column 0 is empty AND it is not itself a state/territory marker row.
    header_rows = [values_rows[4]]
    data_start = 5
    if len(values_rows) > 5 and values_rows[5] and values_rows[5][0] is None:
        # Could be a continuation header (units row, e.g. "(no.)") or the
        # first state marker row ("Australia"). Continuation header rows are
        # short unit strings repeated across many columns; a state marker row
        # has exactly one non-blank cell (at column 1) naming the state.
        non_blank = [v for v in values_rows[5][1:] if v not in (None, "")]
        looks_like_state_marker = len(non_blank) == 1 and str(non_blank[0]).strip() in STATE_NAMES_RE
        if not looks_like_state_marker:
            header_rows.append(values_rows[5])
            data_start = 6
    n_cols = max(
        max((len(r) for r in header_rows), default=0),
        max((len(r) for r in values_rows[data_start:] if r), default=0),
    )
    columns = build_header(header_rows, n_cols)

    section_rows = {
        idx for idx, r in enumerate(values_rows)
        if r and r[0] is None and len(r) > 1 and r[1] not in (None, "") and idx >= data_start
        and clean(r[1]) in STATE_NAMES_RE
    }

    out_rows = []
    section = "Australia"  # Table 78 has no state marker row at all - implicit single section
    for idx in range(data_start, len(values_rows)):
        r = values_rows[idx]
        if idx in section_rows:
            # Marker row: column 0 is blank, the state/territory name is in column 1.
            section = clean(r[1])
            continue
        if not r or r[0] is None or str(r[0]).strip() == "":
            continue
        label_raw = clean(r[0])
        if is_footnote_or_copyright(label_raw):
            continue
        has_data = len(r) > 1 and any(v is not None and str(v).strip() != "" for v in r[1:])
        if not has_data:
            continue
        for i in range(1, len(r)):
            v = r[i]
            if v is None or str(v).strip() == "":
                continue
            out_rows.append({
                "section": section,
                "row_group": "",
                "row_label": label_raw,
                "column": columns[i] if i < len(columns) else f"col_{i}",
                "value": v,
            })
    return out_rows


STATE_NAMES_RE = {
    "Australia", "New South Wales", "Victoria", "Queensland", "South Australia",
    "Western Australia", "Tasmania", "Northern Territory", "Australian Capital Territory",
}


def write_csv(path, fieldnames, rows):
    with path.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
    print(f"wrote {path.name}: {len(rows)} rows")


def main():
    index_rows = []
    all_shape_a = []
    all_shape_b = []

    # --- SA file: Tables 39-44 ---
    wb = openpyxl.load_workbook(SA_FILE, data_only=True)

    ws39 = wb["Table 39"]
    title39 = get_title([trim_row(r) for r in ws39.iter_rows(values_only=True)])
    rows39 = convert_table_39(ws39)
    for row in rows39:
        row["table"] = "Table 39"
        row["table_title"] = title39
    write_csv(OUT_DIR / "table-39-summary-characteristics-by-court-level.csv", FIELDS_A, rows39)
    index_rows.append({"table": "Table 39", "table_title": title39, "source_file": SA_FILE.name, "geography": "South Australia", "rows": len(rows39), "output_file": "table-39-summary-characteristics-by-court-level.csv"})
    all_shape_a.extend(rows39)

    ws40 = wb["Table 40"]
    title40 = get_title([trim_row(r) for r in ws40.iter_rows(values_only=True)])
    rows40 = convert_table_40(ws40)
    for row in rows40:
        row["table"] = "Table 40"
        row["table_title"] = title40
    write_csv(OUT_DIR / "table-40-principal-offence-by-method-of-finalisation.csv", FIELDS_A, rows40)
    index_rows.append({"table": "Table 40", "table_title": title40, "source_file": SA_FILE.name, "geography": "South Australia", "rows": len(rows40), "output_file": "table-40-principal-offence-by-method-of-finalisation.csv"})
    all_shape_a.extend(rows40)

    court_table_map = {
        "Table 41": ("table-41-summary-outcomes-all-courts.csv", "All Courts"),
        "Table 42": ("table-42-summary-outcomes-higher-courts.csv", "Higher Courts"),
        "Table 43": ("table-43-summary-outcomes-magistrates-courts.csv", "Magistrates' Courts"),
        "Table 44": ("table-44-summary-outcomes-childrens-courts.csv", "Children's Courts"),
    }
    for table_name, (out_name, court_level) in court_table_map.items():
        ws = wb[table_name]
        title = get_title([trim_row(r) for r in ws.iter_rows(values_only=True)])
        rows = convert_year_block_table(ws)
        for row in rows:
            row["table"] = table_name
            row["table_title"] = title
        write_csv(OUT_DIR / out_name, FIELDS_B, rows)
        index_rows.append({"table": table_name, "table_title": title, "source_file": SA_FILE.name, "geography": f"South Australia - {court_level}", "rows": len(rows), "output_file": out_name})
        all_shape_b.extend(rows)

    # --- National sentence-length/fine-amount file: Tables 78-84 ---
    wb2 = openpyxl.load_workbook(NATIONAL_FILE, data_only=True)

    ws78 = wb2["Table 78"]
    title78 = get_title([trim_row(r) for r in ws78.iter_rows(values_only=True)])
    rows78 = convert_state_breakdown_table(ws78)  # no state marker rows -> single implicit "Australia" section
    for row in rows78:
        row["table"] = "Table 78"
        row["table_title"] = title78
    write_csv(OUT_DIR / "table-78-sentence-length-by-court-level-australia.csv", FIELDS_A, rows78)
    index_rows.append({"table": "Table 78", "table_title": title78, "source_file": NATIONAL_FILE.name, "geography": "Australia (national totals only - no state breakdown in source)", "rows": len(rows78), "output_file": "table-78-sentence-length-by-court-level-australia.csv"})
    all_shape_a.extend(rows78)

    state_table_map = {
        "Table 79": "table-79-custody-sentence-length-by-offence-all-courts",
        "Table 80": "table-80-custody-sentence-length-by-offence-higher-courts",
        "Table 81": "table-81-custody-sentence-length-by-offence-magistrates-courts",
        "Table 82": "table-82-custody-sentence-length-by-offence-childrens-courts",
        "Table 83": "table-83-community-service-hours-by-offence-and-court-level",
        "Table 84": "table-84-fine-amount-by-offence-and-court-level",
    }
    for table_name, out_stub in state_table_map.items():
        ws = wb2[table_name]
        title = get_title([trim_row(r) for r in ws.iter_rows(values_only=True)])
        rows = convert_state_breakdown_table(ws)
        for row in rows:
            row["table"] = table_name
            row["table_title"] = title
        all_states_name = f"{out_stub}-all-states.csv"
        sa_only_name = f"{out_stub}-south-australia.csv"
        write_csv(OUT_DIR / all_states_name, FIELDS_A, rows)
        sa_rows = [r for r in rows if r["section"].startswith("South Australia")]
        write_csv(OUT_DIR / sa_only_name, FIELDS_A, sa_rows)
        index_rows.append({"table": table_name, "table_title": title, "source_file": NATIONAL_FILE.name, "geography": "All states/territories", "rows": len(rows), "output_file": all_states_name})
        index_rows.append({"table": table_name, "table_title": title, "source_file": NATIONAL_FILE.name, "geography": "South Australia (filtered)", "rows": len(sa_rows), "output_file": sa_only_name})
        all_shape_a.extend(rows)

    # --- Combined long files (one per schema shape) ---
    write_csv(OUT_DIR / "all-tables-long-section-based.csv", FIELDS_A, all_shape_a)
    write_csv(OUT_DIR / "all-tables-long-year-block-based.csv", FIELDS_B, all_shape_b)

    with (OUT_DIR / "table-index.csv").open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["table", "table_title", "source_file", "geography", "rows", "output_file"])
        writer.writeheader()
        for row in index_rows:
            writer.writerow(row)
    print(f"wrote table-index.csv: {len(index_rows)} rows")


if __name__ == "__main__":
    main()
