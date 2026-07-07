"""
Flatten the AIHW "Family, domestic and sexual violence in Australia" supplementary
data-tables workbook (49 sheets: 1 contents page, 10 collection "notes" pages, 38
data tables) into:

  - data/tables/<table-code>.csv   one tidy CSV per source table, columns as published
  - data/all-tables-long.csv        every table stacked into one long/tidy file
  - data/table-index.csv            table code -> title/collection/custodian/coverage

No figure is recalculated or reinterpreted. The only changes made are: comma
thousands-separators stripped from numeric values, column headers converted to
snake_case, and (for the one wide-format table, Kids Helpline 2) melting year
columns into Year/Value rows so it matches the tidy shape of every other table.
Source data-quality flags (n.a., n.p., --, *, **) are preserved exactly as published.
"""
import csv
import os
import re

import openpyxl

RAW = os.path.join(os.path.dirname(__file__), "raw", "AIHW-FDSV-all-data-download.xlsx")
DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
TABLES_DIR = os.path.join(DATA_DIR, "tables")

NUMERIC_RE = re.compile(r"^-?[\d,]+(\.\d+)?$")

# Kids Helpline and 1800RESPECT have no dedicated "notes" sheet in the workbook;
# their custodian/type/frequency/coverage is taken from each table's own "Source:" line.
MANUAL_COLLECTION_META = {
    "Kids Helpline": {
        "full_name": "Kids Helpline",
        "data_custodian": "Kids Helpline (operated by yourtown)",
        "type": "Administrative (service data)",
        "frequency": "Annual",
        "coverage": "National",
    },
    "1800RESPECT": {
        "full_name": "1800RESPECT",
        "data_custodian": "Australian Government Department of Social Services (unpublished data)",
        "type": "Administrative (service data)",
        "frequency": "Annual",
        "coverage": "National",
    },
}


def slug(s):
    s = str(s).strip().lower()
    s = s.replace("–", "-").replace("—", "-")
    s = re.sub(r"[^0-9a-z]+", "_", s)
    return re.sub(r"_+", "_", s).strip("_")


def clean_value(v):
    """Strip thousands separators from genuine numbers; leave data-quality flags as-is."""
    if v is None:
        return ""
    s = str(v).strip()
    if NUMERIC_RE.match(s):
        return s.replace(",", "")
    return s


def parse_collection_notes(wb):
    """Read each '<CODE> notes' sheet for custodian/type/frequency/coverage metadata."""
    meta = dict(MANUAL_COLLECTION_META)
    for name in wb.sheetnames:
        if not name.endswith(" notes"):
            continue
        code = name[: -len(" notes")]
        ws = wb[name]
        rows = [r[0] for r in ws.iter_rows(min_row=1, max_row=8, values_only=True)]
        full_name = rows[1] if len(rows) > 1 else code
        entry = {"full_name": full_name, "data_custodian": "", "type": "", "frequency": "", "coverage": ""}
        for r in rows[2:8]:
            if not isinstance(r, str):
                continue
            for label, key in (
                ("Data custodian:", "data_custodian"),
                ("Type:", "type"),
                ("Frequency:", "frequency"),
                ("Coverage:", "coverage"),
            ):
                if r.strip().startswith(label):
                    entry[key] = r.split(":", 1)[1].strip()
        meta[code] = entry
    return meta


def find_header_row(ws):
    """Header row is the first row containing a 'Value'/'Mean score' cell, or (for the
    one wide-format table) a 'Unit' cell followed by bare-year numeric columns."""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if not row:
            continue
        if any(isinstance(c, str) and c.strip() in ("Value", "Mean score") for c in row):
            return i, [str(c).strip() if c is not None else "" for c in row]
        has_unit = any(isinstance(c, str) and c.strip() == "Unit" for c in row)
        has_year_cols = any(isinstance(c, (int, float)) and 1900 < c < 2100 for c in row)
        if has_unit and has_year_cols:
            return i, [str(int(c)) if isinstance(c, (int, float)) else (str(c).strip() if c is not None else "") for c in row]
    return None, None


def table_title(ws):
    for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
        return row[0]
    return ""


STANDARD_EXTRA_COLS = ["rse", "95%_moe", "lower_ci", "upper_ci", "data_flag", "data_notes"]


def process_standard_table(ws, header_idx, header):
    """Long-format tables: [...dimension cols..., Unit|Mean score, Value|Mean score, optional quality cols]."""
    value_label = "Value" if "Value" in header else "Mean score"
    value_col = header.index(value_label)
    unit_col = header.index("Unit") if "Unit" in header else None
    extra_cols = {h: idx for idx, h in enumerate(header) if h and h not in ("Unit", "Value", "Mean score") and idx > value_col}
    dim_cols = [(h, idx) for idx, h in enumerate(header) if idx not in (value_col, unit_col) and idx not in extra_cols.values() and h]

    out_header = [slug(h) for h, _ in dim_cols]
    if unit_col is not None:
        out_header.append("unit")
    out_header.append("value")
    for h in extra_cols:
        out_header.append(slug(h))

    rows = []
    for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
        if row is None or all(c is None for c in row):
            continue
        if not any(row[idx] is not None for _, idx in dim_cols):
            continue
        out_row = [row[idx] if idx < len(row) and row[idx] is not None else "" for _, idx in dim_cols]
        if unit_col is not None:
            out_row.append(row[unit_col] if unit_col < len(row) and row[unit_col] is not None else "")
        out_row.append(clean_value(row[value_col]) if value_col < len(row) else "")
        for h, idx in extra_cols.items():
            out_row.append(row[idx] if idx < len(row) and row[idx] is not None else "")
        rows.append(out_row)

    dims_for_long = [slug(h) for h, _ in dim_cols]
    return out_header, rows, dims_for_long, ("unit" if unit_col is not None else None), extra_cols


def process_wide_table(ws, header_idx, header):
    """Kids Helpline 2 only: [Concerns, Unit, 2012, 2013, ..., 2024] -> melt years to rows."""
    id_cols = []
    year_cols = []
    for idx, h in enumerate(header):
        if h == "":
            continue
        if isinstance(h, (int, float)) or (isinstance(h, str) and h.strip().isdigit()):
            year_cols.append((idx, str(int(h)) if not isinstance(h, str) else h.strip()))
        else:
            id_cols.append((h, idx))

    out_header = [slug(h) for h, _ in id_cols] + ["year", "unit", "value"]
    unit_pos = next((i for i, (h, _) in enumerate(id_cols) if h == "Unit"), None)

    rows = []
    for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
        if row is None or all(c is None for c in row):
            continue
        base = [row[idx] if idx < len(row) and row[idx] is not None else "" for _, idx in id_cols]
        for idx, year in year_cols:
            val = row[idx] if idx < len(row) else None
            if val is None:
                continue
            out_row = list(base)
            unit_val = base[unit_pos] if unit_pos is not None else ""
            out_row_no_unit = [v for i, v in enumerate(base) if i != unit_pos] if unit_pos is not None else base
            rows.append(out_row_no_unit + [year, unit_val, clean_value(val)])

    dims_for_long = [slug(h) for h, idx in id_cols if h != "Unit"]
    out_header_no_unit = [slug(h) for h, idx in id_cols if h != "Unit"] + ["year", "unit", "value"]
    return out_header_no_unit, rows, dims_for_long + ["year"], "unit", {}


def main():
    os.makedirs(TABLES_DIR, exist_ok=True)
    wb = openpyxl.load_workbook(RAW, data_only=True)
    collection_meta = parse_collection_notes(wb)

    index_rows = []
    long_rows = []
    long_header = ["collection", "table_code", "table_title", "data_custodian", "dimensions", "unit", "value",
                   "rse", "95pct_moe", "lower_ci", "upper_ci", "data_flag", "data_notes"]

    for name in wb.sheetnames:
        if name == "TOC" or name.endswith("notes"):
            continue
        ws = wb[name]
        header_idx, header = find_header_row(ws)
        if header_idx is None:
            print(f"SKIP (no header found): {name}")
            continue

        collection = name.rsplit(" ", 1)[0] if name[-1].isdigit() and " " in name else name
        title = table_title(ws)
        meta = collection_meta.get(collection, {"data_custodian": "", "type": "", "frequency": "", "coverage": ""})

        if name == "Kids Helpline 2":
            out_header, rows, dim_names, unit_name, extra_cols = process_wide_table(ws, header_idx, header)
        else:
            out_header, rows, dim_names, unit_name, extra_cols = process_standard_table(ws, header_idx, header)

        table_slug = slug(name)
        with open(os.path.join(TABLES_DIR, f"{table_slug}.csv"), "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(out_header)
            w.writerows(rows)

        index_rows.append([name, collection, title, meta.get("data_custodian", ""), meta.get("type", ""),
                            meta.get("frequency", ""), meta.get("coverage", ""), len(rows)])

        extra_positions = {slug(h): i for i, h in enumerate(extra_cols)}
        n_dims = len(dim_names)
        for row in rows:
            dims_str = "; ".join(f"{d}={row[i]}" for i, d in enumerate(dim_names) if row[i] != "")
            unit_val = row[n_dims] if unit_name else ""
            value_val = row[n_dims + 1] if unit_name else row[n_dims]
            base_extra_idx = n_dims + 2 if unit_name else n_dims + 1
            extras = {k: "" for k in ("rse", "95pct_moe", "lower_ci", "upper_ci", "data_flag", "data_notes")}
            key_map = {"rse": "rse", "95_moe": "95pct_moe", "lower_ci": "lower_ci", "upper_ci": "upper_ci",
                       "data_flag": "data_flag", "data_notes": "data_notes"}
            for j, orig_h in enumerate(extra_cols):
                sk = slug(orig_h)
                target = key_map.get(sk, sk)
                if target in extras:
                    extras[target] = row[base_extra_idx + j]
            long_rows.append([
                collection, name, title, meta.get("data_custodian", ""), dims_str, unit_val, value_val,
                extras["rse"], extras["95pct_moe"], extras["lower_ci"], extras["upper_ci"],
                extras["data_flag"], extras["data_notes"],
            ])

        print(f"{name}: {len(rows)} rows")

    with open(os.path.join(DATA_DIR, "all-tables-long.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(long_header)
        w.writerows(long_rows)

    with open(os.path.join(DATA_DIR, "table-index.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["table_code", "collection", "title", "data_custodian", "type", "frequency", "coverage", "row_count"])
        w.writerows(index_rows)

    print(f"\nTotal long rows: {len(long_rows)}")


if __name__ == "__main__":
    main()
