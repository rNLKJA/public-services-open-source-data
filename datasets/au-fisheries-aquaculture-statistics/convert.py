#!/usr/bin/env python3
"""Convert ABARES' "Australian fisheries and aquaculture statistics: Data to
2024-25" workbook (mirrored in raw/) into tidy long-format CSVs.

The source workbook has one sheet per numbered table (Table 1 .. Table 19).
Tables 5-11 each report one state/territory's fisheries and aquaculture
production, in the same layout: a "Value" section then a "Quantity" section,
each split into a "Wild-caught" half (grouped under commodity categories
such as Crustaceans/Molluscs/Finfish, each ending in a category "Total" row)
and an "Aquaculture" half (same shape), with a "Total wild-caught" and
"Total production" row closing out each section. Table 1 reports the same
wild-catch/aquaculture split as state-level and Commonwealth-fishery totals
only (no per-commodity detail).

This script only unpivots each sheet's wide year columns into long rows and
labels each row from the sheet's own section/category headers - it does not
recalculate, re-derive or reinterpret any figure. Trailing single-letter
footnote markers (e.g. "Aquaculture b", "Marron c") are stripped from labels
for readability; the footnotes themselves are quoted in this dataset's
README where they materially affect interpretation. Cells published as "na"
are kept as the literal string "na", not treated as zero or blank.

Tables 2-4 (national commodity-group aggregates), 12 (Commonwealth-managed
fisheries, a structurally distinct by-fishery layout) and 13-19 (trade,
consumption and employment - Australia-wide only, no jurisdiction
breakdown) are left unconverted; they remain in raw/ only. See table-index.csv.
"""
import csv
import re
from collections import Counter
from pathlib import Path

import openpyxl

RAW = Path(__file__).parent / "raw"
OUT = Path(__file__).parent / "data"
SRC = RAW / "afas-statistical-tables-data-to-2024-25-v1.0.0.xlsx"

STATE_TABLES = {
    "Table 5": "New South Wales",
    "Table 6": "Victoria",
    "Table 7": "Queensland",
    "Table 8": "South Australia",
    "Table 9": "Western Australia",
    "Table 10": "Tasmania",
    "Table 11": "Northern Territory",
}

FOOTNOTE_RE = re.compile(r"\s+[a-z]$")


def clean_label(s):
    return FOOTNOTE_RE.sub("", s.strip()).strip()


def find_header_row(rows):
    for i, row in enumerate(rows):
        if row[1] and isinstance(row[1], str) and row[1].strip() == "Commodity":
            return i
    raise RuntimeError("header row not found")


def get_years(rows, header_idx):
    years = []
    for v in rows[header_idx][3:]:
        if v is None or v == "":
            break
        years.append(str(v).strip())
    return years


def is_all_none(vals):
    return all(v is None or v == "" for v in vals)


def convert_state_tables(wb):
    records = []
    for tbl, jurisdiction in STATE_TABLES.items():
        rows = list(wb[tbl].iter_rows(values_only=True))
        header_idx = find_header_row(rows)
        years = get_years(rows, header_idx)
        n_years = len(years)

        measure = None
        production_type = None
        category = None

        for row in rows[header_idx + 1 :]:
            label_raw = row[1]
            unit = row[2]
            data = row[3 : 3 + n_years]

            if label_raw is None or (isinstance(label_raw, str) and label_raw.strip() == ""):
                continue
            label = label_raw.strip() if isinstance(label_raw, str) else str(label_raw)
            if len(label) > 120:
                break  # source-note paragraph at the foot of the sheet

            if label in ("Value", "Quantity"):
                measure, production_type, category = label, "Wild-caught", None
                continue

            has_unit = unit is not None and str(unit).strip() != ""
            if not has_unit and is_all_none(data):
                clean = clean_label(label)
                if clean.startswith("Aquaculture"):
                    production_type, category = "Aquaculture", None
                else:
                    category = clean
                continue

            commodity = clean_label(label)
            unit_str = str(unit).strip() if unit is not None else None
            row_production_type, row_category = production_type, category
            if commodity.startswith("Total production"):
                row_production_type, row_category = "All", None
            elif commodity.startswith("Total wild-caught"):
                row_production_type, row_category = "Wild-caught", None

            for yi, yr in enumerate(years):
                records.append(
                    {
                        "jurisdiction": jurisdiction,
                        "measure": measure,
                        "production_type": row_production_type,
                        "category": row_category or "",
                        "commodity": commodity,
                        "unit": unit_str,
                        "financial_year": yr.replace("–", "-").rstrip("p"),
                        "is_preliminary": yr.endswith("p"),
                        "amount": data[yi] if yi < len(data) else None,
                    }
                )

            if commodity == "Total":
                category = None
            if commodity.startswith("Total production"):
                production_type, category = "Wild-caught", None
    return records


def convert_gross_value_table(wb):
    rows = list(wb["Table 1"].iter_rows(values_only=True))
    header_idx = find_header_row(rows)
    years = get_years(rows, header_idx)
    n_years = len(years)

    records = []
    production_type = None
    for row in rows[header_idx + 1 :]:
        label_raw = row[1]
        unit = row[2]
        data = row[3 : 3 + n_years]
        if label_raw is None or (isinstance(label_raw, str) and label_raw.strip() == ""):
            continue
        label = label_raw.strip() if isinstance(label_raw, str) else str(label_raw)
        if len(label) > 120:
            break
        has_unit = unit is not None and str(unit).strip() != ""
        if not has_unit and is_all_none(data):
            production_type = clean_label(label)
            continue
        jurisdiction = clean_label(label)
        unit_str = str(unit).strip() if unit is not None else None
        for yi, yr in enumerate(years):
            records.append(
                {
                    "production_type": production_type,
                    "jurisdiction": jurisdiction,
                    "unit": unit_str,
                    "financial_year": yr.replace("–", "-").rstrip("p"),
                    "is_preliminary": yr.endswith("p"),
                    "value_aud_thousands": data[yi] if yi < len(data) else None,
                }
            )
    return records


def write_csv(path, fieldnames, records):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(records)
    print(f"wrote {path} ({len(records)} rows)")


def main():
    OUT.mkdir(exist_ok=True)
    wb = openpyxl.load_workbook(SRC, data_only=True)

    state_records = convert_state_tables(wb)
    print("jurisdictions:", Counter(r["jurisdiction"] for r in state_records))
    write_csv(
        OUT / "fisheries-aquaculture-production-by-state-1998-99-to-2024-25.csv",
        ["jurisdiction", "measure", "production_type", "category", "commodity", "unit", "financial_year", "is_preliminary", "amount"],
        state_records,
    )

    sa_only = [r for r in state_records if r["jurisdiction"] == "South Australia"]
    write_csv(
        OUT / "south-australia-fisheries-aquaculture-production-1998-99-to-2024-25.csv",
        ["jurisdiction", "measure", "production_type", "category", "commodity", "unit", "financial_year", "is_preliminary", "amount"],
        sa_only,
    )

    gross_value_records = convert_gross_value_table(wb)
    write_csv(
        OUT / "gross-value-of-production-by-jurisdiction-1998-99-to-2024-25.csv",
        ["production_type", "jurisdiction", "unit", "financial_year", "is_preliminary", "value_aud_thousands"],
        gross_value_records,
    )


if __name__ == "__main__":
    main()
