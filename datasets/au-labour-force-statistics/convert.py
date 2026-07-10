"""
Convert ABS Labour Force, Australia (May 2026 release) raw time-series
workbooks into tidy CSVs focused on the South Australia breakdown.

Source tables (raw/, all CC BY 4.0, ABS catalogue 6202.0):
  62020005  Table 005. Labour force status by Sex, South Australia -
            Trend, Seasonally adjusted and Original
  62020010  Table 010. Labour force status by Sex, State and Territory -
            Trend, Seasonally adjusted and Original

Each is a standard ABS "Time Series Workbook": an Index sheet describing
every series (one row per series, format
"<measure> ;  <sex> ;  [<state> ;]"), and one or more Data sheets where
row 1 (0-indexed row 0) repeats that description per column, row 3
(0-indexed row 2) holds the Series Type (Trend / Seasonally Adjusted /
Original), row 10 (0-indexed row 9) holds the Series ID, and data rows
below (0-indexed row 10 onward) are (date, value) pairs.

Table 005 is a South-Australia-only workbook, so its description strings
have no state/territory segment (everything in it already is SA). Table
010 is the combined all-states-in-one-workbook version, whose description
strings carry a third segment naming the state/territory (or "Australia"
for the national figure) -- used here only to build the small headline
unemployment-rate comparison file.
"""
import csv
from pathlib import Path

import openpyxl

RAW = Path(__file__).parent / "raw"
OUT = Path(__file__).parent / "data"
OUT.mkdir(exist_ok=True)

# The six headline labour-force measures published for South Australia at
# all three series types (Trend / Seasonally Adjusted / Original). Finer
# sub-items (e.g. "Unemployed looked for full-time work") exist in the raw
# workbook only as Original series and are out of scope for this dataset.
CORE_MEASURES = {
    "Employed total": "employed_total_000",
    "Unemployed total": "unemployed_total_000",
    "Labour force total": "labour_force_total_000",
    "Unemployment rate": "unemployment_rate_pct",
    "Participation rate": "participation_rate_pct",
    "Employment to population ratio": "employment_population_ratio_pct",
}


def parse_description(desc):
    """Split an ABS 'Data Item Description' string into its parts, stripping
    the '>' indent markers ABS uses to show sub-item/sub-region nesting."""
    return [p.strip().lstrip(">").strip() for p in desc.split(";") if p.strip()]


def read_workbook_series(path):
    """Return a list of (measure, sex, state_or_none, series_type, {date: value})
    tuples for every series in every Data sheet of one ABS Time Series Workbook."""
    wb = openpyxl.load_workbook(path, data_only=True)
    out = []
    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("Data"):
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        desc_row = rows[0]
        type_row = rows[2]
        data_rows = rows[10:]
        for col_idx in range(1, len(desc_row)):
            desc = desc_row[col_idx]
            if not desc or "Commonwealth of Australia" in desc:
                continue
            parts = parse_description(desc)
            if len(parts) == 2:
                measure, sex = parts
                state = None
            elif len(parts) == 3:
                measure, sex, state = parts
            else:
                continue
            series_type = type_row[col_idx]
            series = {}
            for row in data_rows:
                date = row[0]
                value = row[col_idx] if col_idx < len(row) else None
                if date is None or value is None:
                    continue
                series[date.date().isoformat()[:7]] = value
            out.append((measure, sex, state, series_type, series))
    return out


def build_sa_labour_force_table():
    """South Australia's headline labour force measures, all three ABS
    series types (Trend / Seasonally Adjusted / Original), by sex."""
    series_list = read_workbook_series(RAW / "62020005-labour-force-status-by-sex-south-australia.xlsx")

    # key = (date, sex, series_type) -> {output_field: value}
    table = {}
    keys_order = []
    for measure, sex, state, series_type, series in series_list:
        field = CORE_MEASURES.get(measure)
        if field is None:
            continue
        for date, value in series.items():
            key = (date, sex, series_type)
            if key not in table:
                table[key] = {"date": date, "sex": sex, "series_type": series_type}
                keys_order.append(key)
            table[key][field] = value

    fieldnames = ["date", "sex", "series_type"] + list(CORE_MEASURES.values())
    rows = [table[k] for k in sorted(table, key=lambda k: (k[0], k[1], k[2]))]

    out_path = OUT / "sa-labour-force-monthly.csv"
    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in rows:
            writer.writerow({k: r.get(k, "") for k in fieldnames})
    print(f"Wrote {out_path} ({len(rows)} rows)")
    return rows


def build_headline_unemployment_comparison():
    """All states/territories + national headline (Persons) unemployment
    rate, all three series types, for comparing South Australia against
    the rest of the country -- same idea as au-cost-of-living-cpi's
    all-capital-cities-headline-cpi.csv."""
    series_list = read_workbook_series(RAW / "62020010-labour-force-status-by-sex-state-and-territory.xlsx")

    table = {}
    for measure, sex, state, series_type, series in series_list:
        if measure != "Unemployment rate" or sex != "Persons" or state is None:
            continue
        for date, value in series.items():
            key = (date, state)
            if key not in table:
                table[key] = {"date": date, "location": state}
            field = {
                "Trend": "unemployment_rate_trend_pct",
                "Seasonally Adjusted": "unemployment_rate_seasonally_adjusted_pct",
                "Original": "unemployment_rate_original_pct",
            }.get(series_type)
            if field:
                table[key][field] = value

    fieldnames = ["date", "location", "unemployment_rate_trend_pct",
                  "unemployment_rate_seasonally_adjusted_pct", "unemployment_rate_original_pct"]
    rows = [table[k] for k in sorted(table, key=lambda k: (k[0], k[1]))]

    out_path = OUT / "all-states-headline-unemployment-rate.csv"
    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in rows:
            writer.writerow({k: r.get(k, "") for k in fieldnames})
    print(f"Wrote {out_path} ({len(rows)} rows)")
    return rows


if __name__ == "__main__":
    sa_rows = build_sa_labour_force_table()
    headline_rows = build_headline_unemployment_comparison()

    # Spot checks against the raw workbooks (values sourced directly from the
    # Index/Data sheets above, not recomputed) before trusting the output.
    sa_may2026_sa_persons = [r for r in sa_rows
                              if r["date"] == "2026-05" and r["sex"] == "Persons"
                              and r["series_type"] == "Seasonally Adjusted"][0]
    assert round(sa_may2026_sa_persons["unemployment_rate_pct"], 2) == 4.25, sa_may2026_sa_persons
    assert sa_may2026_sa_persons["employed_total_000"] > 0

    headline_sa_may2026 = [r for r in headline_rows
                             if r["location"] == "South Australia" and r["date"] == "2026-05"][0]
    assert round(headline_sa_may2026["unemployment_rate_seasonally_adjusted_pct"], 2) == round(
        sa_may2026_sa_persons["unemployment_rate_pct"], 2)

    print("Spot checks passed.")
