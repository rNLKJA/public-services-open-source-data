"""
Flatten the Department of Home Affairs' Settlement Database export
(raw/settlement-data-reports-jan-2016-to-dec-2025-last-10-cy-by-migration-streams.xlsx)
into one tidy long-format CSV under data/.

Each of the workbook's three stream sheets (Humanitarian, Family, Skilled) stacks
13 independent two-way breakdown tables one after another (Visa Subclass, Country
of Birth, Ethnicity, Religion, Language, Gender, Year of Settlement, Financial Year
Settlement, Age, Age Band, Marital Status, English Proficiency, Local Government
Area), each cross-tabulated against the 8 states/territories plus "Not Recorded".
This script detects each table by its "Current State/Territory" marker row, reads
its header for the exact state/territory column order (which differs sheet to
sheet), and emits one row per (stream, dimension, category, state) - no figure is
recalculated, and the source's own "Grand Total" category row is preserved as-is
(useful because it includes small-cell counts the source masks as "<5" in the
individual category rows). A workbook-provided "row total" column that only
appears on Grand Total rows is dropped, since it is trivially the sum of the
per-state values already carried over in the same row.

Visa Subclass is the only dimension published as a bare numeric code rather than
a readable label; a `category_description` column is joined in from the workbook's
own "Visa Subclases by Streams" lookup sheet (keyed by stream + code, since one
code - 105 - means different things in the Family vs Skilled streams).
"""
import csv
from pathlib import Path

import openpyxl

RAW = Path(__file__).parent / "raw" / "settlement-data-reports-jan-2016-to-dec-2025-last-10-cy-by-migration-streams.xlsx"
OUT = Path(__file__).parent / "data" / "au-migrant-settlement-statistics.csv"

STREAMS = ["Humanitarian", "Family", "Skilled"]


def build_visa_subclass_lookup(wb):
    ws = wb["Visa Subclases by Streams"]
    rows = list(ws.iter_rows(values_only=True))
    lookup = {}
    stream = None
    for row in rows[1:]:
        if row[0]:
            stream = row[0]
        subclass, description = row[2], row[3]
        if subclass is None:
            continue
        try:
            code = int(subclass)
        except (ValueError, TypeError):
            continue  # a handful of footnote/historical rows use non-numeric codes (e.g. "R")
        lookup[(stream, code)] = description
    return lookup


def is_blank_row(row):
    return all(cell is None or cell == "" for cell in row)


def parse_stream_sheet(ws, stream, visa_lookup):
    rows = list(ws.iter_rows(values_only=True))
    records = []
    i = 0
    while i < len(rows):
        row = rows[i]
        if row[0] == " " and row[1] == "Current State/Territory":
            header = rows[i + 1]
            dimension = header[0]
            states = []
            col = 1
            while col < len(header) and header[col] not in (None, ""):
                states.append(header[col])
                col += 1

            j = i + 2
            while j < len(rows):
                data_row = rows[j]
                if is_blank_row(data_row):
                    break
                if data_row[0] == " " and len(data_row) > 1 and data_row[1] == "Current State/Territory":
                    break
                category = data_row[0]
                if category not in (None, ""):
                    category_description = ""
                    if dimension == "Visa Subclass" and category != "Grand Total":
                        try:
                            category_description = visa_lookup.get((stream, int(category)), "")
                        except (ValueError, TypeError):
                            pass
                    for k, state in enumerate(states):
                        value = data_row[1 + k] if (1 + k) < len(data_row) else None
                        if value is None:
                            continue
                        records.append([stream, dimension, str(category), category_description, state, value])
                j += 1
            i = j
        else:
            i += 1
    return records


def main():
    wb = openpyxl.load_workbook(RAW, data_only=True)
    visa_lookup = build_visa_subclass_lookup(wb)

    out_rows = []
    for stream in STREAMS:
        out_rows.extend(parse_stream_sheet(wb[stream], stream, visa_lookup))

    OUT.parent.mkdir(exist_ok=True)
    with OUT.open("w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([
            "migration_stream", "breakdown_dimension", "category",
            "category_description", "state_territory", "settler_count",
        ])
        writer.writerows(out_rows)

    print(f"Wrote {len(out_rows)} rows to {OUT}")


if __name__ == "__main__":
    main()
