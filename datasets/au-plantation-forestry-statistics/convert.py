#!/usr/bin/env python3
"""Convert ABARES' "Australian forest and wood products statistics: Production
and trade to 2024-25 - Dashboard data tables - Plantation area and log
production" workbook (mirrored in raw/) into tidy CSVs.

The source workbook already publishes three of its four sheets in long
(tidy) format - one row per observation - so this script does not unpivot
anything. It only:
  - standardises each sheet's own column names into one shared schema
  - normalises the FY_Year label ("1996-97") and adds an ISO date column
  - concatenates all three measures (plantation area, log production
    volume, log production value) into one tidy table with a `measure`
    column identifying which slice each row came from, since a user
    comparing area against production would otherwise have to open three
    separate sheets/files and already know they belong together
  - writes a South-Australia-filtered companion of the same merged table
  - collapses the source's inconsistent internal spacing in the volume
    sheet's unit label ("'000  m3", double space) to match the other two
    sheets' single-space style - cosmetic only, not a figure

No figures are recalculated, re-derived or reinterpreted.
"""
import csv
import re
from pathlib import Path

import openpyxl

RAW = Path(__file__).parent / "raw"
OUT = Path(__file__).parent / "data"
SRC = RAW / "afwps-plantation-area-and-log-production-2024-25-v1.0.0.xlsx"

FIELDNAMES = [
    "measure",
    "forest_type",
    "log_type",
    "state",
    "financial_year",
    "date",
    "units",
    "value",
]


def clean_units(u):
    return re.sub(r"\s+", " ", u.strip()) if isinstance(u, str) else u


def convert_plantations(wb):
    ws = wb["Plantations"]
    header = [c.value for c in ws[1]]
    idx = {name: header.index(name) for name in ("FY_Year", "Category_1", "State", "Units", "Value", "Date")}
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        records.append(
            {
                "measure": "Plantation area",
                "forest_type": row[idx["Category_1"]],
                "log_type": "",
                "state": row[idx["State"]],
                "financial_year": row[idx["FY_Year"]],
                "date": row[idx["Date"]].date().isoformat(),
                "units": clean_units(row[idx["Units"]]),
                "value": row[idx["Value"]],
            }
        )
    return records


def convert_production(wb, sheet_name, measure_label):
    ws = wb[sheet_name]
    header = [c.value for c in ws[1]]
    idx = {
        name: header.index(name)
        for name in ("FY_Year", "Category_2", "Category_3", "State", "Units", "Value", "Date")
    }
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        records.append(
            {
                "measure": measure_label,
                "forest_type": row[idx["Category_2"]],
                "log_type": row[idx["Category_3"]] or "",
                "state": row[idx["State"]],
                "financial_year": row[idx["FY_Year"]],
                "date": row[idx["Date"]].date().isoformat(),
                "units": clean_units(row[idx["Units"]]),
                "value": row[idx["Value"]],
            }
        )
    return records


def write_csv(path, records):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=FIELDNAMES)
        w.writeheader()
        w.writerows(records)
    print(f"wrote {path} ({len(records)} rows)")


def main():
    OUT.mkdir(exist_ok=True)
    wb = openpyxl.load_workbook(SRC, data_only=True)

    records = []
    records += convert_plantations(wb)
    records += convert_production(wb, "Volume of production", "Volume of log production")
    records += convert_production(wb, "Value of production", "Value of log production")

    write_csv(OUT / "plantation-area-and-log-production-by-state-1938-39-to-2024-25.csv", records)

    sa_records = [r for r in records if r["state"] == "South Australia"]
    write_csv(
        OUT / "south-australia-plantation-area-and-log-production-1938-39-to-2024-25.csv",
        sa_records,
    )


if __name__ == "__main__":
    main()
