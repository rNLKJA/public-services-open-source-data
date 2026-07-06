#!/usr/bin/env python3
"""
Converts the ABS "Prisoners in Australia, 2025 — Prisoner characteristics,
States and territories (Tables 15-35)" workbook into tidy, directly-loadable
CSVs, without altering, recalculating or reinterpreting any figures.

Source file (mirrored verbatim in raw/):
  prisoner-characteristics-states-territories-tables-15-35.xlsx

Approach:
  Each of the 21 data tables (Table 15 .. Table 35) in the source workbook has
  its own row/column layout (some are State x characteristic grids for a single
  2025 reference date, some are State x year time series, some nest two or
  three row-group levels e.g. Indigenous status > Sex > Legal status). To avoid
  forcing 21 structurally different tables into one artificial schema (which
  would require reinterpreting the source), each table is converted to its own
  long-format CSV:

    table, table_title, row_group_1, row_group_2, row_label, column, value

  - `table` / `table_title` identify which of the 21 ABS tables the row came
    from (the "slice" column requested for multi-file/multi-table merges).
  - `row_group_1` / `row_group_2` are the section headers the source itself
    prints above a block of data rows (e.g. "Aboriginal and Torres Strait
    Islander" > "Males"), forward-filled onto every row underneath them so the
    grouping survives outside the spreadsheet's visual indentation. Left blank
    where a table has no such grouping.
  - `row_label` is the table's own left-most row label (e.g. "SA", "2019",
    "Under 1 month", a specific offence description).
  - `column` is the table's own column header, reconstructed by joining its
    (up to 3) header rows (e.g. "SA", "Aust. - Proportion (%)", "Mean (years)").
  - `value` is the cell value, copied verbatim (numbers stay numbers; ABS
    footnote markers like "(a)" stay attached to labels/headers, not to values).

  A second file, all-tables-long.csv, is the union of all 21 per-table CSVs
  stacked together (one row per data cell across every table), for anyone who
  wants to filter/pivot across the whole release in one file (e.g.
  `df[df.row_label == "SA"]` or `df[df.column == "SA"]` depending on table
  orientation).

No totals are recomputed, no percentages derived, no cell values changed —
this script only reshapes layout (unpivots wide ABS tables to long rows) and
forward-fills the section-header labels the source already prints.
"""
import csv
import re
from pathlib import Path

import openpyxl

RAW = Path(__file__).parent / "raw" / "prisoner-characteristics-states-territories-tables-15-35.xlsx"
OUT_DIR = Path(__file__).parent / "data"
OUT_DIR.mkdir(parents=True, exist_ok=True)

FOOTNOTE_RE = re.compile(r"\s*\([a-z]\)")  # strips trailing " (a)", " (b)(c)" etc. from titles only


def trim_row(row):
    vals = list(row)
    while vals and vals[-1] is None:
        vals.pop()
    return vals


def build_columns(header_rows, n_cols):
    """Join up to 3 stacked header rows into one column label per column index,
    forward-filling merged-cell blanks left-to-right (openpyxl returns None for
    the non-top-left cells of a merged range when data_only=True).

    Forward-fill in a subordinate header row is bounded by the group
    boundaries the top-level header row (header_rows[0]) already defines: a
    merged top-level header (e.g. "Aggregate sentence length" spanning several
    columns) may have several distinct sub-headers underneath it (e.g. "Under
    1 year", "Median"), but a subordinate value must never leak across into the
    next top-level group's columns (e.g. "Median" under "Aggregate sentence
    length" must not carry into the next, separate top-level column
    "Sentenced in last 12 months")."""
    # First, forward-fill the top-level row on its own (defines group spans).
    top = []
    last = None
    for i in range(n_cols):
        v = header_rows[0][i] if i < len(header_rows[0]) else None
        if v is not None and str(v).strip() != "":
            last = str(v).strip()
        top.append(last if i > 0 else (str(v).strip() if v else None))

    filled_rows = [top]
    for hr in header_rows[1:]:
        filled = []
        last = None
        last_group = None
        for i in range(n_cols):
            v = hr[i] if i < len(hr) else None
            group = top[i]
            if group != last_group:
                # entered a new top-level group; forward-fill must not cross
                # this boundary regardless of whether this row has a value here
                last = None
                last_group = group
            if v is not None and str(v).strip() != "":
                last = str(v).strip()
                filled.append(last)
            else:
                filled.append(last if i > 0 else None)
        filled_rows.append(filled)

    columns = []
    for i in range(n_cols):
        parts = []
        for fr in filled_rows:
            v = fr[i]
            if v and v not in parts:
                parts.append(v)
        columns.append(" - ".join(parts) if parts else f"col_{i}")
    return columns


def convert_sheet(ws):
    values_rows = [trim_row(r) for r in ws.iter_rows(values_only=True)]
    # rows[0] = ABS auto-generated tab description, rows[1] = "Australian Bureau
    # of Statistics", rows[2] = table title, rows[3] = "Prisoners in Australia, 2025"
    title_raw = values_rows[2][0] if len(values_rows) > 2 and values_rows[2] else ""
    table_title = FOOTNOTE_RE.sub("", title_raw).strip()

    # Determine header block: row 4 (index 4) always present; rows 5 and 6 are
    # continuation header rows only if their first cell is None/blank (a real
    # first data row always has a non-empty row label in column 0).
    header_start = 4
    header_rows = [values_rows[header_start]]
    data_start = header_start + 1
    while data_start < len(values_rows) and (not values_rows[data_start] or values_rows[data_start][0] in (None, "")):
        # A continuation header row still has values in columns >0; a blank
        # section-group row (data row with only col 0 filled) has nothing else.
        candidate = values_rows[data_start]
        if candidate and len(candidate) > 1:
            header_rows.append(candidate)
            data_start += 1
        else:
            break

    n_cols = max((len(r) for r in header_rows), default=0)
    n_cols = max(n_cols, max((len(r) for r in values_rows[data_start:] if r), default=0))
    columns = build_columns(header_rows, n_cols)

    # Row-group headers (e.g. "Aboriginal and Torres Strait Islander", "Males",
    # "New South Wales") are tracked as a stack. This workbook's spreadsheet
    # formatting doesn't carry a reliable per-row indent/outline signal, so
    # depth is inferred from the length of each *consecutive run* of
    # header-only rows (rows with a row label but no data) between data
    # blocks: a run of 2+ header rows in a row establishes a brand new set of
    # ancestor levels from scratch (e.g. "Non-Indigenous" then "Males" —
    # confirmed against the source: this always precedes a state/territory or
    # top-level-category restart), while a run of exactly 1 header row is a
    # sibling replacing only the innermost currently-open level (e.g. "Females"
    # following a "Males" data block, both still under the same outer
    # "Aboriginal and Torres Strait Islander" group). This rule was derived
    # and verified directly against every nested table in this workbook
    # (Tables 20, 21, 29, 30, 33, 34, 35).
    out_rows = []
    stack = []  # list of labels, outermost first
    pending_headers = []  # consecutive header-only labels seen since last data row
    for r in values_rows[data_start:]:
        if not r or r[0] is None or str(r[0]).strip() == "":
            continue
        label_raw = str(r[0]).strip()
        if label_raw.startswith("©") or label_raw.lower().startswith("("):
            continue  # copyright / footnote lines
        has_data = len(r) > 1 and any(v is not None and str(v).strip() != "" for v in r[1:])
        if not has_data:
            pending_headers.append(label_raw)
            continue
        if pending_headers:
            if len(pending_headers) >= 2:
                stack = list(pending_headers)
            else:
                # single header: replace only the innermost open level
                if stack:
                    stack = stack[:-1] + pending_headers
                else:
                    stack = list(pending_headers)
            pending_headers = []
        group1 = stack[0] if len(stack) > 0 else ""
        group2 = " > ".join(stack[1:]) if len(stack) > 1 else ""
        for i in range(1, len(r)):
            v = r[i]
            if v is None or str(v).strip() == "":
                continue
            out_rows.append({
                "row_group_1": group1,
                "row_group_2": group2,
                "row_label": FOOTNOTE_RE.sub("", label_raw).strip(),
                "column": columns[i] if i < len(columns) else f"col_{i}",
                "value": v,
            })
    return table_title, out_rows


def main():
    wb = openpyxl.load_workbook(RAW, data_only=True)
    all_rows = []
    index_rows = []
    for name in wb.sheetnames:
        if name in ("Contents", "Further information"):
            continue
        ws = wb[name]
        table_title, out_rows = convert_sheet(ws)
        table_num = name.replace("Table ", "").strip()
        for row in out_rows:
            row["table"] = name
            row["table_title"] = table_title
        all_rows.extend(out_rows)
        index_rows.append({"table": name, "table_title": table_title, "rows": len(out_rows)})

        out_path = OUT_DIR / f"table-{int(table_num):02d}.csv"
        with out_path.open("w", newline="") as f:
            writer = csv.DictWriter(
                f, fieldnames=["table", "table_title", "row_group_1", "row_group_2", "row_label", "column", "value"]
            )
            writer.writeheader()
            for row in out_rows:
                writer.writerow(row)
        print(f"wrote {out_path.name}: {len(out_rows)} rows")

    combined_path = OUT_DIR / "all-tables-long.csv"
    with combined_path.open("w", newline="") as f:
        writer = csv.DictWriter(
            f, fieldnames=["table", "table_title", "row_group_1", "row_group_2", "row_label", "column", "value"]
        )
        writer.writeheader()
        for row in all_rows:
            writer.writerow(row)
    print(f"wrote {combined_path.name}: {len(all_rows)} rows total")

    index_path = OUT_DIR / "table-index.csv"
    with index_path.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["table", "table_title", "rows"])
        writer.writeheader()
        for row in index_rows:
            writer.writerow(row)
    print(f"wrote {index_path.name}")


if __name__ == "__main__":
    main()
