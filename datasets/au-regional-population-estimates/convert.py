#!/usr/bin/env python3
"""Convert ABS Regional Population, 2024-25 data cubes into tidy CSVs.

Source files (raw/):
  - abs-population-estimates-by-lga-2001-2025.xlsx (Table 1: national ERP time series by LGA)
  - abs-population-components-by-lga-2021-25.xlsx (Table 1: national births/deaths/migration by LGA, 4 financial years)
  - abs-population-estimates-and-components-by-lga-2024-25.xlsx (Table 4: SA-only LGA snapshot with
    components/area/density; Table 8: state/territory comparison)

LGA codes follow the ABS ASGS convention where the leading digit identifies the
state/territory (1 NSW, 2 VIC, 3 QLD, 4 SA, 5 WA, 6 TAS, 7 NT, 8 ACT, 9 Other
Territories) -- confirmed directly against every code in the source file rather
than assumed, and used here to add a `state_name` column so a South Australia
extract doesn't require a separate lookup.
"""
import csv
import openpyxl

RAW = "raw"
DATA = "data"

STATE_PREFIX = {
    "1": "New South Wales",
    "2": "Victoria",
    "3": "Queensland",
    "4": "South Australia",
    "5": "Western Australia",
    "6": "Tasmania",
    "7": "Northern Territory",
    "8": "Australian Capital Territory",
    "9": "Other Territories",
}


def write_csv(path, header, rows):
    with open(f"{DATA}/{path}", "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)
    print(f"wrote {path}: {len(rows)} rows")


def lga_erp_2001_2025():
    wb = openpyxl.load_workbook(f"{RAW}/abs-population-estimates-by-lga-2001-2025.xlsx", data_only=True)
    ws = wb["Table 1"]
    years = list(range(2001, 2026))
    rows = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        code, name = row[0], row[1]
        if isinstance(code, int):
            is_total = False
            state_name = STATE_PREFIX.get(str(code)[0], "Unknown")
        elif code is None and name == "Total Australia":
            is_total = True
            state_name = "Australia"
        else:
            continue  # footnotes / blank / copyright rows
        for i, year in enumerate(years):
            erp = row[2 + i]
            rows.append([code, name, state_name, year, erp, is_total])
    write_csv(
        "au-lga-population-estimates-2001-2025.csv",
        ["lga_code", "lga_name", "state_name", "year", "erp", "is_total_row"],
        rows,
    )
    sa_rows = [r for r in rows if r[2] == "South Australia" and not r[5]]
    assert len(sa_rows) == 71 * 25, f"expected 71 SA LGAs x 25 years, got {len(sa_rows)} rows"
    adelaide_2025 = [r for r in rows if r[0] == 40070 and r[3] == 2025]
    assert adelaide_2025 and adelaide_2025[0][4] == 30173, f"Adelaide (C) 2025 spot-check failed: {adelaide_2025}"


def lga_components_2021_2025():
    wb = openpyxl.load_workbook(f"{RAW}/abs-population-components-by-lga-2021-25.xlsx", data_only=True)
    ws = wb["Table 1"]
    fin_years = ["2021-22", "2022-23", "2023-24", "2024-25"]
    rows = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        st_code = row[0]
        if not isinstance(st_code, int):
            continue  # footnotes / blank / copyright rows
        st_name, lga_code, lga_name = row[1], row[2], row[3]
        for bi, fy in enumerate(fin_years):
            base = 4 + bi * 9
            metrics = row[base:base + 9]
            rows.append([st_code, st_name, lga_code, lga_name, fy, *metrics])
    write_csv(
        "au-lga-population-components-2021-2025.csv",
        ["st_code", "st_name", "lga_code", "lga_name", "financial_year", "births", "deaths",
         "natural_increase", "internal_arrivals", "internal_departures", "net_internal_migration",
         "overseas_arrivals", "overseas_departures", "net_overseas_migration"],
        rows,
    )
    albury = [r for r in rows if r[2] == 10050 and r[4] == "2021-22"]
    assert albury and albury[0][5] == 720, f"Albury 2021-22 births spot-check failed: {albury}"


def sa_lga_estimates_and_components_2024_25():
    wb = openpyxl.load_workbook(
        f"{RAW}/abs-population-estimates-and-components-by-lga-2024-25.xlsx", data_only=True
    )
    ws = wb["Table 4"]
    rows = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        lga_code, lga_name = row[0], row[1]
        if lga_name is None:
            continue  # footnotes / blank / copyright rows
        is_total = lga_code is None
        metrics = row[2:11]
        rows.append([lga_code, lga_name, *metrics, is_total])
    write_csv(
        "sa-lga-population-estimates-and-components-2024-25.csv",
        ["lga_code", "lga_name", "erp_2024", "erp_2025", "erp_change_no", "erp_change_pct",
         "natural_increase", "net_internal_migration", "net_overseas_migration",
         "area_km2", "population_density_2025", "is_total_row"],
        rows,
    )
    adelaide = [r for r in rows if r[0] == 40070]
    assert adelaide and adelaide[0][3] == 30173, f"Adelaide 2025 ERP spot-check failed: {adelaide}"
    total = [r for r in rows if r[11]]
    assert total and total[0][3] == 1902665, f"Total South Australia 2025 ERP spot-check failed: {total}"


def states_territories_2024_25():
    wb = openpyxl.load_workbook(
        f"{RAW}/abs-population-estimates-and-components-by-lga-2024-25.xlsx", data_only=True
    )
    ws = wb["Table 8"]
    rows = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        st_code, st_name = row[0], row[1]
        if not isinstance(st_code, int):
            continue
        metrics = row[2:11]
        rows.append([st_code, st_name, *metrics])
    write_csv(
        "au-states-territories-population-components-2024-25.csv",
        ["st_code", "st_name", "erp_2024", "erp_2025", "erp_change_no", "erp_change_pct",
         "natural_increase", "net_internal_migration", "net_overseas_migration",
         "area_km2", "population_density_2025"],
        rows,
    )
    sa = [r for r in rows if r[1] == "South Australia"]
    assert sa and sa[0][3] == 1902665, f"SA 2025 ERP spot-check failed: {sa}"


if __name__ == "__main__":
    lga_erp_2001_2025()
    lga_components_2021_2025()
    sa_lga_estimates_and_components_2024_25()
    states_territories_2024_25()
    print("All spot-checks passed.")
