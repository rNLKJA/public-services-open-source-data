#!/usr/bin/env python3
"""
Converts the Clean Energy Regulator's Small-scale Renewable Energy Scheme
(SRES) postcode-level Installations and Capacity workbooks into tidy,
directly-loadable CSVs, without altering, recalculating or reinterpreting any
figures.

Source files (mirrored verbatim in raw/):
  sres-postcode-data-installations-2011-to-present-and-totals.xlsx
  sres-postcode-data-capacity-2011-to-present-and-totals.xlsx

Each workbook has one sheet per equipment category (e.g. SGU-Solar,
SWH-Solar, SGU-Battery). Each sheet is a wide postcode x period grid: one row
per postcode, one column per historic-total/month/grand-total period.

Column header parsing:
  - "Historic Total ... (2001 - 2010)" columns become period "2001-2010",
    period_type "historic_total".
  - "<Mon> <YYYY> - <metric text>" columns become period "YYYY-MM",
    period_type "monthly".
  - "Total <metric text>" columns (no month) become period "all-time",
    period_type "grand_total" -- this is the source workbook's own
    pre-computed cumulative-to-date figure, not something this script derives.
  - The trailing "<metric text>" (e.g. "Installation Quantity", "Rated Power
    Output in kW", "Usable capacity in kWh") is kept verbatim as `metric_text`
    so the unit is never assumed, only read from the source's own header.

Two kinds of output are written, both directly from source cells with no
recalculation:

  1. *-totals-by-postcode.csv -- national, every postcode in Australia, one
     row per (category, postcode), using ONLY the source's own "Total"
     (all-time, all-time cumulative) column. Small (source publishes ~13,000
     postcode/category combinations for installations, ~9,500 for capacity)
     because it carries one cumulative figure per postcode-category rather
     than a 15-year monthly series.
  2. sa-*-long.csv -- the full historic-total/monthly/grand-total time series
     (same rows the national series would have), filtered to South Australian
     postcodes (Australia Post ranges 5000-5799 and 5800-5999) only.

The full-Australia monthly time series (every postcode x every month back to
2011 x all 6 categories) is not written to a committed file here -- melted out
long-form it is ~700,000 / ~380,000 rows (tens of MB), far larger than any
other file in this repository. It doesn't need a network fetch to reproduce:
re-run this script with WRITE_NATIONAL_LONG = True below and it will melt the
same already-mirrored raw/ workbooks in full for any postcode, not just SA.

Rows where the source value is 0 are dropped from the monthly/historic-total
files (not from the *-totals-by-postcode.csv grand-total rows, which are kept
regardless of value) -- an omitted (postcode, category, month) combination
means "0, per the source workbook", exactly as an explicit 0 cell would.
"""
import csv
import re
from pathlib import Path

import openpyxl

RAW_DIR = Path(__file__).parent / "raw"
OUT_DIR = Path(__file__).parent / "data"
OUT_DIR.mkdir(parents=True, exist_ok=True)

WRITE_NATIONAL_LONG = False  # set True to also reproduce the full-Australia monthly series locally

CATEGORY_LABELS = {
    "SGU-Solar": "Small Generation Unit - Solar Panel",
    "SGU-Wind": "Small Generation Unit - Wind",
    "SGU-Hydro": "Small Generation Unit - Hydro",
    "SGU-Battery": "Small Generation Unit - Battery",
    "SWH-Air Source Heat Pump": "Solar Water Heater - Air Source Heat Pump",
    "SWH-Solar": "Solar Water Heater - Solar",
}

MONTHS = {m: i + 1 for i, m in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
)}
MONTH_RE = re.compile(r"^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(\d{4})\s*-\s*(.+)$")
HISTORIC_RE = re.compile(r"^Historic Total (.+?)\s*\(2001\s*-\s*2010\)$")
TOTAL_RE = re.compile(r"^Total\s+(.+)$")

FIELDNAMES = ["category_code", "category", "postcode", "period", "period_type", "metric_text", "value"]


def trim_row(row):
    vals = list(row)
    while vals and vals[-1] is None:
        vals.pop()
    return vals


def find_header_row(rows):
    for i, r in enumerate(rows[:6]):
        if r and r[0] is not None and "postcode" in str(r[0]).lower():
            return i
    raise ValueError("no header row found in first 6 rows")


def parse_column(raw_header):
    """Returns (period, period_type, metric_text) for one data column header."""
    h = re.sub(r"\s+", " ", str(raw_header).replace("\xa0", " ")).strip()
    m = HISTORIC_RE.match(h)
    if m:
        return "2001-2010", "historic_total", m.group(1).strip()
    m = MONTH_RE.match(h)
    if m:
        month, year, metric = m.groups()
        return f"{year}-{MONTHS[month]:02d}", "monthly", metric.strip()
    m = TOTAL_RE.match(h)
    if m:
        return "all-time", "grand_total", m.group(1).strip()
    raise ValueError(f"unrecognised column header: {raw_header!r}")


def is_sa_postcode(pc):
    if not pc.isdigit():
        return False
    n = int(pc)
    return 5000 <= n <= 5999


def melt_workbook(path):
    """Yields one dict per non-empty data cell across every sheet in the workbook."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = [trim_row(r) for r in ws.iter_rows(values_only=True)]
        header_idx = find_header_row(rows)
        header = rows[header_idx]
        data_rows = [r for r in rows[header_idx + 1:] if r and r[0] is not None]
        parsed_cols = [None] + [parse_column(header[i]) for i in range(1, len(header))]
        category = CATEGORY_LABELS.get(sheet_name, sheet_name)

        for r in data_rows:
            postcode = str(r[0]).strip()
            for i in range(1, len(header)):
                if i >= len(r) or r[i] is None or str(r[i]).strip() == "":
                    continue
                value = r[i]
                period, period_type, metric_text = parsed_cols[i]
                yield {
                    "category_code": sheet_name,
                    "category": category,
                    "postcode": postcode,
                    "period": period,
                    "period_type": period_type,
                    "metric_text": metric_text,
                    "value": value,
                }
        print(f"  {sheet_name}: {len(data_rows)} postcodes")


def convert_workbook(path, out_name):
    print(f"{path.name}:")
    totals_rows = []
    sa_rows = []
    national_long_rows = [] if WRITE_NATIONAL_LONG else None

    for row in melt_workbook(path):
        if row["period_type"] == "grand_total":
            totals_rows.append(row)
        elif row["value"] == 0:
            continue  # omitted zero cell == "0, per the source" (see module docstring)
        if is_sa_postcode(row["postcode"]):
            sa_rows.append(row)
        if WRITE_NATIONAL_LONG:
            national_long_rows.append(row)

    totals_path = OUT_DIR / f"{out_name}-totals-by-postcode.csv"
    with totals_path.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        writer.writeheader()
        writer.writerows(totals_rows)
    print(f"wrote {totals_path.name}: {len(totals_rows)} rows (national, all-time totals only)")

    sa_path = OUT_DIR / f"sa-{out_name}-long.csv"
    with sa_path.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        writer.writeheader()
        writer.writerows(sa_rows)
    print(f"wrote {sa_path.name}: {len(sa_rows)} rows (SA postcodes 5000-5999, full time series)")

    if WRITE_NATIONAL_LONG:
        national_path = OUT_DIR / f"{out_name}-long-national.csv"
        with national_path.open("w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
            writer.writeheader()
            writer.writerows(national_long_rows)
        print(f"wrote {national_path.name}: {len(national_long_rows)} rows (national, full time series -- NOT committed, see docstring)")


def main():
    convert_workbook(RAW_DIR / "sres-postcode-data-installations-2011-to-present-and-totals.xlsx", "installations")
    convert_workbook(RAW_DIR / "sres-postcode-data-capacity-2011-to-present-and-totals.xlsx", "capacity")


if __name__ == "__main__":
    main()
