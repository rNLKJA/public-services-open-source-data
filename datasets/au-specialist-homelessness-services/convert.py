"""Reshape AIHW's Specialist Homelessness Services Collection historical
tables workbook (16 HIST.* sheets, 2011-12 to 2024-25) into tidy CSVs.

Every sheet shares the same 4-row preamble (a "Table of contents" back-link
cell, the table's own title, a blank row, then the real header) followed by
data rows and a trailing footnote block. Leading columns are dimension
columns (State/territory, Data type, Sex, Indigenous status, Remoteness
area, Client characteristics — the exact set varies by table); the columns
after them are financial-year value columns plus one "Average annual
change" summary column. This script locates that header generically per
sheet (no fixed row/column layout assumed across the differently-shaped
tables), keeps each table's own native wide shape in data/tables/, and also
melts every table into one combined long file for cross-table filtering —
the same two-tier approach already used for au-youth-justice-supervision-statistics
in this repository.
"""
import csv
import re
from pathlib import Path

import openpyxl

RAW = Path(__file__).parent / "raw"
OUT = Path(__file__).parent / "data"
(OUT / "tables").mkdir(parents=True, exist_ok=True)

SOURCE = RAW / "aihw-hou-343-specialist-homelessness-services-historical-tables-2011-12-to-2024-25.xlsx"

YEAR_RE = re.compile(r"^(\d{4})[–-](\d{2})(\(\w+\))?$")


def snake_case(label):
    label = str(label).strip()
    label = re.sub(r"\s+", " ", label)
    label = label.replace("%", "pct")
    label = re.sub(r"[^0-9a-zA-Z]+", "_", label).strip("_").lower()
    return label


def year_label(header_cell):
    m = YEAR_RE.match(str(header_cell).strip())
    if not m:
        return None
    return f"{m.group(1)}-{m.group(2)}"


def is_change_col(header_cell):
    return "change" in str(header_cell).lower()


wb = openpyxl.load_workbook(SOURCE, data_only=True)

table_titles = {}
ws = wb["Contents"]
for row in ws.iter_rows(values_only=True):
    if row[0] and str(row[0]).startswith("Table HIST."):
        code = str(row[0]).split(":")[0].replace("Table ", "").strip()
        title = str(row[0]).split(":", 1)[1].strip()
        table_titles[code] = title

index_rows = []
long_rows = []

for sheet_name in wb.sheetnames:
    if sheet_name in ("Contents", "Explanatory notes"):
        continue

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[3]

    dim_idx, year_idx, change_idx = [], [], []
    for i, cell in enumerate(header):
        if cell is None:
            continue
        if year_label(cell) is not None:
            year_idx.append(i)
        elif is_change_col(cell):
            change_idx.append(i)
        else:
            dim_idx.append(i)

    dim_names = [snake_case(header[i]) for i in dim_idx]

    data_rows = []
    notes_lines = []
    seen_blank = False
    for r in rows[4:]:
        if not seen_blank and all(c is None for c in r):
            seen_blank = True
            continue
        if seen_blank:
            if r[0]:
                notes_lines.append(str(r[0]))
            continue
        data_rows.append(r)

    # --- native wide-shape per-table CSV ---
    wide_fieldnames = dim_names + [f"fy_{year_label(header[i])}" for i in year_idx] + [
        f"avg_annual_change_pct__{snake_case(header[i])}" for i in change_idx
    ]
    wide_path = OUT / "tables" / f"{sheet_name}.csv"
    with wide_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(wide_fieldnames)
        for r in data_rows:
            writer.writerow([r[i] for i in dim_idx] + [r[i] for i in year_idx] + [r[i] for i in change_idx])

    # --- long-format rows for the combined file ---
    for r in data_rows:
        dims_str = "; ".join(f"{header[i]}={r[i]}" for i in dim_idx if r[i] is not None)
        for i in year_idx:
            long_rows.append({
                "table_code": sheet_name,
                "table_title": table_titles.get(sheet_name, ""),
                "dimensions": dims_str,
                "financial_year": year_label(header[i]),
                "value": r[i],
            })
        for i in change_idx:
            long_rows.append({
                "table_code": sheet_name,
                "table_title": table_titles.get(sheet_name, ""),
                "dimensions": dims_str,
                "financial_year": f"avg_annual_change_pct",
                "value": r[i],
            })

    index_rows.append({
        "table_code": sheet_name,
        "table_title": table_titles.get(sheet_name, ""),
        "n_data_rows": len(data_rows),
        "dimension_columns": "; ".join(dim_names),
        "financial_years_covered": ", ".join(year_label(header[i]) for i in year_idx),
        "has_south_australia": any(
            r[dim_idx[0]] == "South Australia" for r in data_rows
        ) if dim_idx else False,
        "notes": " | ".join(notes_lines),
    })

with (OUT / "table-index.csv").open("w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=list(index_rows[0].keys()))
    writer.writeheader()
    writer.writerows(index_rows)

with (OUT / "all-tables-long.csv").open("w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=["table_code", "table_title", "dimensions", "financial_year", "value"])
    writer.writeheader()
    writer.writerows(long_rows)

print(f"Wrote {len(index_rows)} per-table CSVs to data/tables/, table-index.csv ({len(index_rows)} rows), "
      f"all-tables-long.csv ({len(long_rows)} rows)")
