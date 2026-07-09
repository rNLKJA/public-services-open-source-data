"""Convert DCCEEW's State and Territory Greenhouse Gas Inventories 2024 workbook
(raw/state-territory-inventories-2024-emission-data-tables.xlsx) into tidy long-format CSVs.

The source has one sheet per jurisdiction (plus a national "Australia" sheet), each a wide
table: IPCC category rows x financial-year columns (1989-90 to 2023-24), plus a trailing
"Change from 2004-05 to latest reported year (%)" summary column. This script melts each
sheet into long format and concatenates all jurisdictions into one table, decodes the
UNFCCC/IPCC notation keys (NO/NE/NA/IE/C) used in place of a number, and writes a second
table for the change-% summary column since it is a different kind of observation (a
period-to-period comparison, not an annual value) and would misrepresent the annual series
if melted alongside it.

Requires: openpyxl
"""

import csv
import openpyxl

SOURCE = "raw/state-territory-inventories-2024-emission-data-tables.xlsx"
SHEETS = ["Australia", "NSW", "Vic", "Qld", "SA", "WA", "Tas", "NT", "ACT", "External Territories"]
FIRST_YEAR_COL = 2   # column B = 1989-90
LAST_YEAR_COL = 36   # column AJ = 2023-24
CHANGE_COL = 37      # column AK = change 2004-05 to latest (%)
FIRST_DATA_ROW = 8
LAST_DATA_ROW = 66

NOTATION_MEANINGS = {
    "NO": "Not occurring",
    "NE": "Not estimated",
    "NA": "Not applicable",
    "IE": "Included elsewhere",
    "C": "Confidential (aggregated to avoid disclosing confidential information)",
}


def split_value(raw):
    """Return (numeric_value_or_blank, notation_key, notation_meaning) for a source cell."""
    if raw is None:
        return "", "", ""
    if isinstance(raw, (int, float)):
        return raw, "", ""
    code = str(raw).strip()
    return "", code, NOTATION_MEANINGS.get(code, "")


def financial_year_start(fy):
    # "1989-90" -> 1989
    return int(fy.split("-")[0])


def main():
    wb = openpyxl.load_workbook(SOURCE, data_only=True)

    annual_rows = []
    change_rows = []

    for jurisdiction in SHEETS:
        ws = wb[jurisdiction]
        year_headers = [ws.cell(row=7, column=c).value for c in range(FIRST_YEAR_COL, LAST_YEAR_COL + 1)]

        for r in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
            category = ws.cell(row=r, column=1).value
            if category is None:
                continue
            category = category.strip()

            for c, fy in zip(range(FIRST_YEAR_COL, LAST_YEAR_COL + 1), year_headers):
                raw = ws.cell(row=r, column=c).value
                value, notation_key, notation_meaning = split_value(raw)
                annual_rows.append({
                    "jurisdiction": jurisdiction,
                    "category": category,
                    "financial_year": fy,
                    "financial_year_start": financial_year_start(fy),
                    "emissions_gg_co2e": value,
                    "notation_key": notation_key,
                    "notation_meaning": notation_meaning,
                })

            raw_change = ws.cell(row=r, column=CHANGE_COL).value
            change_value, change_key, change_meaning = split_value(raw_change)
            change_rows.append({
                "jurisdiction": jurisdiction,
                "category": category,
                "change_2004_05_to_2023_24_pct": change_value,
                "notation_key": change_key,
                "notation_meaning": change_meaning,
            })

    annual_fields = [
        "jurisdiction", "category", "financial_year", "financial_year_start",
        "emissions_gg_co2e", "notation_key", "notation_meaning",
    ]
    with open("data/au-state-territory-ghg-emissions.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=annual_fields)
        w.writeheader()
        w.writerows(annual_rows)

    sa_annual = [row for row in annual_rows if row["jurisdiction"] == "SA"]
    with open("data/au-state-territory-ghg-emissions-sa.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=annual_fields)
        w.writeheader()
        w.writerows(sa_annual)

    change_fields = ["jurisdiction", "category", "change_2004_05_to_2023_24_pct", "notation_key", "notation_meaning"]
    with open("data/au-state-territory-ghg-change-2004-05-to-2023-24.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=change_fields)
        w.writeheader()
        w.writerows(change_rows)

    sa_change = [row for row in change_rows if row["jurisdiction"] == "SA"]
    with open("data/au-state-territory-ghg-change-2004-05-to-2023-24-sa.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=change_fields)
        w.writeheader()
        w.writerows(sa_change)

    print(f"annual rows: {len(annual_rows)} (SA: {len(sa_annual)})")
    print(f"change rows: {len(change_rows)} (SA: {len(sa_change)})")


if __name__ == "__main__":
    main()
