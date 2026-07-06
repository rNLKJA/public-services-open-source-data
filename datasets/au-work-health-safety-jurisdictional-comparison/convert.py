"""
Unpivot Safe Work Australia's "Jurisdictional Comparison detailed data file" (raw/*.xlsx)
into tidy long-format CSVs under data/. No values are recalculated or reinterpreted -
each source table is reshaped from a wide year-by-jurisdiction (and, for some tables,
category-by-jurisdiction) grid into one row per observation.

Source layout per "Table X.Y" sheet: a title row, a metadata row, an indicator-code row,
a header row of financial years, then data rows. Most tables are flat (every data row is
a jurisdiction). Five tables (1.7, 1.11, 1.12, 3.2, 5.2) nest a category (duration bin,
mechanism of incident, or industry) above per-jurisdiction rows; the category row itself
carries the same figure for all jurisdictions combined (labelled here as jurisdiction "AUST"),
exactly as it appears in the source - not a recomputation.
"""
import csv
import re
import openpyxl
from pathlib import Path

RAW = Path(__file__).parent / "raw" / "jurisdictional-comparison-detailed-data-file-2023-24.xlsx"
OUT = Path(__file__).parent / "data"

YEAR_RE = re.compile(r"^20\d{2}-\d{2}p?$")
JURISDICTIONS = {
    "NSW", "VIC", "QLD", "SA", "WA", "TAS", "NT", "ACT", "ACTPrivate",
    "Aus Gov", "Comcare", "Seacare", "AUST", "AUST Total",
}


def find_header_row(rows):
    for i, row in enumerate(rows):
        if any(isinstance(v, str) and YEAR_RE.match(v) for v in row if v is not None):
            return i
    raise ValueError("no year header row found")


def parse_table(ws):
    rows = list(ws.iter_rows(min_row=1, max_row=200, values_only=True))
    title = rows[0][0]
    hdr_idx = find_header_row(rows)
    years = rows[hdr_idx]
    year_cols = [(i, y) for i, y in enumerate(years) if isinstance(y, str) and YEAR_RE.match(y)]

    records = []
    category = None
    for row in rows[hdr_idx + 1:]:
        label = row[0]
        if label is None:
            break
        label = str(label).strip()
        if label in JURISDICTIONS:
            jurisdiction = label
        else:
            category = label
            jurisdiction = "AUST"
        for col_idx, year in year_cols:
            value = row[col_idx]
            if value is None:
                continue
            records.append({
                "category": category,
                "jurisdiction": jurisdiction,
                "financial_year": year,
                "value": value,
            })
    return title, records


def first_cell(row):
    """Cell text lands in different columns depending on merges; take the first non-empty one."""
    for v in row:
        if v is not None:
            return v
    return None


def parse_cover_page(ws):
    """Extract the table index (id, title, section) from the Contents block."""
    rows = list(ws.iter_rows(min_row=1, max_row=250, values_only=True))
    contents_start = next(i for i, r in enumerate(rows) if first_cell(r) == "Contents")
    desc_start = next(i for i, r in enumerate(rows) if first_cell(r) == "Description and Notes")

    index = []
    section = None
    table_re = re.compile(r"^(Table \d+\.\d+)\.\s*(.+)$")
    for row in rows[contents_start + 1:desc_start]:
        cell = first_cell(row)
        if cell is None:
            continue
        m = table_re.match(cell)
        if m:
            index.append({"table": m.group(1), "title": m.group(2), "section": section})
        else:
            section = cell.strip()
    return index


def main():
    OUT.mkdir(exist_ok=True)
    wb = openpyxl.load_workbook(RAW, read_only=True, data_only=True)

    index = parse_cover_page(wb["Cover page"])

    all_rows = []
    for tname in wb.sheetnames:
        if not tname.startswith("Table "):
            continue
        table_id = tname.replace("Table ", "")
        title, records = parse_table(wb[tname])
        slug = "table-" + table_id.replace(".", "-")
        with open(OUT / f"{slug}.csv", "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=["category", "jurisdiction", "financial_year", "value"])
            w.writeheader()
            w.writerows(records)
        for r in records:
            all_rows.append({"table": tname, "title": title, **r})
        print(f"{tname}: {len(records)} rows -> {slug}.csv")

    with open(OUT / "all-tables-long.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["table", "title", "category", "jurisdiction", "financial_year", "value"])
        w.writeheader()
        w.writerows(all_rows)
    print(f"all-tables-long.csv: {len(all_rows)} rows")

    with open(OUT / "table-index.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["table", "title", "section"])
        w.writeheader()
        for e in index:
            w.writerow(e)
    print(f"table-index.csv: {len(index)} rows")


if __name__ == "__main__":
    main()
