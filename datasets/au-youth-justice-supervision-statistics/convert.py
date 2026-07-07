import openpyxl
import re
import csv
import os

RAW_DIR = "raw"
OUT_TABLES_DIR = "data/tables"
OUT_INDEX = "data/table-index.csv"
OUT_LONG = "data/all-tables-long.csv"

os.makedirs(OUT_TABLES_DIR, exist_ok=True)

WORKBOOKS = [
    ("s1-s33.xlsx", "Characteristics of young people under supervision"),
    ("s34-s71.xlsx", "Characteristics of young people under community-based supervision"),
    ("s72-s125.xlsx", "Characteristics of young people in detention"),
    ("s126-s141.xlsx", "State and territory summary"),
    ("s142-s151.xlsx", "Population numbers"),
]

TITLE_RE = re.compile(r"^Table\s+(S\d+[a-zA-Z]?)\s*:\s*(.*)$")

VALUE_TOKENS = {
    "nsw", "vic", "qld", "wa", "sa", "tas", "act", "nt", "aust",
    "australia", "australia excl nt", "australia excl. nt", "total",
}

def is_value_header(text):
    if text is None:
        return False
    t = str(text).strip().lower()
    if t == "":
        return False
    if t in VALUE_TOKENS:
        return True
    if re.fullmatch(r"\d+", t):
        return True
    if re.fullmatch(r"\d+\+", t):
        return True
    if re.fullmatch(r"\d{1,4}\s*[-–]\s*\d{1,4}", t):
        return True
    return False

def compute_split(header):
    split = len(header)
    for i, h in enumerate(header):
        if is_value_header(h):
            split = i
            break
    if split == len(header):
        split = max(len(header) - 1, 0)
    return split

def slug(text):
    s = str(text).strip()
    s = s.replace("–", "-").replace("—", "-")
    s = re.sub(r"[^\w\s+-]", "", s)
    s = s.replace("+", "plus")
    s = re.sub(r"[\s-]+", "_", s)
    return s.lower().strip("_")

def is_blank_row(row):
    return all(c is None or (isinstance(c, str) and c.strip() == "") for c in row)

NOTE_RE = re.compile(r"^(note|notes|source)\b", re.IGNORECASE)
FOOTNOTE_MARK_RE = re.compile(r"^\(?[a-zA-Z0-9]{1,3}\)?[.)]\s")
DASH_RE = re.compile(r"^[\-–—]\s")

def looks_like_note(text):
    return bool(NOTE_RE.match(text) or FOOTNOTE_MARK_RE.match(text) or DASH_RE.match(text))

def trim_trailing_none(row):
    row = list(row)
    while row and row[-1] is None:
        row.pop()
    return row

def pad(row, width):
    r = list(row[:width])
    while len(r) < width:
        r.append(None)
    return r

tables = []

for fname, workbook_label in WORKBOOKS:
    path = os.path.join(RAW_DIR, fname)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() in ("contents", "notes and symbols", "notes and symbol"):
            continue
        ws = wb[sheet_name]
        state = "seek_title"
        cur = None

        def finalize():
            if cur is not None and cur.get("header"):
                tables.append(cur)

        for row in ws.iter_rows(values_only=True):
            cell0 = row[0] if row and len(row) > 0 else None
            m = TITLE_RE.match(str(cell0).strip()) if isinstance(cell0, str) else None

            if m:
                finalize()
                cur = {
                    "code": m.group(1),
                    "title": m.group(2).strip(),
                    "source_file": fname,
                    "workbook_label": workbook_label,
                    "sheet": sheet_name,
                    "header": None,
                    "dim_idx": None,
                    "val_idx": None,
                    "rows": [],
                    "notes": [],
                    "section": None,
                    "has_section": False,
                }
                state = "seek_header"
                continue

            reprocess = True
            while reprocess:
                reprocess = False
                if state == "seek_title":
                    break
                elif state == "seek_header":
                    if is_blank_row(row):
                        break
                    non_none = [c for c in row if c is not None and str(c).strip() != ""]
                    if len(non_none) < 2:
                        cur["title"] = (cur["title"] + ": " + str(non_none[0]).strip()) if non_none else cur["title"]
                        break
                    header = trim_trailing_none(row)
                    cur["header"] = header
                    split = compute_split(header)
                    cur["dim_idx"] = list(range(0, split))
                    cur["val_idx"] = list(range(split, len(header)))
                    state = "in_data"
                    break
                elif state == "in_data":
                    width = len(cur["header"])
                    r = pad(row, width)
                    val_idx = cur["val_idx"]
                    dim_idx = cur["dim_idx"]
                    if dim_idx and r[dim_idx[-1]] is None:
                        # spreadsheet artifact row (no genuine row label at all)
                        break
                    if all(r[i] is None for i in val_idx):
                        non_none = [c for c in r if c is not None and str(c).strip() != ""]
                        text0 = str(non_none[0]).strip() if non_none else ""
                        if not text0 or looks_like_note(text0):
                            state = "seek_notes"
                            reprocess = True
                            continue
                        # mid-table section heading (e.g. "Unsentenced"), not a footnote
                        cur["section"] = text0
                        cur["has_section"] = True
                        break
                    cur["rows"].append((r, cur["section"]))
                    break
                elif state == "seek_notes":
                    if is_blank_row(row):
                        break
                    text = " ".join(str(c).strip() for c in row if c is not None and str(c).strip() != "")
                    if text:
                        cur["notes"].append(text)
                    break
        finalize()
    wb.close()

print(f"Parsed {len(tables)} sub-tables total")

index_rows = []
long_rows = []

for t in tables:
    header = t["header"]
    dim_idx = t["dim_idx"]
    val_idx = t["val_idx"]

    last_vals = [None] * len(dim_idx)
    filled_rows = []
    sections = []
    for r, section in t["rows"]:
        newr = list(r)
        for k, di in enumerate(dim_idx):
            if newr[di] is None:
                newr[di] = last_vals[k]
            else:
                last_vals[k] = newr[di]
        filled_rows.append(newr)
        sections.append(section)

    dim_headers = [slug(header[i]) for i in dim_idx]
    val_headers = [slug(header[i]) for i in val_idx]
    orig_dim_headers = [str(header[i]) for i in dim_idx]
    orig_val_headers = [str(header[i]) for i in val_idx]

    # long format: built from the un-shifted rows, using original index positions
    for row_i, r in enumerate(filled_rows):
        dims_str = "; ".join(
            f"{orig_dim_headers[k]}={r[dim_idx[k]]}" for k in range(len(dim_idx))
            if r[dim_idx[k]] is not None and str(r[dim_idx[k]]).strip() != ""
        )
        if sections[row_i] is not None:
            dims_str = f"Section={sections[row_i]}" + ("; " + dims_str if dims_str else "")
        for k, vi in enumerate(val_idx):
            val = r[vi]
            long_rows.append({
                "table_code": t["code"],
                "table_title": t["title"],
                "dimensions": dims_str,
                "series": orig_val_headers[k],
                "value": "" if val is None else val,
            })

    if t["has_section"]:
        dim_headers = ["section"] + dim_headers
        filled_rows = [[sections[i]] + r for i, r in enumerate(filled_rows)]

    out_path = os.path.join(OUT_TABLES_DIR, f"{t['code']}.csv")
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(dim_headers + val_headers)
        for r in filled_rows:
            w.writerow(["" if v is None else v for v in r])

    index_rows.append({
        "table_code": t["code"],
        "table_title": t["title"],
        "source_file": t["source_file"],
        "workbook_label": t["workbook_label"],
        "sheet": t["sheet"],
        "n_rows": len(filled_rows),
        "dimension_columns": "; ".join(dim_headers),
        "value_columns": "; ".join(val_headers),
        "notes": " | ".join(t["notes"]),
    })

with open(OUT_INDEX, "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=[
        "table_code", "table_title", "source_file", "workbook_label", "sheet",
        "n_rows", "dimension_columns", "value_columns", "notes",
    ])
    w.writeheader()
    for row in index_rows:
        w.writerow(row)

with open(OUT_LONG, "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=["table_code", "table_title", "dimensions", "series", "value"])
    w.writeheader()
    for row in long_rows:
        w.writerow(row)

print(f"Wrote {len(index_rows)} table CSVs, {len(long_rows)} long-format rows")
