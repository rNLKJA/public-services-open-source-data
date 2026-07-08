#!/usr/bin/env python3
"""Convert DIT's raw Boat Registrations and Boat Licence Statistics workbooks into tidy CSVs."""
import csv
from pathlib import Path
from collections import defaultdict

import openpyxl

HERE = Path(__file__).parent
RAW_REG = HERE / "raw" / "boat-registrations"
RAW_LIC = HERE / "raw" / "boat-licence-statistics"
DATA = HERE / "data"

# Only the 10 "current era" annual editions share a consistent 18/19-sheet, %-column
# layout (FY2015-16 onward). The bundled 2007-2015 zip uses an incompatible, simpler
# legacy layout (no ABP/fee-group/secondary-engine/postcode/location sheets, no %
# columns) and is kept in raw/ for provenance only -- see README "Known limitations".
REGISTRATION_FILES = [
    ("2015-16", "rba011-analysis-of-registered-motor-boats_2015_2016.xlsx"),
    ("2016-17", "rba011-analysis-of-registered-motor-boats_2016_2017.xlsx"),
    ("2017-18", "rba011-analysis-of-registered-motor-boats_2017_2018.xlsx"),
    ("2018-19", "rba011-analysis-of-registered-motor-boats_2018_2019.xlsx"),
    ("2019-20", "rba011-analysis-of-registered-motor-boats_2019_2020.xlsx"),
    ("2020-21", "rba011-analysis-of-registered-motor-boats-3.xlsx"),
    ("2021-22", "rba011-analysis-of-registered-motor-boats.xlsx"),
    ("2022-23", "rba011-analysis-of-registered-motor-boats-14.xlsx"),
    ("2023-24", "boat-registrations-2023-2024.xlsx"),
    ("2024-25", "boat-registrations-2024-2025.xlsx"),
]

METRIC_MAP = {
    "New Boat Reg": "new_boat_registrations",
    "% New Boat Reg": "new_boat_registrations_pct",
    "Boats Currently Reg": "boats_currently_registered",
    "% Boats Currently Reg": "boats_currently_registered_pct",
}


def convert_registrations():
    rows_out = []
    for fy, fname in REGISTRATION_FILES:
        wb = openpyxl.load_workbook(RAW_REG / fname, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_rows = list(ws.iter_rows(values_only=True))
            header_row_idx = next(
                (i for i, row in enumerate(all_rows) if "New Boat Reg" in row), None
            )
            if header_row_idx is None:
                continue
            header = all_rows[header_row_idx]
            col_map = {idx: METRIC_MAP[val] for idx, val in enumerate(header) if val in METRIC_MAP}
            for row in all_rows[header_row_idx + 1:]:
                category_value = row[0]
                if category_value is None:
                    continue
                rec = {
                    "financial_year": fy,
                    "category": sheet_name,
                    "category_value": category_value,
                    "new_boat_registrations": None,
                    "new_boat_registrations_pct": None,
                    "boats_currently_registered": None,
                    "boats_currently_registered_pct": None,
                }
                for col_idx, field in col_map.items():
                    if col_idx < len(row):
                        rec[field] = row[col_idx]
                rows_out.append(rec)

    # Validate: for each (year, category), count/pct should imply a consistent grand total.
    checks = defaultdict(list)
    for r in rows_out:
        if r["new_boat_registrations"] and r["new_boat_registrations_pct"]:
            checks[(r["financial_year"], r["category"])].append(
                r["new_boat_registrations"] / r["new_boat_registrations_pct"]
            )
    for k, totals in checks.items():
        assert max(totals) - min(totals) <= 2, f"inconsistent implied total for {k}"

    fields = ["financial_year", "category", "category_value", "new_boat_registrations",
              "new_boat_registrations_pct", "boats_currently_registered",
              "boats_currently_registered_pct"]
    out_path = DATA / "boat-registrations-by-category-2015-16-to-2024-25.csv"
    with open(out_path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows_out)
    print(f"Wrote {len(rows_out)} rows to {out_path.name}")


def to_int(s):
    s = s.strip().replace(",", "")
    return int(s) if s else None


def to_pct(s):
    s = s.strip().replace("%", "")
    return round(float(s) / 100, 6) if s else None


def convert_licence_statistics():
    with open(RAW_LIC / "boat-licence-statistics.csv") as f:
        reader = list(csv.reader(f))

    header_idx = next(i for i, row in enumerate(reader) if row and row[0] == "Issue Year")

    rows_out = []
    for row in reader[header_idx + 1:]:
        if not row or not row[0].strip():
            continue
        issue_year = row[0].strip()
        is_total = issue_year == "Report Totals"
        if not (issue_year.isdigit() or is_total):
            continue  # skips the trailing "RBA014 / Page:1" footer row
        rows_out.append({
            "issue_year": None if is_total else int(issue_year),
            "is_total_row": is_total,
            "year_sub_total": to_int(row[1]),
            "female": to_int(row[2]),
            "female_pct_of_year": to_pct(row[3]),
            "male": to_int(row[5]),
            "male_pct_of_year": to_pct(row[8]),
            "unknown": to_int(row[9]),
            "unknown_pct_of_year": to_pct(row[10]),
        })

    # Validate: female + male + unknown should reconstruct each row's own sub-total.
    for r in rows_out:
        parts = [v for v in (r["female"], r["male"], r["unknown"]) if v is not None]
        assert sum(parts) == r["year_sub_total"], f"mismatch for {r['issue_year']}"

    fields = ["issue_year", "is_total_row", "year_sub_total", "female", "female_pct_of_year",
              "male", "male_pct_of_year", "unknown", "unknown_pct_of_year"]
    out_path = DATA / "boat-licence-issues-by-year-and-gender.csv"
    with open(out_path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows_out)
    print(f"Wrote {len(rows_out)} rows to {out_path.name}")


if __name__ == "__main__":
    DATA.mkdir(exist_ok=True)
    convert_registrations()
    convert_licence_statistics()
