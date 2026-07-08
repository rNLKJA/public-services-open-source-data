"""Reshape the raw Dogs and Cats Online (DACO) workbooks into tidy CSVs.

The source publishes one workbook per measure per financial-year edition
(2021-22 through 2024-25). Animals/Owners/MostPopular/Top5Breeds are each a
point-in-time snapshot (no year field of their own), so the 4 editions are
merged into one long table per measure with an `edition`/`extract_date`
column identifying which annual snapshot a row came from. Incidents already
carries its own `FinancialYear` column and each edition republishes the full
cumulative history (counts revised upward in later editions as late-reported
incidents are added) — so only the latest (2024-25) edition is used for
Incidents, to avoid duplicate/inconsistent historical rows.

Run: python3 convert.py (requires openpyxl)
"""
import csv
from pathlib import Path

import openpyxl

RAW = Path(__file__).parent / "raw"
DATA = Path(__file__).parent / "data"
DATA.mkdir(exist_ok=True)

EDITIONS = [
    ("2021-22", "2022-05-30"),
    ("2022-23", "2023-05-30"),
    ("2023-24", "2024-05-30"),
    ("2024-25", "2025-05-30"),
]


def rows(path, sheet_name=None):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else [wb[s] for s in wb.sheetnames if s != "Notes"][0]
    data = list(ws.iter_rows(values_only=True))
    wb.close()
    return data


def find_file(edition, *name_fragments):
    folder = RAW / edition
    for f in folder.iterdir():
        name_lower = f.name.lower()
        if all(frag.lower() in name_lower for frag in name_fragments):
            return f
    raise FileNotFoundError(f"No file matching {name_fragments} in {folder}")


# ---------------------------------------------------------------------------
# 1. Animal population by council (Animals workbooks)
# ---------------------------------------------------------------------------
animal_rows = []
for edition, extract_date in EDITIONS:
    f = find_file(edition, "animal")
    data = rows(f)
    header = [str(h).strip() for h in data[0]]
    # RegistrationStatus renamed to Reg_status from 2022-23 onward; same meaning either way.
    for row in data[1:]:
        if not row or row[0] is None:
            continue
        council, species, gender, reg_status, microchipped, desexed, count = row
        animal_rows.append({
            "edition": edition,
            "extract_date": extract_date,
            "council": council,
            "animal_species": species,
            "gender": gender,
            "registration_status": reg_status,
            "microchipped": "Yes" if microchipped == "Y" else "No",
            "desexed": "Yes" if desexed == "Y" else "No",
            "count": count,
        })

with open(DATA / "animal-population-by-council.csv", "w", newline="", encoding="utf-8") as fh:
    w = csv.DictWriter(fh, fieldnames=["edition", "extract_date", "council", "animal_species",
                                        "gender", "registration_status", "microchipped", "desexed", "count"])
    w.writeheader()
    w.writerows(animal_rows)
print(f"Wrote {len(animal_rows)} rows to animal-population-by-council.csv")

# ---------------------------------------------------------------------------
# 2. Owners by council (Owners workbooks)
# ---------------------------------------------------------------------------
owner_rows = []
for edition, extract_date in EDITIONS:
    f = find_file(edition, "owner")
    data = rows(f)
    for row in data[1:]:
        if not row or row[0] is None:
            continue
        council, gender, age_category, num_dogs, num_cats, count = row
        owner_rows.append({
            "edition": edition,
            "extract_date": extract_date,
            "council": council,
            "gender": gender,
            "age_category": age_category,
            "num_of_dogs": num_dogs,
            "num_of_cats": num_cats,
            "count": count,
        })

with open(DATA / "owners-by-council.csv", "w", newline="", encoding="utf-8") as fh:
    w = csv.DictWriter(fh, fieldnames=["edition", "extract_date", "council", "gender",
                                        "age_category", "num_of_dogs", "num_of_cats", "count"])
    w.writeheader()
    w.writerows(owner_rows)
print(f"Wrote {len(owner_rows)} rows to owners-by-council.csv")

# ---------------------------------------------------------------------------
# 3. Incidents by financial year — latest (2024-25) edition only (cumulative superset)
# ---------------------------------------------------------------------------
incident_rows = []
f = find_file("2024-25", "incident")
data = rows(f)
for row in data[1:]:
    if not row or row[0] is None:
        continue
    location_type, time, season, leash_status, incident_type, victim_type, financial_year, count = row
    incident_rows.append({
        "financial_year": financial_year,
        "location_type": location_type,
        "time": time,
        "season": season,
        "offending_animal_leash_status": leash_status,
        "incident_type": incident_type,
        "victim_type": victim_type,
        "incident_count": count,
    })

with open(DATA / "incidents-by-financial-year.csv", "w", newline="", encoding="utf-8") as fh:
    w = csv.DictWriter(fh, fieldnames=["financial_year", "location_type", "time", "season",
                                        "offending_animal_leash_status", "incident_type",
                                        "victim_type", "incident_count"])
    w.writeheader()
    w.writerows(incident_rows)
print(f"Wrote {len(incident_rows)} rows to incidents-by-financial-year.csv "
      f"(2024-25 edition only — republishes full cumulative history back to before FY2018, "
      f"see README)")

# ---------------------------------------------------------------------------
# 4. Most popular breeds and names (MostPopular workbooks, 4 sheets each)
#
# Breed sheets changed shape between editions: 2021-22/2022-23/2023-24 report
# (Primary breed, Secondary breed, Count) — DACO records a primary+secondary
# breed pair per animal, secondary='Cross Breed' for mixed breeds, or repeats
# the primary breed for purebreds. 2024-25 collapses this to (Breed, Sum of
# Count) — primary breed only, secondary-breed dimension dropped by the
# source. Both are kept as-is per edition; `secondary_breed` is blank for 2024-25.
# ---------------------------------------------------------------------------
breed_rows = []
name_rows = []
for edition, extract_date in EDITIONS:
    f = find_file(edition, "popular")
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    sheet_map = {s.lower(): s for s in wb.sheetnames}
    for species, key in [("Dog", "dog breeds - sa"), ("Cat", "cat breeds - sa")]:
        ws = wb[sheet_map[key]]
        data = list(ws.iter_rows(values_only=True))
        header = [str(h).strip().lower() for h in data[0]]
        has_secondary = "secondary breed" in header
        for row in data[1:]:
            if not row or row[0] is None:
                continue
            if has_secondary:
                primary_breed, secondary_breed, count = row
            else:
                primary_breed, count = row
                secondary_breed = ""
            breed_rows.append({
                "edition": edition, "extract_date": extract_date, "species": species,
                "primary_breed": primary_breed, "secondary_breed": secondary_breed, "count": count,
            })
    for species, key in [("Dog", "top 100 dog names"), ("Cat", "top 100 cat names")]:
        ws = wb[sheet_map[key]]
        for row in list(ws.iter_rows(values_only=True))[1:]:
            if not row or row[0] is None:
                continue
            name_rows.append({
                "edition": edition, "extract_date": extract_date,
                "species": species, "name": row[0], "count": row[1],
            })
    wb.close()

with open(DATA / "most-popular-breeds.csv", "w", newline="", encoding="utf-8") as fh:
    w = csv.DictWriter(fh, fieldnames=["edition", "extract_date", "species", "primary_breed",
                                        "secondary_breed", "count"])
    w.writeheader()
    w.writerows(breed_rows)
print(f"Wrote {len(breed_rows)} rows to most-popular-breeds.csv")

with open(DATA / "most-popular-names.csv", "w", newline="", encoding="utf-8") as fh:
    w = csv.DictWriter(fh, fieldnames=["edition", "extract_date", "species", "name", "count"])
    w.writeheader()
    w.writerows(name_rows)
print(f"Wrote {len(name_rows)} rows to most-popular-names.csv")

# ---------------------------------------------------------------------------
# 5. Top 5 dog breeds by council (long format: one row per council/rank)
# ---------------------------------------------------------------------------
top5_rows = []
for edition, extract_date in EDITIONS:
    f = find_file(edition, "5", "dog", "breed")
    data = rows(f, sheet_name=[s for s in openpyxl.load_workbook(f, read_only=True).sheetnames
                                if s != "Notes"][0])
    # 2021-22/2022-23/2023-24 have an extra title row above the real header ('Popular breeds').
    header_row_idx = 0
    for i, row in enumerate(data[:2]):
        if row and row[0] and "council" in str(row[0]).lower():
            header_row_idx = i
            break
    for row in data[header_row_idx + 1:]:
        if not row or row[0] is None:
            continue
        council = row[0]
        for rank, breed in enumerate(row[1:6], start=1):
            if breed is None:
                continue
            top5_rows.append({
                "edition": edition, "extract_date": extract_date,
                "council": council, "rank": rank, "breed": breed,
            })

with open(DATA / "top-5-dog-breeds-by-council.csv", "w", newline="", encoding="utf-8") as fh:
    w = csv.DictWriter(fh, fieldnames=["edition", "extract_date", "council", "rank", "breed"])
    w.writeheader()
    w.writerows(top5_rows)
print(f"Wrote {len(top5_rows)} rows to top-5-dog-breeds-by-council.csv")
