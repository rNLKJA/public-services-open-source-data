"""
Builds the tidy CSVs in this data/ folder from the raw ICAC/OPI source files in ../raw/.

Source files span two eras because SA's anti-corruption body was restructured (the single
"Independent Commissioner Against Corruption" plus a separate Office for Public Integrity
became the "Independent Commission Against Corruption" in 2021):
  - icac-and-opi-*  (predecessor org, 2013-14 to 2020-21 CKAN time-series CSVs)
  - icac-*.xlsx     (current Commission org, multi-year xlsx resources through 2022-23)

Values below are transcribed directly from each raw file's cells (verified via csv.reader /
openpyxl with column position preserved) -- run this script's assertions to cross-check
against the source "Total" rows.
"""
import csv
import openpyxl
from pathlib import Path

RAW = Path(__file__).parent.parent / "raw"
OUT = Path(__file__).parent


def write_csv(name, header, rows):
    path = OUT / name
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)
    print(f"wrote {path} ({len(rows)} rows)")


def clean(s):
    """Collapse embedded newlines/whitespace from merged Excel cells into single spaces."""
    if not isinstance(s, str):
        return s
    return " ".join(s.split())


# ---------------------------------------------------------------------------
# 1. Complaints and reports received under the ICAC Act, 2013-14 to 2020-21
#    (source: icac-and-opi ... section-i, "Complaints and reports received under the ICAC Act")
# ---------------------------------------------------------------------------
complaints_reports = [
    ("2020-21", 596, 48.1, 642, 51.9, 1238, 100),
    ("2019-20", 495, 39.4, 762, 60.6, 1257, 100),
    ("2018-19", 487, 40, 742, 60, 1229, 100),
    ("2017-18", 415, 37, 706, 63, 1121, 100),
    ("2016-17", 428, 36, 772, 64, 1200, 100),
    ("2015-16", 463, 44, 600, 56, 1063, 100),
    ("2014-15", 453, 49, 474, 51, 927, 100),
    ("2013-14", 462, 50, 461, 50, 923, 100),
]
for r in complaints_reports:
    assert r[1] + r[3] == r[5], r
write_csv(
    "complaints_and_reports_received_icac_act.csv",
    ["financial_year", "complaints_number", "complaints_pct", "reports_number", "reports_pct", "total_number", "total_pct"],
    complaints_reports,
)

# ---------------------------------------------------------------------------
# 2. Enquiries received by the OPI outside the Commissioner's jurisdiction
# ---------------------------------------------------------------------------
enquiries = [
    ("2020-21", 1059, 186, 1245, ""),
    ("2019-20", 1136, 313, 1449, ""),
    ("2018-19", 1157, 306, 1463, ""),
    ("2017-18", 1147, 174, 1321, ""),
    ("2016-17", 374, 59, 433, ""),
    ("2015-16", 251, 44, 295, ""),
    ("2014-15", 289, 40, 329, ""),
    ("2013-14", 624, "", 624, "reported as a single combined number in 2013-14, not split into general enquiries/contacts"),
]
write_csv(
    "enquiries_received_by_opi_outside_jurisdiction.csv",
    ["financial_year", "general_enquiries", "contacts", "total", "note"],
    enquiries,
)

# ---------------------------------------------------------------------------
# 3. Subject of complaints and reports received under the ICAC Act
#    Full Number+% breakdown exists for 2018-19 to 2020-21 only; earlier years (2013-14 to
#    2017-18) were published as percentages only, split by Complaints/Reports (no Total%, no counts).
# ---------------------------------------------------------------------------
subject_full = [
    # financial_year, subject_category, complaints_number, complaints_pct, reports_number, reports_pct, total_number, total_pct
    ("2020-21", "Ombudsman", 2, 0.3, 1, 0.2, 3, 0.2),
    ("2020-21", "Inquiry Agency", None, None, None, None, None, None),
    ("2020-21", "Local government", 112, 18.8, 141, 22.0, 253, 20.4),
    ("2020-21", "Member of Parliament", 35, 5.9, 14, 2.2, 49, 4.0),
    ("2020-21", "Statutory Authority", 112, 18.8, 99, 15.4, 211, 17.0),
    ("2020-21", "State government", 334, 56.0, 386, 60.1, 720, 58.2),
    ("2020-21", "SA Police", None, None, None, None, None, None),
    ("2020-21", "Private / unknown", 1, 0.2, 1, 0.2, 2, 0.2),
    ("2020-21", "Total", 596, 100, 642, 100.1, 1238, 100.0),
    ("2019-20", "Ombudsman", 2, 0.4, 1, 0.1, 3, 0.2),
    ("2019-20", "Inquiry Agency", None, None, None, None, None, None),
    ("2019-20", "Local government", 89, 18, 201, 26.4, 290, 23.1),
    ("2019-20", "Member of Parliament", 13, 2.6, 9, 1.2, 22, 1.8),
    ("2019-20", "Statutory Authority", 105, 21.2, 120, 15.7, 225, 17.9),
    ("2019-20", "State government", 285, 57.6, 429, 56.3, 714, 56.8),
    ("2019-20", "SA Police", None, None, None, None, None, None),
    ("2019-20", "Private / unknown", 1, 0.2, 2, 0.3, 3, 0.2),
    ("2019-20", "Total", 495, 100, 762, 100, 1257, 100),
    ("2018-19", "Ombudsman", 2, 0.4, 0, 0, 2, 0.2),
    ("2018-19", "Inquiry Agency", None, None, None, None, None, None),
    ("2018-19", "Local government", 105, 21.6, 136, 18.3, 241, 19.6),
    ("2018-19", "Member of Parliament", 16, 3.3, 14, 1.9, 30, 2.4),
    ("2018-19", "Statutory Authority", 126, 25.9, 86, 11.6, 212, 17.2),
    ("2018-19", "State government", 237, 48.7, 504, 67.9, 741, 60.3),
    ("2018-19", "SA Police", None, None, None, None, None, None),
    ("2018-19", "Private / unknown", 1, 0.2, 2, 0.3, 3, 0.2),
    ("2018-19", "Total", 487, 100, 742, 100, 1229, 99.9),
]
write_csv(
    "subject_of_complaints_and_reports_2018_19_to_2020_21.csv",
    ["financial_year", "subject_category", "complaints_number", "complaints_pct", "reports_number", "reports_pct", "total_number", "total_pct"],
    subject_full,
)

# percentage-only years (2013-14 to 2017-18): Complaints% and Reports% only, no counts/totals published.
# "SA Police" was reported as "Separate report" for 2018-19 to 2020-21 (see PCD Act files, added after
# the Police Complaints and Discipline Act 2016 commenced 4 Sept 2017) and included within this same
# table's ordinary categories before that (footnote 1: "Received prior to 4 September 2017").
subject_pct_only = [
    # financial_year, subject_category, complaints_pct, reports_pct
    ("2017-18", "Ombudsman", None, None),
    ("2017-18", "Inquiry Agency", 1.0, None),
    ("2017-18", "Local government", 21.8, 15.6),
    ("2017-18", "Member of Parliament", 1.2, 1.0),
    ("2017-18", "Statutory Authority", 16.9, 10.3),
    ("2017-18", "State government", 48.7, 70.4),
    ("2017-18", "SA Police (received prior to 4 September 2017)", 9.2, 2.3),
    ("2017-18", "Private / unknown", 1.2, 0.4),
    ("2016-17", "Ombudsman", None, None),
    ("2016-17", "Inquiry Agency", 0.5, 0.5),
    ("2016-17", "Local government", 16.1, 15.7),
    ("2016-17", "Member of Parliament", 0.9, 0.5),
    ("2016-17", "Statutory Authority", 8.2, 6.2),
    ("2016-17", "State government", 32.2, 69.2),
    ("2016-17", "SA Police (received prior to 4 September 2017)", 40.7, 7.6),
    ("2016-17", "Private / unknown", 1.4, 0.3),
    ("2015-16", "Ombudsman", None, None),
    ("2015-16", "Inquiry Agency", None, None),
    ("2015-16", "Local government", 22, 15),
    ("2015-16", "Member of Parliament", 2, None),
    ("2015-16", "Statutory Authority", 11, 7),
    ("2015-16", "State government", 64, 77),
    ("2015-16", "SA Police (received prior to 4 September 2017)", None, None),
    ("2015-16", "Private / unknown", 1, None),
    ("2014-15", "Ombudsman", None, None),
    ("2014-15", "Inquiry Agency", None, None),
    ("2014-15", "Local government", 21, 19),
    ("2014-15", "Member of Parliament", 2, 1),
    ("2014-15", "Statutory Authority", 9, 10),
    ("2014-15", "State government", 67, 69),
    ("2014-15", "SA Police (received prior to 4 September 2017)", None, None),
    ("2014-15", "Private / unknown", 1, 1),
    ("2013-14", "Ombudsman", None, None),
    ("2013-14", "Inquiry Agency", None, None),
    ("2013-14", "Local government", 19, 13),
    ("2013-14", "Member of Parliament", 5, 2),
    ("2013-14", "Statutory Authority", 12, 5),
    ("2013-14", "State government", 64, 79),
    ("2013-14", "SA Police (received prior to 4 September 2017)", None, None),
    ("2013-14", "Private / unknown", None, 1),
]
write_csv(
    "subject_of_complaints_and_reports_2013_14_to_2017_18_pct_only.csv",
    ["financial_year", "subject_category", "complaints_pct", "reports_pct"],
    subject_pct_only,
)

# ---------------------------------------------------------------------------
# 4. Timeframe for assessment of all complaints and reports (average working days)
# ---------------------------------------------------------------------------
timeframe = [
    ("2020-21", "18.5"),
    ("2019-20", "17.4"),
    ("2018-19", "14.1"),
    ("2017-18", "15.3"),
    ("2016-17", "18.3"),
    ("2015-16", "25.6"),
    ("2014-15", ""),  # not numerically reported; source notes a 20% improvement vs prior year instead
    ("2013-14", "18"),
]
write_csv(
    "assessment_timeframe_working_days.csv",
    ["financial_year", "average_assessment_working_days"],
    timeframe,
)

# ---------------------------------------------------------------------------
# 5. ICAC's own-initiative matters
# ---------------------------------------------------------------------------
own_initiative_years = ["2020-21", "2019-20", "2018-19", "2017-18", "2016-17", "2015-16", "2014-15", "2013-14"]
own_initiative_categories = {
    "progressed_to_corruption_investigation": [8, 10, 6, 1, "-", 1, 1, 3],
    "referred_as_potential_misconduct_or_maladministration": [2, 3, 2, "-", "-", 1, "-", 1],
    "referred_as_raising_some_other_issue": [3, 1, 1, "-", "-", "-", "-", "-"],
    "resulted_in_no_further_action": [9, 1, 1, "-", "-", "-", "-", "-"],
    "total": [22, 15, 10, 1, 5, "-", "-", "-"],
}
own_initiative_rows = []
for cat, vals in own_initiative_categories.items():
    for year, v in zip(own_initiative_years, vals):
        own_initiative_rows.append((year, cat, v))
write_csv(
    "own_initiative_matters.csv",
    ["financial_year", "category", "count"],
    own_initiative_rows,
)

# ---------------------------------------------------------------------------
# 6. No further action outcomes (complaints/reports assessed and closed with no further action)
# ---------------------------------------------------------------------------
nfa_years = ["2020-21", "2019-20", "2018-19", "2017-18", "2016-17", "2015-16", "2014-15", "2013-14"]
nfa_rows = [
    # category, [(number, pct) per year] -- 2013-14 has number only, no pct published
    ("Reports", [(350, 42.0), (324, 45.4), (288, 41.8), (244, 42), (310, 47), (201, 38), (186, 35), (145, None)]),
    ("Complaints", [(474, 56.9), (388, 54.4), (400, 58.1), (335, 58), (350, 53), (326, 62), (342, 65), (327, None)]),
    ("Own initiative", [(9, 1.1), (1, 0.1), (1, 0.1), ("-", "-"), (1, 1), ("-", "-"), ("-", "-"), ("-", "-")]),
    ("Total", [(833, 100), (713, 99.9), (689, 100), (579, 100), (661, 100), (527, 100), (528, 100), (472, None)]),
]
nfa_long = []
for cat, per_year in nfa_rows:
    for year, (number, pct) in zip(nfa_years, per_year):
        nfa_long.append((year, cat, number, pct))
write_csv(
    "no_further_action_outcomes.csv",
    ["financial_year", "category", "number", "pct"],
    nfa_long,
)

# ---------------------------------------------------------------------------
# 7. General nature of matters investigated by the Commission (formal corruption
#    investigations), 2014-15 to 2022-23 -- current org's self-contained series.
#    Category wording/classification changed between years (ICAC's own footnotes below);
#    kept as separately labelled rows per year rather than force-mapped to one fixed taxonomy.
# ---------------------------------------------------------------------------
wb = openpyxl.load_workbook(RAW / "icac-general-nature-of-matters-investigated-2014-15-to-2022-23.xlsx", data_only=True)
ws = wb.worksheets[0]

nature_rows = []
# Block 1: rows 6-16 (0-indexed) = categories; years at row4/5: 2022-23,2021-22,2020-21,2019-20,2018-19,2017-18(pct only)
years_block1 = ["2022-23", "2021-22", "2020-21", "2019-20", "2018-19", "2017-18"]
for row in ws.iter_rows(min_row=7, max_row=17, values_only=True):
    cat = row[0]
    if not cat:
        continue
    cat = clean(cat)
    # columns: 0=label,1,2 blank, 3=2022-23 Number,4=% ... 13=2017-18 % (only one col)
    vals = row
    idx = 3
    for yi, year in enumerate(years_block1):
        if year != "2017-18":
            number, pct = vals[idx], vals[idx + 1]
            idx += 2
        else:
            number, pct = None, vals[idx]
            idx += 1
        if number == "" or number is None:
            number = None
        if pct == "" :
            pct = None
        if number is not None or pct is not None:
            nature_rows.append((year, cat, number, pct))

# Block 2: rows 30-39 (0-indexed) = categories for 2016-17, 2015-16, 2014-15 (% only)
years_block2 = ["2016-17", "2015-16", "2014-15"]
for row in ws.iter_rows(min_row=31, max_row=40, values_only=True):
    cat = row[0]
    if not cat:
        continue
    cat = clean(cat)
    vals = row[3:6]
    for year, pct in zip(years_block2, vals):
        if pct not in (None, ""):
            nature_rows.append((year, cat, None, pct))

write_csv(
    "general_nature_of_investigations.csv",
    ["financial_year", "category", "number", "pct"],
    nature_rows,
)

# ---------------------------------------------------------------------------
# 8. Corruption investigations completed within 12 months of allocation (file closure KPI)
# ---------------------------------------------------------------------------
wb = openpyxl.load_workbook(RAW / "icac-file-closure-corruption-investigations-completed-2014-15-to-2022-23.xlsx", data_only=True)
ws = wb.worksheets[0]
rows_all = list(ws.iter_rows(values_only=True))
years_a = rows_all[4][2:8]  # 2022-23 .. 2017-18
vals_a = rows_all[5][2:8]
kpi_a = rows_all[6][2:8]
years_b = ("2016-17", "2015-16", "2014-15")
vals_b = (rows_all[11][0], rows_all[11][2], rows_all[11][4])
kpi_b = (rows_all[12][0], rows_all[12][2], rows_all[12][4])

file_closure = []
for y, v, k in zip(list(years_a) + list(years_b), list(vals_a) + list(vals_b), list(kpi_a) + list(kpi_b)):
    pct = round(v * 100, 1) if isinstance(v, (int, float)) else v
    kpi_clean = " ".join(str(k).split()) if k else k
    file_closure.append((y, pct, kpi_clean))
write_csv(
    "corruption_investigations_completed_within_12_months.csv",
    ["financial_year", "pct_completed_within_12_months_of_allocation", "kpi_benchmark"],
    file_closure,
)

# ---------------------------------------------------------------------------
# 9. Matters referred to SA Police for investigation -- year totals only (2013-14 to 2022-23).
#    Category-level breakdown kept separately (10) for the two years it's unambiguously labelled.
# ---------------------------------------------------------------------------
referred_totals = [
    ("2022-23", 29, 100),
    ("2021-22", 33, 100),
    ("2020-21", 31, 100),
    ("2019-20", 44, 100),
    ("2018-19", 76, 100),
    ("2017-18", None, 100),
    ("2016-17", None, 100),
    ("2015-16", None, 100),
    ("2014-15", None, 100),
    ("2013-14", None, 100),
]
write_csv(
    "matters_referred_to_sapol_totals_by_year.csv",
    ["financial_year", "total_number", "total_pct"],
    referred_totals,
)

# ---------------------------------------------------------------------------
# 10. Matters referred to SA Police, by category -- 2021-22 and 2022-23 only (the two years
#     the source publishes with unambiguous category labels alongside the numbers).
# ---------------------------------------------------------------------------
referred_categories = [
    ("Abuse of power for personal or financial gain / Bribery", None, None, 21, 72.73),
    ("Assault", None, None, 1, 3.03),
    ("Breach of ICAC Act provisions regarding disclosure of information", None, None, 1, 3.03),
    ("Bribery", None, None, 3, 9.09),
    ("Failing to act honestly", None, None, 4, 12.12),
    ("Theft / misappropriation / fraud / deception", None, None, None, None),
    ("Improper use or disclosure of information/systems", None, None, 2, 6.06),
    ("Dishonest dealings with documents", None, None, None, None),
    ("Offences associated with the use, sale or supply of prescription and/or illicit drugs", None, None, 1, 3.03),
    ("Miscellaneous (Other)", None, None, None, None),
    ("Deception or abuse of power associated with employment or appointment to public office", 29, 100, None, None),
]
referred_cat_rows = []
for cat, n2223, p2223, n2122, p2122 in referred_categories:
    if n2223 is not None or p2223 is not None:
        referred_cat_rows.append(("2022-23", cat, n2223, p2223))
    if n2122 is not None or p2122 is not None:
        referred_cat_rows.append(("2021-22", cat, n2122, p2122))
write_csv(
    "matters_referred_to_sapol_by_category_2021_22_and_2022_23.csv",
    ["financial_year", "category", "number", "pct"],
    referred_cat_rows,
)

# ---------------------------------------------------------------------------
# 11. Complaints and reports received under the Police Complaints and Discipline Act 2016
#     (i.e. about SA Police specifically), 2017-18 to 2020-21 -- old org, not replicated in
#     the current org's PCD Act resource (which only carries the sanctions table, see 12).
# ---------------------------------------------------------------------------
pcd_intake = [
    ("2020-21", "Complaints", 400, 1593, 1993),
    ("2020-21", "Reports", 397, 23, 420),
    ("2020-21", "Protective Security Officers", None, None, None),  # "Reported separately"
    ("2020-21", "Total", 797, 1616, 2413),
    ("2019-20", "Complaints", 313, 1806, 2119),
    ("2019-20", "Reports", 416, 13, 429),
    ("2019-20", "Protective Security Officers", None, None, None),
    ("2019-20", "Total", 729, 1819, 2548),
    ("2018-19", "Complaints", 279, 1655, 1934),
    ("2018-19", "Reports", 359, 10, 369),
    ("2018-19", "Protective Security Officers", 6, 0, 6),
    ("2018-19", "Total", 644, 1665, 2309),
    ("2017-18", "Complaints", 216, 1459, 1675),
    ("2017-18", "Reports", 283, 27, 310),
    ("2017-18", "Protective Security Officers", 5, 0, 5),
    ("2017-18", "Total", 504, 1486, 1990),
]
write_csv(
    "pcd_act_complaints_and_reports_received_about_sapol.csv",
    ["financial_year", "type", "received_by_iis", "received_by_opi", "total"],
    pcd_intake,
)

# ---------------------------------------------------------------------------
# 12. Sanctions imposed under the PCD Act -- individual case-level outcomes, Sept 2017 to
#     June 2023 (current org's self-contained series). "SA Police reference no." / "Officer
#     No." is a per-case reporting number reused each period, not a persistent personnel ID
#     (numbers restart/repeat across years) -- no officer names appear anywhere in the source.
# ---------------------------------------------------------------------------
sanctions = [
    # (reporting_period, breach_category, case_reference_no, breach_description, sanction_outcome)
    ("2022-23", "Criminal offence (s26(1)(a))", "10", "Criminal - Sexual offence", "Employment with SA Police terminated after criminal conviction."),
    ("2022-23", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "1", "Clause 4 Performance of Orders / Duties - Disobey Orders and Failure to carry out a lawful order", "Fine / Recorded Reprimand"),
    ("2022-23", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "2", "Clause 3 Conduct Prejudicial - Corrupt or Improper Practice (33 counts); Clause 4 Performance of Orders / Duties - Failure to carry out a General Order (2 counts)", "Fine / Recorded Reprimand"),
    ("2022-23", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "3", "Clause 8 Conflict of Interest - Conflict of Interest; Clause 10 Confidentiality of Information - Improper Disclosure of Information", "Fine / Recorded Reprimand"),
    ("2022-23", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "4", "Clause 4 Performance of Orders / Duties - Failure to carry out a General Order; Clause 10 Confidentiality of Information - Improper release / Access", "Fine / Recorded Reprimand / Training and Education"),
    ("2022-23", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "5", "Clause 4 Performance of Orders / Duties - Disobey Orders", "Fine / Recorded Reprimand"),
    ("2022-23", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "6", "Clause 4 Performance of Orders / Duties - Failure to carry out a lawful order", "Unrecorded Reprimand"),
    ("2022-23", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "6", "Clause 4 Performance of Orders / Duties - Failure to carry out a lawful order; Clause 7 Conduct Public / SA Police - Conduct towards Employee", "Fine / Recorded Reprimand"),
    ("2022-23", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "7", "Clause 3 Conduct Prejudicial - Reflects Adversely", "Recorded Reprimand / Suspension without Pay (Penalty)"),
    ("2022-23", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "8", "Clause 4 Performance of Orders / Duties - Failure to carry out a lawful order; Clause 5 Negligence - Neglect of Duty", "Fine / Recorded Reprimand"),
    ("2022-23", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "9", "Clause 2 Honesty and Integrity - Falsehood / prevarication; Clause 4 Performance of Orders / Duties - Disobey Orders; Clause 8 Conflict of Interest - Conflict of Interest", "Fine / Recorded Reprimand / Suspension without Pay (Penalty)"),
    ("2022-23", "Code of Conduct - Police Regulations 2014 (s26(1)(b)/(c))", None, "NIL", None),
    ("2022-23", "Code of Conduct - Police Regulations 1999 (s26(1)(b)/(c))", None, "NIL", None),
    ("2021-22", "Criminal offence (s26(1)(a))", None, "NIL", None),
    ("2021-22", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "1", "Clause 4 Performance of orders / duties - Failure to carry out a lawful order", "Fine / recorded reprimand / training and education"),
    ("2021-22", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "2", "Clause 10 Confidentiality of information - improper release/access", "Recorded reprimand / fine"),
    ("2021-22", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "3", "Clause 3 Conduct prejudicial - reflects adversely (2 counts); Clause 4 Performance of orders / duties - Failure to carry out a general order (2 counts)", "Fine / recorded reprimand"),
    ("2021-22", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "4", "Clause 7 Conduct public / SAPOL - conduct towards employee", "Fine / transfer / recorded reprimand / counselling"),
    ("2021-22", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "5", "Clause 4 Performance of orders / duties - failure to carry out a general order", "Fine / transfer / recorded reprimand"),
    ("2021-22", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "6", "Clause 2 Honesty and Integrity - Honesty; Clause 3 Conduct Prejudicial - reflects adversely; Clause 5 Negligence - neglect of duty", "Fine / recorded reprimand"),
    ("2021-22", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "7", "Clause 4 Performance of orders / duties - disobey orders; Clause 10 Confidentiality of information - improper disclosure of information", "Fine / recorded reprimand"),
    ("2021-22", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "8", "Clause 4 Performance of orders / duties - disobey orders; Clause 10 Confidentiality of information - improper disclosure of information", "Fine / recorded reprimand"),
    ("2021-22", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "9", "Clause 3 Conduct prejudicial - reflects adversely", "Fine / transfer / recorded reprimand"),
    ("2021-22", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "10", "Clause 4 Performance of orders / duties - disobey orders", "Recorded reprimand"),
    ("2021-22", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "11", "Clause 4 Performance of orders / duties - disobey orders; Clause 4 Performance of orders / duties - failure to carry out a general order", "Fine / recorded reprimand"),
    ("2021-22", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "12", "Clause 4 Performance of orders / duties - disobey orders", "Transfer / recorded reprimand / reduction in rank"),
    ("2021-22", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "13", "Clause 6 Proper exercise of authority - excessive force", "Fine / recorded reprimand"),
    ("2021-22", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "14", "Clause 3 Conduct prejudicial - reflects adversely", "Fine / recorded reprimand"),
    ("2021-22", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "15", "Clause 10 Confidentiality of information - improper release / access", "Fine / recorded reprimand"),
    ("2021-22", "Code of Conduct - Police Regulations 2014 (s26(1)(b)/(c))", None, "NIL", None),
    ("2021-22", "Code of Conduct - Police Regulations 1999 (s26(1)(b)/(c))", None, "NIL", None),
    ("2020-21", "Criminal offence (s26(1)(a))", "12", "Criminal - Drug Offence", "Employment Terminated"),
    ("2020-21", "Criminal offence (s26(1)(a))", "18", "Criminal - Fraud Offence", "Employment Terminated"),
    ("2020-21", "Criminal offence (s26(1)(a))", "19", "Criminal - Offence Against the Person", "Employment Terminated"),
    ("2020-21", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "1", "Clause 7 Conduct Public/ SAPOL - Oppressive, offensive, abusive", "Fine/Recorded Reprimand/Transfer"),
    ("2020-21", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "2", "Clause 3 Conduct Prejudicial - Good Order and Discipline", "Recorded Reprimand"),
    ("2020-21", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "3", "Clause 6 Proper Exercise of Authority - Excessive Force", "Fine/Recorded Reprimand"),
    ("2020-21", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "4", "Clause 4 Performance of Orders / Duties - Failure to carry out a lawful order", "Transfer/Administrative Order - Penalty Officer"),
    ("2020-21", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "5", "Clause 3 Conduct Prejudicial - Good Order and Discipline; Clause 4 Performance of Orders/Duties - Failure to carry out a General Order", "Fine/Recorded Reprimand"),
    ("2020-21", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "7", "Clause 10 Confidentiality of Information - Improper Release/Access", "Fine/Recorded Reprimand"),
    ("2020-21", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "8", "Clause 3 Conduct Prejudicial - Good Order and Discipline; Clause 4 Performance of Orders / Duties - Disobey Orders", "Fine/Recorded Reprimand"),
    ("2020-21", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "9", "Clause 7 Conduct Public/SAPOL - Conduct towards employee", "Fine/Training and Education/Transfer"),
    ("2020-21", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "10", "Clause 5 Negligence - Neglect of Duty", "Fine/Recorded Reprimand"),
    ("2020-21", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "11", "Clause 10 Confidentiality of Information - Improper Release/Access", "Fine/Recorded Reprimand"),
    ("2020-21", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "12", "Clause 3 Conduct Prejudicial - Reflects Adversely", "Employment Terminated"),
    ("2020-21", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "13", "Clause 3 Conduct Prejudicial - Good Order and Discipline; Clause 11 Responsibility for Property - Fail to comply Property General Order", "Fine/Training and Education/Transfer/Administration Order - Penalty Officer"),
    ("2020-21", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "15", "Clause 10 Confidentiality of Information - Improper Release/Access", "Fine/Recorded Reprimand"),
    ("2020-21", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "16", "Clause 7 Conduct Public/SAPOL - Conduct towards public; Clause 10 Confidentiality of Information - Improper Release/Access", "Fine/Recorded Reprimand/Fine"),
    ("2020-21", "Code of Conduct - PCD Regulations 2017 (s26(1)(b)/(c))", "17", "Clause 2 Honesty and Integrity - Integrity", "Fine"),
    ("2020-21", "Code of Conduct - PCD Regulations 2014 (s26(1)(b)/(c))", "6", "Regulation 20 Confidentiality of Information - Improper Disclosure of Information (2 counts); Regulation 20 Confidentiality of Information - Improper Release/Access (6 counts)", "Fine/Recorded Reprimand"),
    ("2020-21", "Code of Conduct - PCD Regulations 2014 (s26(1)(b)/(c))", "14", "Regulation 13 Conduct Prejudicial - Good Order and Discipline", "Recorded Reprimand"),
    ("2019-20", "Code of Conduct - Police Regulations 2014", "6", "Reg 14 Conduct Prejudicial - Reflects Adversely", "Fine"),
    ("2019-20", "Code of Conduct - Police Regulations 2014", "7", "Reg 14 Conduct Prejudicial - Good Order & Discipline", "Reduction in Rank / Penalty - Administrative Order"),
    ("2019-20", "Code of Conduct - Police Regulations 2014", "8", "Reg 18 Conduct Public / SAPOL - Conduct towards Employee (three counts)", "Reduction in Rank / Transfer"),
    ("2019-20", "Code of Conduct - PCD Regulations 2017", "1", "Clause 3 Conduct Prejudicial - Reflects Adversely (three counts); Clause 4 Performance of Orders / Duties - Failure to carry out a lawful order; Clause 8 Conflict of Interest - Conflict of Interest", "Fine / Recorded Reprimand / Transfer"),
    ("2019-20", "Code of Conduct - PCD Regulations 2017", "2", "Clause 10 Confidentiality of Information - Improper Release / Access (five counts)", "Fine / Recorded Reprimand"),
    ("2019-20", "Code of Conduct - PCD Regulations 2017", "3", "Clause 10 Confidentiality of Information - Improper Release / Access (nine counts)", "Fine / Recorded Reprimand / Penalty - Administrative Order"),
    ("2019-20", "Code of Conduct - PCD Regulations 2017", "4", "Clause 8 Conflict of Interest - Conflict of Interest; Clause 10 Confidentiality of Information - Improper Release / Access", "Fine / Recorded Reprimand"),
    ("2019-20", "Code of Conduct - PCD Regulations 2017", "5", "Clause 10 Confidentiality of Information - Improper Release / Access (seven counts)", "Fine / Recorded Reprimand"),
    ("2019-20", "Code of Conduct - PCD Regulations 2017", "8", "Clause 4 Performance of Orders / Duties - Disobey Orders", "Reduction in Rank / Transfer"),
    ("2019-20", "Code of Conduct - PCD Regulations 2017", "9", "Clause 4 Performance of Orders / Duties - Failure to carry out a lawful order; Clause 7 Conduct Public/ SAPOL - Oppressive, offensive, abusive", "Fine / Recorded Reprimand / Penalty - Administrative Order"),
    ("2019-20", "Code of Conduct - PCD Regulations 2017", "10", "Clause 4 Performance of Orders/ Duties - Failure to carry out a lawful order (16 counts)", "Recorded Reprimand / Transfer"),
    ("2019-20", "Code of Conduct - PCD Regulations 2017", "11", "Clause 7 Conduct Public / SAPOL - Conduct towards Employee", "Penalty - Suspension without Pay"),
    ("2019-20", "Code of Conduct - PCD Regulations 2017", "12", "Clause 7 Conduct Public / SAPOL - Oppressive, offensive, abusive", "Transfer / Counselling"),
    ("2019-20", "Code of Conduct - PCD Regulations 2017", "13", "Clause 2 Honesty & Integrity - Falsehood / prevarication", "Fine / Recorded Reprimand"),
    ("2019-20", "Code of Conduct - PCD Regulations 2017", "14", "Clause 4 Performance of Orders / Duties - Failure to carry out a lawful order", "Fine / Recorded Reprimand / Transfer / Penalty - Administrative Order"),
    ("2019-20", "Code of Conduct - PCD Regulations 2017", "15", "Clause 9 Obtain benefit or advantage - Improperly seek / obtain benefit / advantage", "Fine / Penalty - Administrative Order"),
    ("2019-20", "Code of Conduct - PCD Regulations 2017", "16", "Clause 5 Negligence - Neglect of Duty", "Fine / Training and Education / Transfer / Penalty - Administrative Order"),
    ("2019-20", "Code of Conduct - PCD Regulations 2017", "17", "Criminal - Traffic", "Fine / Recorded Reprimand / Penalty - Administrative Order"),
    ("2019-20", "Code of Conduct - PCD Regulations 2017", "18", "Clause 2 Honesty & Integrity - Falsehood / prevarication", "Fine / Recorded Reprimand"),
    ("2019-20", "Code of Conduct - PCD Regulations 2017", "19", "Clause 10 Confidentiality of Information - Improper Disclosure of Information", "Fine / Penalty - Administrative Order"),
    ("2019-20", "Code of Conduct - PCD Regulations 2017", "20", "Criminal - Traffic", "Fine / Recorded Reprimand"),
    ("2019-20", "Code of Conduct - PCD Regulations 2017", "21", "Clause 10 Confidentiality of Information - Improper Release / Access", "Fine / Recorded Reprimand"),
    ("2018-19", "Code of Conduct - Police Regulations 1999", "1", "Criminal - Abuse of Public Office", "Reduction in Rank / Transfer / Fine"),
    ("2018-19", "Code of Conduct - Police Regulations 1999", "2", "Reg 20 Confidentiality of Information - Improper Release / Access", "Fine"),
    ("2018-19", "Code of Conduct - Police Regulations 2014", "2", "Reg 15 Performance of Orders / Duties - Disobey Orders; Reg 21 Confidentiality of Information - Improper Release / Access", "Fine"),
    ("2018-19", "Code of Conduct - Police Regulations 2014", "3", "Criminal - Traffic", "Fine / Recorded Reprimand"),
    ("2018-19", "Code of Conduct - Police Regulations 2014", "4", "Criminal - Traffic", "Fine / Recorded Reprimand"),
    ("2018-19", "Code of Conduct - Police Regulations 2014", "5", "Reg 14 Conduct Prejudicial - Reflects Adversely", "Fine"),
    ("2018-19", "Code of Conduct - Police Regulations 2014", "6", "Criminal - Traffic", "Fine / Recorded Reprimand"),
    ("2018-19", "Code of Conduct - Police Regulations 2014", "7", "Reg 15 Performance of Orders / Duties - Disobey Orders; Reg 21 Confidentiality of Information - Improper Release / Access", "Fine / Recorded Reprimand"),
    ("2018-19", "Code of Conduct - Police Regulations 2014", "9", "Reg 14 Conduct Prejudicial - Reflects Adversely", "Reduction in Rank / Transfer"),
    ("2018-19", "Code of Conduct - Police Regulations 2014", "11", "Reg 15 Performance of Orders / Duties - Disobey Orders", "Recorded Reprimand"),
    ("2018-19", "Code of Conduct - PCD Regulations 2017", "8", "Clause 6 Proper Exercise of Authority - Excessive Force", "Fine"),
    ("2018-19", "Code of Conduct - PCD Regulations 2017", "10", "Clause 4 Performance of Orders / Duties - Failure to carry out a lawful order; Clause 5 Negligence - Neglect of Duty", "Fine / Recorded Reprimand / Training and Education"),
    ("2017-18", "Breach of the Code of Conduct (4 Sept 2017 to 30 Jun 2018)", "9", "Police Regulations 1999 - Reg 16 Proper Exercise of Authority - Excessive Force; Police Regulations 2014 - Reg 17 Proper Exercise of Authority - Abuse of Authority", "Fine / Recorded Reprimand"),
    ("2017-18", "Breach of the Code of Conduct (4 Sept 2017 to 30 Jun 2018)", "10", "Police Regulations 1999 - Reg 15 Negligence - Neglect of Duty", "Fine / Recorded Reprimand"),
    ("2017-18", "Breach of the Code of Conduct (4 Sept 2017 to 30 Jun 2018)", "11", "Police Regulations 1999 - Reg 15 Negligence - Neglect of Duty", "Fine / Training and Education"),
    ("2017-18", "Breach of the Code of Conduct (4 Sept 2017 to 30 Jun 2018)", "12", "Police Regulations 1999 - Reg 14 Performance of Orders / Duties - Failure to Carry Out Lawful Order; Reg 15 Negligence - Neglect of Duty; Police Regulations 2014 - Reg 16 Negligence - Neglect of Duty", "Fine / Recorded Reprimand / Training and Education"),
    ("2017-18", "Breach of the Code of Conduct (4 Sept 2017 to 30 Jun 2018)", "13", "Police Regulations 2014 - Reg 14 Conduct Prejudicial - Reflects Adversely", "Unrecorded Reprimand / Counselling"),
    ("2017-18", "Breach of the Code of Conduct (4 Sept 2017 to 30 Jun 2018)", "14", "Police Regulations 2014 - Reg 21 Confidentiality of Information - Improper Release / Access; Reg 19 Conflict of Interest - Conflict of Interest", "Fine / Recorded Reprimand"),
    ("2017-18", "Breach of the Code of Conduct (4 Sept 2017 to 30 Jun 2018)", "15", "Police Regulations 2014 - Reg 13 Honesty & Integrity - Honesty; Reg 21 Confidentiality of Information - Improper Release / Access (2 counts); Reg 13 Honesty & Integrity - Integrity", "Transfer"),
    ("2017-18", "Breach of the Code of Conduct (4 Sept 2017 to 30 Jun 2018)", "16", "Police Regulations 2014 - Reg 14 Conduct Prejudicial - Reflects Adversely", "Fine / Recorded Reprimand"),
    ("2017-18", "Breach of the Code of Conduct (4 Sept 2017 to 30 Jun 2018)", "17", "Police Regulations 2014 - Reg 14 Conduct Prejudicial - Good Order & Discipline; Reg 14 Conduct Prejudicial - Reflects Adversely", "Reduction in Rank / Transfer / Fine"),
    ("2017-18", "Breach of the Code of Conduct (4 Sept 2017 to 30 Jun 2018)", "18", "Police Regulations 2014 - Criminal - Traffic", "Fine / Recorded Reprimand"),
    ("2017-18", "Breach of the Code of Conduct (4 Sept 2017 to 30 Jun 2018)", "19", "Police Regulations 2014 - Criminal - Offence against Person", "Unrecorded Reprimand"),
]
write_csv(
    "pcd_act_sanctions_imposed_on_sapol_officers.csv",
    ["reporting_period", "breach_category", "case_reference_no", "breach_description", "sanction_outcome"],
    sanctions,
)

print("done")
