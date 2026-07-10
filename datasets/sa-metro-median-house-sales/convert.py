"""
Downloads all quarterly resources from DHUD's "Metro median house sales" CKAN
dataset into raw/, then merges them into data/metro_median_house_sales.csv.

Each quarterly file reports its own quarter plus a same-quarter-prior-year
comparison column. Consecutive files' comparison columns duplicate the
previous year's own "current quarter" file, so this script prefers the
row from whichever file the quarter was the *primary* (current) report in,
and only falls back to a comparison column for December quarter 2014, which
has no dedicated file of its own.
"""

import csv
import json
import re
import subprocess
from collections import OrderedDict
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
RAW_DIR = HERE / "raw"
DATA_DIR = HERE / "data"
RAW_DIR.mkdir(parents=True, exist_ok=True)
DATA_DIR.mkdir(parents=True, exist_ok=True)

PACKAGE_URL = "https://data.sa.gov.au/data/api/3/action/package_show?id=metro-median-house-sales"

CITY_CANONICAL = {
    "NORWD PAYNM ST PET": "NORWOOD PAYNEHAM & ST PETERS",
    "PORT ADEL ENFIELD": "PORT ADELAIDE ENFIELD",
}

QUARTER_RE = re.compile(r"Q(\d)\s*(\d{4})$")
NORM_WS_RE = re.compile(r"\s+")
HDR_RE = re.compile(r"^(Sales|Median)\s+(\d)Q\s*(\d{4})$")


def fetch_package():
    out = subprocess.run(["curl", "-s", PACKAGE_URL], capture_output=True, text=True, check=True)
    return json.loads(out.stdout)["result"]["resources"]


def slug_for(name, url):
    m = QUARTER_RE.search(name)
    q, y = m.group(1), m.group(2)
    ext = url.rsplit(".", 1)[-1].lower()
    if ext not in ("xlsx", "xls", "csv"):
        ext = "xlsx"
    return f"metro-median-house-sales-{y}-q{q}.{ext}", int(q), int(y)


def normalize_header(h):
    if h is None:
        return ""
    return NORM_WS_RE.sub(" ", str(h).replace("\n", " ")).strip()


def classify_header(h):
    n = normalize_header(h)
    m = HDR_RE.match(n)
    if m:
        return m.group(1).lower(), int(m.group(2)), int(m.group(3))
    if n.lower().startswith("median change"):
        return "change", None, None
    if n.lower() == "city":
        return "city", None, None
    if n.lower() == "suburb":
        return "suburb", None, None
    return None, None, None


def clean_number(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip().replace(",", "")
    if s == "":
        return None
    is_pct = s.endswith("%")
    if is_pct:
        s = s[:-1]
    try:
        num = float(s)
    except ValueError:
        return v
    if is_pct:
        num = num / 100.0
    if num == int(num):
        num = int(num)
    return num


def load_rows(path):
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        return list(ws.iter_rows(min_row=1, values_only=True))
    except Exception:
        with open(path, newline="", encoding="utf-8-sig") as f:
            return [tuple(row) for row in csv.reader(f)]


def main():
    resources = fetch_package()
    all_rows = []
    manifest_lines = []

    for r in resources:
        fname, cur_q, cur_y = slug_for(r["name"], r["url"])
        out_path = RAW_DIR / fname
        if not out_path.exists():
            subprocess.run(["curl", "-sL", r["url"], "-o", str(out_path)], check=True)
        manifest_lines.append(f"{fname}\t{out_path.stat().st_size} bytes\t{r['url']}")

        rows_iter = load_rows(out_path)
        header = rows_iter[0]
        col_map = {i: classify_header(h) for i, h in enumerate(header)}
        col_map = {i: v for i, v in col_map.items() if v[0]}

        city_idx = next(i for i, v in col_map.items() if v[0] == "city")
        suburb_idx = next(i for i, v in col_map.items() if v[0] == "suburb")
        change_idx = next((i for i, v in col_map.items() if v[0] == "change"), None)

        quarters_present = sorted(
            {(v[1], v[2]) for v in col_map.values() if v[0] in ("sales", "median")},
            key=lambda t: (t[1], t[0]),
        )

        for q, y in quarters_present:
            sales_idx = next((i for i, v in col_map.items() if v == ("sales", q, y)), None)
            median_idx = next((i for i, v in col_map.items() if v == ("median", q, y)), None)
            is_primary = q == cur_q and y == cur_y

            for row in rows_iter[1:]:
                if row is None or len(row) <= max(city_idx, suburb_idx):
                    continue
                city = (str(row[city_idx]).strip() if row[city_idx] is not None else "")
                suburb = (str(row[suburb_idx]).strip() if row[suburb_idx] is not None else "")
                city = CITY_CANONICAL.get(city, city)
                if not city and not suburb:
                    continue

                sales = clean_number(row[sales_idx]) if sales_idx is not None and sales_idx < len(row) else None
                median = clean_number(row[median_idx]) if median_idx is not None and median_idx < len(row) else None
                change = None
                if is_primary and change_idx is not None and change_idx < len(row):
                    change = clean_number(row[change_idx])

                all_rows.append({
                    "year": y, "quarter": q, "quarter_label": f"{y}-Q{q}",
                    "city": city, "suburb": suburb,
                    "sales_count": sales, "median_price_aud": median,
                    "median_change_yoy_pct": change,
                    "source_file": fname, "is_primary": is_primary,
                })

    dedup = OrderedDict()
    for row in all_rows:
        key = (row["year"], row["quarter"], row["city"], row["suburb"])
        if key not in dedup or (row["is_primary"] and not dedup[key]["is_primary"]):
            dedup[key] = row

    final_rows = sorted(dedup.values(), key=lambda r: (r["year"], r["quarter"], r["city"], r["suburb"]))

    out_csv = DATA_DIR / "metro_median_house_sales.csv"
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["year", "quarter", "quarter_label", "city", "suburb", "sales_count",
                    "median_price_aud", "median_change_yoy_pct", "reported_in_file"])
        for row in final_rows:
            w.writerow([row["year"], row["quarter"], row["quarter_label"], row["city"], row["suburb"],
                        row["sales_count"], row["median_price_aud"], row["median_change_yoy_pct"],
                        row["source_file"]])

    (RAW_DIR / "MANIFEST.txt").write_text("\n".join(manifest_lines) + "\n")
    print(f"wrote {out_csv} ({len(final_rows)} rows)")


if __name__ == "__main__":
    main()
