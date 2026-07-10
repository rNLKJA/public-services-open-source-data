#!/usr/bin/env python3
"""
Converts raw/ CBS occupational licence source files into tidy, ready-to-use CSVs.

Two source families:
1. occupational-licences-september-2018.xlsx - the full licensee-level register
   (194,620 records). Split into three tidy tables and redacted per this
   repository's standing "no individual-identifying fields" rule (see README's
   Privacy check section): individual licensee names are withheld, and street-
   level address detail is dropped for every row (kept at suburb/postcode/state
   only), regardless of entity type.
2. annual-report-cbs-*.csv - six CBS Annual Report tables (2014-15 to 2017-18
   aggregate counts by Act), each with its own layout. Merged into one tidy
   long-format table.
"""
import csv
import datetime
import html
import re
from pathlib import Path

RAW = Path(__file__).parent.parent / "raw"
OUT = Path(__file__).parent

REDACTED = "[individual licensee - name withheld for privacy]"

JURISDICTIONS = {
    "BLD": ("Building Work Contractors Act 1995", "building work contractors and supervisors"),
    "PGE": ("Plumbers, Gas fitters and Electricians Act 1995", "plumbers, gas fitters and electricians"),
    "ISL": ("Security and Investigation Industry Act 1995", "security and investigation agents"),
    "MVD": ("Second-hand Vehicle Dealers Act 1995", "second-hand motor vehicle/motor cycle dealers"),
    "RCO": ("Conveyancers Act 1994", "registered conveyancers"),
    "RLA": ("Land Agents Act 1994", "registered land agents"),
    "RSR": ("Land Agents Act 1994", "registered sales representatives and auctioneers"),
}

ENTITY_TYPES = {
    "P": "Individual person",
    "CO": "Body corporate (company)",
    "IN": "Incorporated association",
}

# Best-effort inference only - the source's own metadata DOC file (linked from
# the CKAN record) 404s, and CBS's public licence-check page blocked automated
# fetch (403). Not confirmed against an official code dictionary; disclosed as
# such in the README rather than presented as fact.
LICENCE_STATUS = {
    "L": "Current (licensed) [inferred, unconfirmed]",
    "C": "Cancelled [inferred, unconfirmed]",
    "SR": "Surrendered [inferred, unconfirmed]",
    "SP": "Suspended [inferred, unconfirmed]",
}

ADDR_RE = re.compile(r"<Suburb>(.*?)</Suburb>|<Postcode>(.*?)</Postcode>|<State>(.*?)</State>")
CATSUBCAT_RE = re.compile(
    r"<CatSubCat><CatCode>(.*?)</CatCode><CondCode>(.*?)</CondCode>"
    r"<Details>(.*?)</Details><ItemType>(.*?)</ItemType><SubCatCode>(.*?)</SubCatCode></CatSubCat>"
)
CONDS_RE = re.compile(
    r"<Conds><CatCode>(.*?)</CatCode><CondCode>(.*?)</CondCode>"
    r"<Details>(.*?)</Details><ItemType>(.*?)</ItemType><SubCatCode>(.*?)</SubCatCode></Conds>"
)


def parse_address(raw):
    if not raw or raw == "NULL":
        return "", "", ""
    suburb = postcode = state = ""
    for m in ADDR_RE.finditer(raw):
        if m.group(1) is not None:
            suburb = m.group(1)
        elif m.group(2) is not None:
            postcode = m.group(2)
        elif m.group(3) is not None:
            state = m.group(3)
    return suburb, postcode, state


def convert_register():
    import openpyxl

    wb = openpyxl.load_workbook(RAW / "occupational-licences-september-2018.xlsx", read_only=True)
    ws = wb[wb.sheetnames[0]]

    licences_f = open(OUT / "occupational_licences.csv", "w", newline="", encoding="utf-8")
    categories_f = open(OUT / "occupational_licence_categories.csv", "w", newline="", encoding="utf-8")
    conditions_f = open(OUT / "occupational_licence_conditions.csv", "w", newline="", encoding="utf-8")

    licences_w = csv.writer(licences_f)
    categories_w = csv.writer(categories_f)
    conditions_w = csv.writer(conditions_f)

    licences_w.writerow([
        "record_id", "licence_number", "jurisdiction_code", "act", "jurisdiction_desc",
        "entity_type_code", "entity_type_desc", "licensee_name", "date_granted",
        "licence_status_code", "licence_status_desc", "address_suburb",
        "address_postcode", "address_state", "has_conditions",
    ])
    categories_w.writerow([
        "record_id", "licence_number", "jurisdiction_code", "item_type",
        "category_code", "category_description", "condition_code", "subcategory_code",
    ])
    conditions_w.writerow([
        "record_id", "licence_number", "jurisdiction_code", "condition_code", "condition_detail",
    ])

    record_id = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        record_id += 1
        (licensee_name, jur_code, licence_number, entity_type, licence_status,
         date_granted, service_address, catsubcats, conds) = row

        act, jur_desc = JURISDICTIONS.get(jur_code, ("", ""))
        entity_desc = ENTITY_TYPES.get(entity_type, entity_type)
        status_desc = LICENCE_STATUS.get(licence_status, licence_status)

        name_out = REDACTED if entity_type == "P" else (licensee_name or "")

        suburb, postcode, state = parse_address(service_address)

        granted_out = ""
        if isinstance(date_granted, datetime.datetime):
            granted_out = date_granted.date().isoformat()
        elif date_granted:
            granted_out = str(date_granted)

        has_conditions = bool(conds and conds != "NULL")

        licences_w.writerow([
            record_id, licence_number, jur_code, act, jur_desc,
            entity_type, entity_desc, name_out, granted_out,
            licence_status, status_desc, suburb, postcode, state,
            "true" if has_conditions else "false",
        ])

        if catsubcats and catsubcats != "NULL":
            for m in CATSUBCAT_RE.finditer(catsubcats):
                cat_code, cond_code, details, item_type, subcat_code = m.groups()
                categories_w.writerow([
                    record_id, licence_number, jur_code, item_type,
                    cat_code, html.unescape(details), cond_code, subcat_code,
                ])

        if has_conditions:
            for m in CONDS_RE.finditer(conds):
                cat_code, cond_code, details, item_type, subcat_code = m.groups()
                conditions_w.writerow([record_id, licence_number, jur_code, cond_code, html.unescape(details)])

    licences_f.close()
    categories_f.close()
    conditions_f.close()
    print(f"Wrote {record_id} licence records")


def _clean_num(v):
    v = (v or "").strip()
    if v in ("", "#"):
        return "", ("not available - reporting systems changed in 2015-16" if v == "#" else "")
    return v, ""


def convert_annual_reports():
    rows_out = []

    def emit(act, category, metric, year_values, notes=""):
        for year, val in year_values.items():
            value, note = _clean_num(val)
            rows_out.append([act, category, metric, year, value, note or notes])

    # Building Work Contractors Act 1995
    with open(RAW / "annual-report-cbs-building-work-contractors-act.csv", encoding="cp1252") as f:
        r = list(csv.reader(f))
    act = "Building Work Contractors Act 1995"
    years = ["2017-18", "2016-17", "2015-16", "2014-15"]
    for line in r[3:7]:
        emit(act, "", line[0].strip(), dict(zip(years, line[1:5])))
    emit(act, "", r[7][0].strip(), dict(zip(years, r[7][1:5])))

    # Conveyancers Act 1994
    with open(RAW / "annual-report-cbs-conveyancers-act.csv", encoding="cp1252") as f:
        r = list(csv.reader(f))
    act = "Conveyancers Act 1994"
    years = ["2017-18", "2016-17", "2015-16", "2014-15"]
    for line in r[3:6]:
        emit(act, "", line[0].strip(), dict(zip(years, line[1:5])))
    emit(act, "", r[6][0].strip(), dict(zip(years, r[6][1:5])))

    # Land Agents Act 1994
    with open(RAW / "annual-report-cbs-land-agents-act.csv", encoding="cp1252") as f:
        r = list(csv.reader(f))
    act = "Land Agents Act 1994"
    years = ["2017-18", "2016-17", "2015-16"]
    for line in r[3:7]:
        emit(act, "", line[0].strip(), dict(zip(years, line[1:4])))
    emit(act, "", r[7][0].strip(), dict(zip(years, r[7][1:4])))

    # Plumbers, Gas fitters and Electricians Act 1995 (3 sub-tables)
    with open(RAW / "annual-report-cbs-plumbers-gas-fitters-electricians-act.csv", encoding="cp1252") as f:
        r = list(csv.reader(f))
    act = "Plumbers, Gas fitters and Electricians Act 1995"
    years = ["2017-18", "2016-17", "2015-16"]
    # rows: 2 header, then [category header + 3 metric rows] x 3 categories
    idx = 2
    for _ in range(3):
        category = r[idx][0].strip()
        idx += 1
        for _ in range(3):
            emit(act, category, r[idx][0].strip(), dict(zip(years, r[idx][1:4])))
            idx += 1

    # Second-hand Vehicle Dealers Act 1995
    with open(RAW / "annual-report-cbs-second-hand-vehicle-dealer-act.csv", encoding="cp1252") as f:
        r = list(csv.reader(f))
    act = "Second-hand Vehicle Dealers Act 1995"
    years = ["2017-18", "2016-17", "2015-16"]
    emit(act, "Held by bodies corporate", r[4][0].strip(), dict(zip(years, r[4][1:4])))
    emit(act, "Held by bodies corporate", r[5][0].strip(), dict(zip(years, r[5][1:4])))
    emit(act, "Held by natural persons", r[7][0].strip(), dict(zip(years, r[7][1:4])))
    emit(act, "Held by natural persons", r[8][0].strip(), dict(zip(years, r[8][1:4])))
    emit(act, "", r[9][0].strip(), dict(zip(years, r[9][1:4])))
    emit(act, "Applications for new licences/registrations received", r[11][0].strip(), dict(zip(years, r[11][1:4])))
    emit(act, "Applications for new licences/registrations received", r[12][0].strip(), dict(zip(years, r[12][1:4])))
    emit(act, "", r[13][0].strip(), dict(zip(years, r[13][1:4])))

    # Security and Investigation Industry Act 1995
    with open(RAW / "annual-report-cbs-security-and-investigation-act.csv", encoding="cp1252") as f:
        r = list(csv.reader(f))
    act = "Security and Investigation Industry Act 1995"
    years = ["2017-18", "2016-17", "2015-16"]
    for line in r[3:6]:
        emit(act, "", line[0].strip(), dict(zip(years, line[1:4])))
    emit(act, "", r[6][0].strip(), dict(zip(years, r[6][1:4])))

    with open(OUT / "annual_report_by_act.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["act", "category", "metric", "year", "value", "note"])
        w.writerows(rows_out)
    print(f"Wrote {len(rows_out)} annual-report rows")


if __name__ == "__main__":
    convert_register()
    convert_annual_reports()
