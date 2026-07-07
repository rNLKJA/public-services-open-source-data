#!/usr/bin/env python3
"""Convert DPTI/DHUD SA population projection workbooks into tidy long-format CSVs.

Source files (raw/):
  - sa-lga-population-projections-2016-2036.xlsx  (LGA-level, 2016 Census base)
  - sa-sa2-population-projections-2016-2036.xlsx  (SA2-level, 2016 Census base)
  - sa-and-regions-population-projections-medium-series-2016-2041.xls (State + 11 Population
    Projection Regions, Medium series only)

Each workbook mixes a genuine LGA/SA2 aggregate ("TOTAL SOUTH AUSTRALIA" or
"TOTAL <region>") into the same column layout as individual areas, using a blank
code cell to mark it. That's preserved here as an explicit `is_total_row` flag
rather than dropped or left ambiguous.
"""
import csv
import openpyxl
import xlrd

RAW = "raw"
DATA = "data"


def write_csv(path, header, rows):
    with open(f"{DATA}/{path}", "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)
    print(f"wrote {path}: {len(rows)} rows")


def lga_total():
    wb = openpyxl.load_workbook(f"{RAW}/sa-lga-population-projections-2016-2036.xlsx", data_only=True)
    ws = wb["Projected population"]
    years = [2016, 2021, 2026, 2031, 2036]
    rows = []
    for row in ws.iter_rows(min_row=11, values_only=True):
        if all(v is None for v in row):
            continue
        lga_code, lga_name, sex = row[0], row[1], row[2]
        if lga_name is None:
            continue
        is_total = lga_code is None
        for i, year in enumerate(years):
            pop = row[3 + i]
            rows.append([lga_code, lga_name, sex, year, pop, is_total])
    write_csv(
        "sa-lga-population-projections-total.csv",
        ["lga_code", "lga_name", "sex", "year", "population", "is_total_row"],
        rows,
    )
    # spot check: Adelaide (C), Persons, 2016 == 23552 (from source row inspected directly)
    check = [r for r in rows if r[0] == 40070 and r[2] == "Persons" and r[3] == 2016]
    assert check and check[0][4] == 23552, f"Adelaide (C) 2016 spot-check failed: {check}"
    # spot check: TOTAL SOUTH AUSTRALIA, Females, 2016 == 865966
    check2 = [r for r in rows if r[5] and r[2] == "Females" and r[3] == 2016]
    assert check2 and check2[0][4] == 865966, f"TOTAL SA 2016 Females spot-check failed: {check2}"


def lga_by_age():
    wb = openpyxl.load_workbook(f"{RAW}/sa-lga-population-projections-2016-2036.xlsx", data_only=True)
    rows = []
    for year in (2016, 2021, 2026, 2031, 2036):
        ws = wb[str(year)]
        header_row = [c.value for c in ws[10]]
        age_labels = [str(v).strip() for v in header_row[4:23]]  # 19 five-year age groups
        for row in ws.iter_rows(min_row=11, values_only=True):
            if all(v is None for v in row):
                continue
            lga_code, lga_name, sex = row[1], row[2], row[3]
            if lga_name is None:
                continue
            is_total = lga_code is None
            age_values = row[4:23]  # excludes the trailing 'Total' column (redundant with the total file)
            for age_label, pop in zip(age_labels, age_values):
                rows.append([lga_code, lga_name, sex, age_label, year, pop, is_total])
    write_csv(
        "sa-lga-population-projections-by-age.csv",
        ["lga_code", "lga_name", "sex", "age_group", "year", "population", "is_total_row"],
        rows,
    )
    # spot check: Adelaide (C), Persons, 2016, 0-4 age group == 584 (from source row inspected directly)
    check = [r for r in rows if r[0] == 40070 and r[2] == "Persons" and r[3] == "0-4" and r[4] == 2016]
    assert check and check[0][5] == 584, f"Adelaide (C) 2016 0-4 spot-check failed: {check}"


def sa2_total():
    wb = openpyxl.load_workbook(f"{RAW}/sa-sa2-population-projections-2016-2036.xlsx", data_only=True)
    ws = wb["Projected population"]
    years = [2016, 2021, 2026, 2031, 2036]
    rows = []
    for row in ws.iter_rows(min_row=11, values_only=True):
        if all(v is None for v in row):
            continue
        region, sa2_code, sa2_name, sex = row[0], row[1], row[2], row[3]
        if sa2_name is None:
            continue
        is_total = sa2_code is None
        for i, year in enumerate(years):
            pop = row[4 + i]
            rows.append([region, sa2_code, sa2_name, sex, year, pop, is_total])
    write_csv(
        "sa-sa2-population-projections-total.csv",
        ["region_code", "sa2_code", "sa2_name", "sex", "year", "population", "is_total_row"],
        rows,
    )
    check = [r for r in rows if r[6] and r[3] == "Females" and r[4] == 2016 and r[2] == "TOTAL SOUTH AUSTRALIA"]
    assert check and check[0][5] == 865966, f"SA2 TOTAL SA 2016 Females spot-check failed: {check}"


def regional_medium_series():
    wb = xlrd.open_workbook(f"{RAW}/sa-and-regions-population-projections-medium-series-2016-2041.xls")
    ws = wb.sheet_by_name("Summary")
    years = [int(y) for y in ws.row_values(8)[2:]]
    rows = []
    for r in range(10, ws.nrows):
        vals = ws.row_values(r)
        if not vals[1]:
            continue
        region = vals[1]
        is_total = region == "SOUTH AUSTRALIA"
        for year, pop in zip(years, vals[2:]):
            rows.append([region, year, int(pop), is_total])
    write_csv(
        "sa-regional-population-projections-medium-series.csv",
        ["region", "year", "population_persons", "is_total_row"],
        rows,
    )
    check = [r for r in rows if r[3] and r[1] == 2016]
    assert check and check[0][2] == 1712844, f"Regional TOTAL SA 2016 spot-check failed: {check}"


if __name__ == "__main__":
    lga_total()
    lga_by_age()
    sa2_total()
    regional_medium_series()
    print("All spot-checks passed.")
