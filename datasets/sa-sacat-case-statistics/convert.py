"""Merge SACAT's per-financial-year annual report CSV/XLSX exports into tidy, multi-year tables.

Source: 3 annual "Annual Report" CKAN packages (2020-21, 2021-22, 2022-23) on data.sa.gov.au, each
publishing the same 10 categories as separate files. Most categories are already tidy per-row CSVs
(the source repeats FinYear/FinYearTotal on every data row); four files are instead pivot-table
XLSX exports with a Fin Year -> Act -> Application row hierarchy, where every real data row is
immediately followed by a duplicate "subtotal" row (blank labels, same count) -- an artifact of
the reporting tool's export, not a second observation. Those four are reconstructed here into the
same tidy shape as their CSV-format siblings (confirmed against the other two years' CSVs for the
same category) before merging all three years together.

No count or percentage is recalculated, aggregated further or reinterpreted -- every number is
carried through exactly as published. `financial_year` is normalised from the source's `22-23`
style to `2022-23` for consistency with the rest of this repository; nothing else is renamed
except column headers (CamelCase -> snake_case).
"""

import csv
from pathlib import Path

import openpyxl

RAW = Path(__file__).parent / "raw"
DATA = Path(__file__).parent / "data"
YEARS = ["2020-21", "2021-22", "2022-23"]


def fy_long(fy_short):
    """'22-23' -> '2022-23'"""
    a, b = fy_short.split("-")
    return f"20{a}-{b}"


def read_csv_rows(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        header = next(reader)
        return header, [row for row in reader if any(cell.strip() for cell in row)]


RENAMES = {
    "FinYear": "financial_year",
    "FinYearTotal": "financial_year_total",
    "FinYearTotalResolved": "financial_year_total_resolved",
    "FinYearPercentage": "financial_year_percentage_resolved",
    "Stream": "stream",
    "StreamTotal": "stream_total",
    "StreamTotalResolved": "stream_total_resolved",
    "StreamPercentage": "stream_percentage_resolved",
    "Act": "act",
    "ActTotal": "act_total",
    "ActTotalResolved": "act_total_resolved",
    "ActPercentage": "act_percentage_resolved",
    "Application": "application",
    "ApplicationTotal": "application_total",
    "ApplicationTotalResolved": "application_total_resolved",
    "ApplicationPercentage": "application_percentage_resolved",
    "ApplicationType": "application_type",
    "HearingType": "hearing_type",
    "Type": "type",
    "TypeTotal": "type_total",
    "Percentage": "percentage",
    "Total": "count",
    "ResolvedConference": "resolved_at_conference",
}


def load_csv_category(category, year):
    header, rows = read_csv_rows(RAW / year / f"{category}.csv")
    header = [RENAMES[h] for h in header]
    fy_idx = header.index("financial_year")
    out = []
    for row in rows:
        row = list(row)
        row[fy_idx] = fy_long(row[fy_idx])
        out.append(row)
    return header, out


def nonblank_rows(ws):
    rows = list(ws.iter_rows(values_only=True))
    return [r for r in rows[2:] if any(c not in (None, "") for c in r)]  # skip title + header row


def parse_single_level_pivot(path, count_col, pct_col=None):
    """Fin Year / Type(-ish label) / count [/ percentage] pivot, e.g. community-hearings,
    admin-disc-hearings. Returns rows of (financial_year, financial_year_total, label, count,
    [percentage])."""
    ws = openpyxl.load_workbook(path).active
    fy = fy_total = None
    out = []
    for r in nonblank_rows(ws):
        if r[0]:
            fy = r[0]
            fy_total = r[count_col]
            continue
        label = r[1]
        if not label:
            continue  # duplicate blank subtotal artifact row
        count = r[count_col]
        row = [fy_long(fy), fy_total, label, count]
        if pct_col is not None:
            row.append(r[pct_col])
        row.append(count)  # trailing Total column, always equal to the leaf count in this source
        out.append(row)
    return out


def parse_two_level_pivot(path):
    """Fin Year / Act / Application / Total pivot, e.g. housing-applications,
    community-applications. Returns rows of (financial_year, financial_year_total, act,
    act_total, application, count)."""
    ws = openpyxl.load_workbook(path).active
    fy = fy_total = None
    act = act_total = None
    out = []
    for r in nonblank_rows(ws):
        if r[0]:
            fy, fy_total = r[0], r[-1]
            continue
        act_label, app_label, total = r[1], r[2], r[-1]
        if act_label and not app_label:
            act, act_total = act_label, total
            continue
        if not act_label and not app_label:
            continue  # duplicate blank subtotal artifact row
        out.append([fy_long(fy), fy_total, act, act_total, app_label, total, total])
    return out


# (year, category) -> reconstruction function producing rows already in canonical column order
XLSX_RECONSTRUCT = {
    ("2021-22", "community-hearings"): lambda p: parse_single_level_pivot(p, count_col=2, pct_col=3),
    ("2022-23", "admin-disc-hearings"): lambda p: parse_single_level_pivot(p, count_col=2),
    ("2021-22", "housing-applications"): parse_two_level_pivot,
    ("2022-23", "community-applications"): parse_two_level_pivot,
}

# Canonical (renamed) column order per category, matching the CSV-format years exactly.
CANONICAL_HEADERS = {
    "admin-disc-applications": ["financial_year", "financial_year_total", "application", "type_total", "percentage", "count"],
    "admin-disc-hearings": ["financial_year", "financial_year_total", "hearing_type", "type_total", "count"],
    "adr": ["financial_year", "financial_year_total", "financial_year_total_resolved", "financial_year_percentage_resolved",
            "stream", "stream_total", "stream_total_resolved", "stream_percentage_resolved",
            "act", "act_total", "act_total_resolved", "act_percentage_resolved",
            "application", "application_total", "application_total_resolved", "application_percentage_resolved",
            "count", "resolved_at_conference"],
    "community-applications": ["financial_year", "financial_year_total", "act", "act_total", "application", "type_total", "count"],
    "community-auto-reviews": ["financial_year", "financial_year_total", "type", "type_total", "count"],
    "community-hearings": ["financial_year", "financial_year_total", "type", "type_total", "percentage", "count"],
    "housing-applications": ["financial_year", "financial_year_total", "act", "act_total", "application", "type_total", "count"],
    "housing-hearings": ["financial_year", "financial_year_total", "application", "type_total", "percentage", "count"],
    "internal-review-applications": ["financial_year", "financial_year_total", "application_type", "type_total", "count"],
    "internal-review-hearings": ["financial_year", "financial_year_total", "hearing_type", "type_total", "count"],
}

OUT_NAMES = {
    "admin-disc-applications": "administrative-disciplinary-applications.csv",
    "admin-disc-hearings": "administrative-disciplinary-hearings.csv",
    "adr": "alternative-dispute-resolution.csv",
    "community-applications": "community-list-applications.csv",
    "community-auto-reviews": "community-list-automatic-reviews.csv",
    "community-hearings": "community-list-hearings.csv",
    "housing-applications": "housing-list-applications.csv",
    "housing-hearings": "housing-list-hearings.csv",
    "internal-review-applications": "internal-review-applications.csv",
    "internal-review-hearings": "internal-review-hearings.csv",
}


def merge_category(category):
    canonical = CANONICAL_HEADERS[category]
    merged = []
    for year in YEARS:
        reconstruct = XLSX_RECONSTRUCT.get((year, category))
        if reconstruct:
            rows = reconstruct(RAW / year / f"{category}.xlsx")
            merged.extend(rows)
        else:
            header, rows = load_csv_category(category, year)
            assert header == canonical, f"{category}/{year} header mismatch: {header} != {canonical}"
            merged.extend(rows)

    DATA.mkdir(exist_ok=True)
    out_path = DATA / OUT_NAMES[category]
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(canonical)
        writer.writerows(merged)
    print(f"wrote {out_path.name}: {len(merged)} rows")


if __name__ == "__main__":
    for category in CANONICAL_HEADERS:
        merge_category(category)
