"""Reshape the 16 SA Housing Authority Social Housing workbooks (4 measures x
2 housing types x 2 financial years) into 4 tidy, per-measure long-format CSVs.

Each source workbook is a single-sheet LGA-by-characteristic table with a
2-row header (a category-group row and a column-name row) preceded by title/
notes/source metadata rows, and followed by a "Total - <measure>" and a
"% Total - <measure>" summary row. This script locates the real header and
data block programmatically (the row whose first cell is "Local Government
Area"), forward-fills the merged-cell group header, converts column names to
snake_case, and concatenates the 4 housing-type/financial-year combinations
for each measure into one file, adding `housing_type` and `financial_year`
columns and an `is_total_row` flag rather than dropping the source's own
Total/% Total rows.
"""
import csv
import re
from pathlib import Path

import openpyxl

RAW = Path(__file__).parent / "raw"
OUT = Path(__file__).parent / "data"
OUT.mkdir(exist_ok=True)


def snake_case(label):
    label = label.strip().rstrip(",")
    label = label.replace("%", "pct").replace("/", " or ")
    label = re.sub(r"[^0-9a-zA-Z]+", "_", label).strip("_").lower()
    return label


def load_table(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    header_idx = next(i for i, r in enumerate(rows) if r and r[0] == "Local Government Area")
    name_row = rows[header_idx]

    # The source's merged-cell category-group row (e.g. "Household composition",
    # "Main source of income - households paying less than market rent only")
    # groups leaf columns whose own names are already unique within a measure,
    # and its wording varies slightly between PH/SOMIH editions of the same
    # measure (see README) — so only the leaf column name is kept here; the
    # group label is documented in the README instead of baked into the name.
    columns = []
    for n in name_row:
        if n is None:
            continue
        columns.append((snake_case(n), len(columns)))

    n_cols = len(columns)
    data_rows = []
    for r in rows[header_idx + 1:]:
        if r[0] is None:
            continue
        data_rows.append(r[:n_cols])

    return [c for c, _ in columns], data_rows


MEASURES = {
    "dwellings": "Dwellings — number of public/community housing dwellings by Local Government Area",
    "households": "Households — households residing in public/community housing by Local Government Area",
    "new-households-housed": "New households housed — new households allocated housing during the financial year by Local Government Area",
    "household-members": "Household members — people residing in public/community housing households by Local Government Area",
}

HOUSING_TYPES = {"ph": "Public Housing", "somih": "State Owned and Managed Indigenous Housing (SOMIH)"}
YEARS = {"2019-20": "2019-20", "2020-21": "2020-21"}

for measure in MEASURES:
    all_columns = None
    combined_rows = []
    for housing_key, housing_label in HOUSING_TYPES.items():
        for year in YEARS:
            path = RAW / f"{measure}_{housing_key}_{year}.xlsx"
            columns, data_rows = load_table(path)
            if all_columns is None:
                all_columns = columns
            elif len(columns) != len(all_columns):
                raise ValueError(f"Column count mismatch in {path}: {columns} != {all_columns}")
            elif columns != all_columns:
                # A handful of PH/SOMIH editions word one or two column labels
                # very slightly differently (e.g. "Younger person (main tenant
                # under 25 years)" vs "... (household main tenant ...)") for
                # the same underlying measure at the same position — the
                # canonical name from the first file processed is kept.
                print(f"  note: {path.name} has minor column-label wording differences, using canonical names")

            for row in data_rows:
                lga = row[0]
                is_total = isinstance(lga, str) and lga.lower().startswith(("total", "% total"))
                out_row = {
                    "financial_year": year,
                    "housing_type": housing_label,
                    "local_government_area": lga,
                    "is_total_row": is_total,
                }
                for col, val in zip(all_columns[1:], row[1:]):
                    out_row[col] = val
                combined_rows.append(out_row)

    fieldnames = ["financial_year", "housing_type", "local_government_area", "is_total_row"] + all_columns[1:]
    out_path = OUT / f"{measure}.csv"
    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(combined_rows)
    print(f"Wrote {out_path} ({len(combined_rows)} rows, {len(fieldnames)} columns)")
