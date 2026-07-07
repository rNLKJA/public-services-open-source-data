#!/usr/bin/env python3
"""Convert ABS Government Finance Statistics, Annual, 2024-25 workbooks
(South Australia general government, Table 234; all-states total, Table 239)
into tidy long-format CSVs.

Source layout per sheet: 6 header rows (title/source/release-date/table-name,
then a year row, then a $m unit row), then a mix of:
  - header-only rows (label, no values) marking a section/category
  - "less" / "equals" / "plus" marker rows with no values, used purely to show
    the additive structure between sections (Revenue less Expenses equals Net
    Operating Balance, etc.) - these carry no figures and are dropped
  - data rows (label + one value per financial year, 2015-16 to 2024-25)
No value is recalculated; this script only reshapes wide-by-year sheets into
one row per line-item x year, and flags aggregate/total rows for convenience.
"""
import csv
import re
from pathlib import Path

import openpyxl

RAW_DIR = Path(__file__).parent.parent / "raw"
OUT_DIR = Path(__file__).parent

SA_FILE = RAW_DIR / "55120do006-south-australia-general-government.xlsx"
TOTAL_STATE_FILE = RAW_DIR / "55120do011-total-state-general-government.xlsx"

MARKER_ROWS = {"less", "equals", "plus"}
TOTAL_KEYWORDS = [
    "net lending",
    "net worth",
    "net financial worth",
    "net operating balance",
    "cash surplus",
    "net change in the stock of cash",
]

SA_SHEETS = [
    ("Table_1", "operating_statement", "sa-operating-statement.csv",
     "South Australia State General Government Operating Statement"),
    ("Table_2", "cash_flow_statement", "sa-cash-flow-statement.csv",
     "South Australia State General Government Cash Flow Statement"),
    ("Table_3", "balance_sheet", "sa-balance-sheet.csv",
     "South Australia State General Government Balance Sheet"),
    ("Table_4", "expenses_by_purpose", "sa-expenses-by-purpose.csv",
     "South Australia State General Government Expenses by Purpose (COFOG)"),
]

TOTAL_STATE_SHEETS = [
    ("Table_1", "operating_statement", "all-states-total-operating-statement.csv",
     "Total State General Government Operating Statement (all states/territories combined)"),
]


def is_total_row(label):
    lower = label.lower()
    if lower.startswith("total"):
        return True
    return any(kw in lower for kw in TOTAL_KEYWORDS)


def parse_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    years = [str(y).strip() for y in rows[4][1:11]]
    records = []
    row_order = 0
    for row in rows[6:]:
        row = list(row)
        while row and row[-1] is None:
            row.pop()
        if not row:
            continue
        label = row[0]
        if label is None:
            continue
        label = str(label).strip()
        if not label or label in MARKER_ROWS or label.startswith("©"):
            continue
        values = row[1:]
        if not values:
            # header-only row: not carried into output, source's own section label
            continue
        row_order += 1
        total_flag = is_total_row(label)
        for year, value in zip(years, values):
            if value is None:
                continue
            records.append({
                "row_order": row_order,
                "line_item": label,
                "is_total_row": total_flag,
                "year": year,
                "value_million_aud": value,
            })
    return records


def write_csv(records, path):
    fieldnames = ["row_order", "line_item", "is_total_row", "year", "value_million_aud"]
    with open(path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in records:
            writer.writerow(r)


def main():
    sa_wb = openpyxl.load_workbook(SA_FILE, data_only=True)
    combined = []
    table_index = []

    for sheet_name, table_key, out_file, title in SA_SHEETS:
        records = parse_sheet(sa_wb[sheet_name])
        write_csv(records, OUT_DIR / out_file)
        for r in records:
            combined.append({"table": table_key, **r})
        table_index.append({
            "table": table_key,
            "title": title,
            "geography": "South Australia",
            "rows": len(records),
            "file": out_file,
        })
        print(f"{out_file}: {len(records)} rows")

    with open(OUT_DIR / "sa-all-tables-long.csv", "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["table", "row_order", "line_item",
                                                "is_total_row", "year", "value_million_aud"])
        writer.writeheader()
        for r in combined:
            writer.writerow(r)
    print(f"sa-all-tables-long.csv: {len(combined)} rows")

    total_wb = openpyxl.load_workbook(TOTAL_STATE_FILE, data_only=True)
    for sheet_name, table_key, out_file, title in TOTAL_STATE_SHEETS:
        records = parse_sheet(total_wb[sheet_name])
        write_csv(records, OUT_DIR / out_file)
        table_index.append({
            "table": table_key,
            "title": title,
            "geography": "Australia (all states/territories total)",
            "rows": len(records),
            "file": out_file,
        })
        print(f"{out_file}: {len(records)} rows")

    with open(OUT_DIR / "table-index.csv", "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["table", "title", "geography", "rows", "file"])
        writer.writeheader()
        for r in table_index:
            writer.writerow(r)

    # Sanity checks against raw values spot-checked by hand
    def lookup(records, line_item, year):
        for r in records:
            if r["line_item"].strip() == line_item and r["year"] == year:
                return r["value_million_aud"]
        return None

    op_records = parse_sheet(sa_wb["Table_1"])
    assert lookup(op_records, "Total GFS revenue", "2024-25") == 27208
    assert lookup(op_records, "GFS Net operating balance", "2024-25") == 65
    exp_records = parse_sheet(sa_wb["Table_4"])
    assert lookup(exp_records, "Total Expenses", "2024-25") == 27143
    assert lookup(exp_records, "Total health", "2024-25") == 9297
    bs_records = parse_sheet(sa_wb["Table_3"])
    assert lookup(bs_records, "GFS NET WORTH", "2024-25") == 63165
    print("Spot checks passed.")


if __name__ == "__main__":
    main()
